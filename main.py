import tkinter as tk
from tkinter import messagebox
from pathlib import Path
from datetime import datetime

import openpyxl

PROJECT_ROOT = Path(__file__).resolve().parent
EXCEL_PATH = PROJECT_ROOT / "resources" / "Formato_Dinamico_.xlsx"
OUTPUT_DIR = PROJECT_ROOT / "salida"
SHEET_NAME = "LSMW "


def export_sheet_to_tsv(
    excel_path: Path,
    sheet_name: str,
    output_dir: Path,
    file_prefix: str = "LSMW",
) -> tuple[Path, int]:
    """Lee `sheet_name` del workbook en `excel_path` y escribe un .txt
    separado por tabulación dentro de `output_dir`. Devuelve (ruta, filas)."""
    if not excel_path.exists():
        raise FileNotFoundError(f"No se encontró el archivo: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"La hoja '{sheet_name.strip()}' no existe en el archivo.")

    ws = wb[sheet_name]
    output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"{file_prefix}_{ts}.txt"

    rows_written = 0
    with output_path.open("w", encoding="utf-8", newline="") as f:
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v) for v in row]
            f.write("\t".join(cells) + "\n")
            rows_written += 1

    return output_path, rows_written


def extraer_lsmw_a_txt(status_var: tk.StringVar) -> None:
    try:
        output_path, rows_written = export_sheet_to_tsv(
            EXCEL_PATH, SHEET_NAME, OUTPUT_DIR
        )
    except FileNotFoundError as exc:
        messagebox.showerror("Archivo no encontrado", str(exc))
        return
    except ValueError as exc:
        messagebox.showerror("Hoja no encontrada", str(exc))
        return
    except Exception as exc:
        messagebox.showerror("Error al exportar", str(exc))
        return

    status_var.set(f"Exportado: {output_path.name} ({rows_written} filas)")
    messagebox.showinfo(
        "Extracción completa",
        f"Se generó el archivo:\n{output_path}\n\nFilas exportadas: {rows_written}",
    )


def main() -> None:
    root = tk.Tk()
    root.title("Extracción LSMW")
    root.geometry("440x180")
    root.resizable(False, False)

    title = tk.Label(
        root,
        text="Extracción de la hoja LSMW",
        font=("Helvetica", 13, "bold"),
    )
    title.pack(pady=(18, 4))

    subtitle = tk.Label(
        root,
        text=f"Origen: resources/{EXCEL_PATH.name}\nDestino: salida/",
        font=("Helvetica", 10),
        fg="#555",
        justify="center",
    )
    subtitle.pack(pady=(0, 10))

    status_var = tk.StringVar(value="")

    btn = tk.Button(
        root,
        text="Extraer información en txt",
        command=lambda: extraer_lsmw_a_txt(status_var),
        font=("Helvetica", 11),
        padx=18,
        pady=6,
    )
    btn.pack()

    status = tk.Label(root, textvariable=status_var, font=("Helvetica", 9), fg="#1a7f37")
    status.pack(pady=(10, 0))

    root.mainloop()


if __name__ == "__main__":
    main()
