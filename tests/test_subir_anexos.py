"""Pruebas unitarias para subir_anexos.py.

Mismo patrón de mocking que test_sap_upload / test_sox_report /
test_extraer_activos_creados: `MockSAPSession` registra cada llamada
`findById(...).method()` en `session.actions` para verificar la secuencia
exacta sin SAP real. Para tests del orquestador se usa `patch.multiple`
sobre los pasos individuales.
"""

import shutil
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import MagicMock, patch

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

import subir_anexos  # noqa: E402
from subir_anexos import (  # noqa: E402
    BTN_CONFIRMAR_WND1,
    CAMPO_ANLN1,
    CAMPO_ANLN2,
    CAMPO_BUKRS,
    CAMPO_DY_PATH,
    CREATE_ATTA,
    GOS_TOOLBOX,
    PCATTA_CREA,
    SHELL_GOS_BAR,
    SHELL_TITULAR,
    T_CODE_AS02,
    adjuntar_archivo,
    get_archivo_activos_mas_reciente,
    leer_activos_del_excel,
    subir_anexos as orquestador,
    validar_y_leer_activos_usuario,
)
from extraer_activos_creados import ACTIVOS_FIJOS_SHEET_NAME


# ---------------------------------------------------------------------------
# MockSAPSession (igual al usado en los otros test_*)
# ---------------------------------------------------------------------------


class MockSAPSession:
    def __init__(self):
        self._elements: dict = {}
        self.actions: list = []

    def findById(self, sap_id):
        if sap_id not in self._elements:
            self._elements[sap_id] = _MockElement(self, sap_id)
        return self._elements[sap_id]


class _MockElement:
    def __init__(self, session, sap_id):
        self._session = session
        self._sap_id = sap_id
        self.text = ""
        self.caretPosition = 0

    def press(self):
        self._session.actions.append((self._sap_id, "press"))

    def setFocus(self):
        self._session.actions.append((self._sap_id, "setFocus"))

    def maximize(self):
        self._session.actions.append((self._sap_id, "maximize"))

    def sendVKey(self, key):
        self._session.actions.append((self._sap_id, "sendVKey", key))

    def pressContextButton(self, ctx_id):
        self._session.actions.append(
            (self._sap_id, "pressContextButton", ctx_id)
        )

    def pressButton(self, btn_id):
        """SAP usa `pressButton` para clicks normales en botones de
        toolbar (distinto de `pressContextButton` que abre menú)."""
        self._session.actions.append(
            (self._sap_id, "pressButton", btn_id)
        )

    def selectContextMenuItem(self, item_id):
        self._session.actions.append(
            (self._sap_id, "selectContextMenuItem", item_id)
        )

    def __setattr__(self, name, value):
        if name in ("_session", "_sap_id"):
            super().__setattr__(name, value)
            return
        super().__setattr__(name, value)
        if name in ("text", "caretPosition"):
            self._session.actions.append(
                (self._sap_id, f"set_{name}", value)
            )


# ---------------------------------------------------------------------------
# get_archivo_activos_mas_reciente
# ---------------------------------------------------------------------------


class GetArchivoMasRecienteTest(unittest.TestCase):
    def setUp(self):
        self.tmpdir = Path(tempfile.mkdtemp(prefix="test_anexos_"))

    def tearDown(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_raises_when_dir_missing(self):
        with self.assertRaises(FileNotFoundError):
            get_archivo_activos_mas_reciente(self.tmpdir / "no_existe")

    def test_raises_when_no_matching_files(self):
        (self.tmpdir / "otro.xlsx").write_text("x")
        with self.assertRaisesRegex(FileNotFoundError, "ActivosCreados"):
            get_archivo_activos_mas_reciente(self.tmpdir)

    def test_returns_most_recent_by_mtime(self):
        import time as _time
        # Crear dos archivos con mtimes distintos
        a = self.tmpdir / "ActivosCreados_USR_20260101_120000.xlsx"
        a.write_text("a")
        _time.sleep(0.05)
        b = self.tmpdir / "ActivosCreados_USR_20260201_120000.xlsx"
        b.write_text("b")

        result = get_archivo_activos_mas_reciente(self.tmpdir)
        self.assertEqual(result, b)


# ---------------------------------------------------------------------------
# leer_activos_del_excel
# ---------------------------------------------------------------------------


class LeerActivosDelExcelTest(unittest.TestCase):
    def setUp(self):
        self.tmpdir = Path(tempfile.mkdtemp(prefix="test_leer_"))

    def tearDown(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _crear_archivo(self, pares=None, incluir_hoja=True) -> Path:
        wb = Workbook()
        ws = wb.active
        if incluir_hoja:
            ws.title = ACTIVOS_FIJOS_SHEET_NAME
            ws.append(("Activos Fijos", "Subnúmero"))
            for activo, sub in (pares or []):
                ws.append((activo, sub))
        else:
            ws.title = "OtraHoja"
        ruta = self.tmpdir / "test.xlsx"
        wb.save(ruta)
        return ruta

    def test_raises_when_file_missing(self):
        with self.assertRaises(FileNotFoundError):
            leer_activos_del_excel(self.tmpdir / "no_existe.xlsx")

    def test_raises_when_sheet_missing(self):
        archivo = self._crear_archivo(incluir_hoja=False)
        with self.assertRaisesRegex(ValueError, ACTIVOS_FIJOS_SHEET_NAME):
            leer_activos_del_excel(archivo)

    def test_returns_pairs_as_ints(self):
        archivo = self._crear_archivo(pares=[(100, 0), (200, 1), (300, 5)])
        result = leer_activos_del_excel(archivo)
        self.assertEqual(result, [(100, 0), (200, 1), (300, 5)])

    def test_skips_non_int_rows(self):
        """Filas con valores no-int (strings, None) se omiten."""
        archivo = self._crear_archivo(pares=[(100, 0)])
        # Añadir filas con datos inválidos
        from openpyxl import load_workbook
        wb = load_workbook(archivo)
        ws = wb[ACTIVOS_FIJOS_SHEET_NAME]
        ws.append(("texto", "no_int"))
        ws.append((None, None))
        ws.append((200, 1))
        wb.save(archivo)

        result = leer_activos_del_excel(archivo)
        self.assertEqual(result, [(100, 0), (200, 1)])

    def test_returns_empty_list_when_no_data_rows(self):
        archivo = self._crear_archivo(pares=[])
        self.assertEqual(leer_activos_del_excel(archivo), [])


# ---------------------------------------------------------------------------
# adjuntar_archivo
# ---------------------------------------------------------------------------


class AdjuntarArchivoTest(unittest.TestCase):
    """Verifica la secuencia exacta de acciones SAP que adjuntan un
    archivo a un activo: AS02 + ANLN1/ANLN2/BUKRS + GOS + cascada F4 +
    DY_PATH + cascada btn[0]."""

    def test_opens_as02_and_sets_asset_fields(self):
        session = MockSAPSession()
        adjuntar_archivo(session, 8048124, 0, "ISA", Path(r"C:\docs\a.pdf"))

        # AS02 en okcd
        self.assertEqual(
            session._elements["wnd[0]/tbar[0]/okcd"].text, T_CODE_AS02
        )
        # ANLN1 + BUKRS (ANLN2 omitido para subnúmero=0; ver
        # test_skips_anln2_when_zero / test_sets_anln2_when_nonzero).
        self.assertEqual(session._elements[CAMPO_ANLN1].text, "8048124")
        self.assertEqual(session._elements[CAMPO_BUKRS].text, "ISA")

    def test_skips_anln2_when_subnumero_is_zero(self):
        """Match exacto al recording: con sub=0 no se toca ANLN2 (su
        default ya es 0). Setearlo descoloca el focus chain en SAP."""
        session = MockSAPSession()
        adjuntar_archivo(session, 8048124, 0, "ISA", Path(r"C:\f.pdf"))

        # No debe haber acción set_text sobre ANLN2.
        anln2_set_actions = [
            a for a in session.actions
            if a[0] == CAMPO_ANLN2 and a[1] == "set_text"
        ]
        self.assertEqual(anln2_set_actions, [])

    def test_sets_anln2_when_subnumero_nonzero(self):
        """Si el subnúmero es != 0, sí debe setearse ANLN2."""
        session = MockSAPSession()
        adjuntar_archivo(session, 8048124, 3, "ISA", Path(r"C:\f.pdf"))

        self.assertEqual(session._elements[CAMPO_ANLN2].text, "3")

    def test_sets_caret_position_on_bukrs_after_focus(self):
        """Línea crítica del recording (caretPosition en BUKRS al final
        del texto) — sin este paso SAP no termina de aceptar el dato
        antes del Enter."""
        session = MockSAPSession()
        adjuntar_archivo(session, 8048124, 0, "ISA", Path(r"C:\f.pdf"))

        # setFocus precede a caretPosition, ambos sobre BUKRS.
        def idx(action):
            return session.actions.index(action)

        self.assertLess(
            idx((CAMPO_BUKRS, "setFocus")),
            idx((CAMPO_BUKRS, "set_caretPosition", 3)),
        )
        # caretPosition = len("ISA") = 3
        self.assertEqual(
            session._elements[CAMPO_BUKRS].caretPosition, 3
        )

    def test_opens_gos_with_pressButton_then_contextButton_then_select(self):
        """Match al recording actualizado (líneas 23-25 del .vbs):
          - pressButton "%GOS_TOOLBOX" en SHELL_TITULAR (NO pressContextButton)
          - pressContextButton "CREATE_ATTA" en SHELL_GOS_BAR (otra shell!)
          - selectContextMenuItem "PCATTA_CREA" (sin prefijo %GOS_)
        """
        session = MockSAPSession()
        adjuntar_archivo(session, 100, 0, "ISA", Path(r"C:\f.pdf"))

        self.assertIn(
            (SHELL_TITULAR, "pressButton", GOS_TOOLBOX),
            session.actions,
        )
        self.assertIn(
            (SHELL_GOS_BAR, "pressContextButton", CREATE_ATTA),
            session.actions,
        )
        self.assertIn(
            (SHELL_GOS_BAR, "selectContextMenuItem", PCATTA_CREA),
            session.actions,
        )

    def test_no_uses_f4_anywhere(self):
        """El recording NO tiene `sendVKey 4` en ninguna parte — todo
        se hace vía press de botones."""
        session = MockSAPSession()
        adjuntar_archivo(session, 100, 0, "ISA", Path(r"C:\f.pdf"))

        f4_actions = [
            a for a in session.actions
            if len(a) >= 3 and a[1] == "sendVKey" and a[2] == 4
        ]
        self.assertEqual(f4_actions, [])

    def test_sets_dy_path_in_wnd1(self):
        """DY_PATH vive en wnd[1], no wnd[2]. SAP abre wnd[1]
        directamente tras selectContextMenuItem("PCATTA_CREA")."""
        session = MockSAPSession()
        ruta = Path(r"C:\Users\xxx\docs\contrato.pdf")
        adjuntar_archivo(session, 100, 0, "ISA", ruta)

        self.assertTrue(CAMPO_DY_PATH.startswith("wnd[1]/"))
        self.assertEqual(
            session._elements[CAMPO_DY_PATH].text, str(ruta)
        )

    def test_does_not_touch_wnd2(self):
        """wnd[2] no existe en este flujo — todo ocurre en wnd[0]
        (asset) y wnd[1] (path dialog)."""
        session = MockSAPSession()
        adjuntar_archivo(session, 100, 0, "ISA", Path(r"C:\f.pdf"))

        wnd2_actions = [
            a for a in session.actions
            if isinstance(a[0], str) and a[0].startswith("wnd[2]")
        ]
        self.assertEqual(wnd2_actions, [])

    def test_does_not_set_dy_filename(self):
        """El recording sólo setea DY_PATH (path completo). No hay
        DY_FILENAME separado."""
        session = MockSAPSession()
        adjuntar_archivo(session, 100, 0, "ISA", Path(r"C:\f.pdf"))

        filename_actions = [
            a for a in session.actions
            if isinstance(a[0], str) and "DY_FILENAME" in a[0]
        ]
        self.assertEqual(filename_actions, [])

    def test_confirms_with_wnd1_btn0_at_end(self):
        """btn[0] de wnd[1] confirma el path inyectado y crea el adjunto.
        Es la ÚNICA confirmación — no hay cascada hacia más ventanas."""
        session = MockSAPSession()
        adjuntar_archivo(session, 100, 0, "ISA", Path(r"C:\f.pdf"))

        # btn[0] de wnd[1] se presiona DESPUÉS de setear DY_PATH
        def idx(action):
            return session.actions.index(action)

        self.assertLess(
            idx((CAMPO_DY_PATH, "set_text", r"C:\f.pdf")),
            idx((BTN_CONFIRMAR_WND1, "press")),
        )

    def test_raises_runtime_error_with_context_when_step_fails(self):
        """Si SAP rechaza algún paso (ej. AS02 no abre), se re-lanza
        RuntimeError con descripción del paso que falló."""
        session = MockSAPSession()
        original = session.findById

        def find_with_error(sap_id):
            if sap_id == CAMPO_ANLN1:
                raise Exception("campo no existe")
            return original(sap_id)

        session.findById = find_with_error

        with self.assertRaisesRegex(RuntimeError, "ANLN1"):
            adjuntar_archivo(session, 100, 0, "ISA", Path(r"C:\f.pdf"))


# ---------------------------------------------------------------------------
# subir_anexos (orquestador con soft-fail)
# ---------------------------------------------------------------------------


class SubirAnexosTest(unittest.TestCase):
    def setUp(self):
        self.tmpdir = Path(tempfile.mkdtemp(prefix="test_orq_"))
        # Archivo de activos con 2 pares
        self.archivo_activos = self.tmpdir / "ActivosCreados_USR_x.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = ACTIVOS_FIJOS_SHEET_NAME
        ws.append(("Activos Fijos", "Subnúmero"))
        ws.append((8048124, 0))
        ws.append((8048125, 0))
        wb.save(self.archivo_activos)

        # Archivos a subir (paths fake)
        self.archivo_a = self.tmpdir / "a.pdf"
        self.archivo_a.write_text("a")
        self.archivo_b = self.tmpdir / "b.pdf"
        self.archivo_b.write_text("b")

    def tearDown(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def test_raises_for_invalid_sociedad(self):
        with self.assertRaises(ValueError):
            orquestador(
                MagicMock(), "XYZ", [self.archivo_a],
                archivo_activos=self.archivo_activos,
            )

    def test_raises_when_no_files(self):
        with self.assertRaisesRegex(ValueError, "archivo"):
            orquestador(
                MagicMock(), "ISA", [],
                archivo_activos=self.archivo_activos,
            )

    def test_raises_when_activos_sheet_empty(self):
        archivo_vacio = self.tmpdir / "vacio.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = ACTIVOS_FIJOS_SHEET_NAME
        ws.append(("Activos Fijos", "Subnúmero"))
        wb.save(archivo_vacio)

        with self.assertRaisesRegex(ValueError, "sin filas"):
            orquestador(
                MagicMock(), "ISA", [self.archivo_a],
                archivo_activos=archivo_vacio,
            )

    def test_calls_adjuntar_for_each_activo_x_archivo(self):
        """Con 2 activos × 2 archivos = 4 llamadas a adjuntar_archivo."""
        with patch("subir_anexos.adjuntar_archivo") as mock_adj:
            stats = orquestador(
                MagicMock(), "ISA",
                [self.archivo_a, self.archivo_b],
                archivo_activos=self.archivo_activos,
            )

        self.assertEqual(mock_adj.call_count, 4)
        self.assertEqual(stats["total_intentos"], 4)
        self.assertEqual(stats["exitosos"], 4)
        self.assertEqual(stats["fallidos"], 0)

    def test_soft_fail_when_adjuntar_raises(self):
        """Si adjuntar_archivo falla para una combinación, se cuenta como
        fallo pero el orquestador sigue con la siguiente."""
        call_count = [0]

        def fake_adjuntar(*args, **kwargs):
            call_count[0] += 1
            if call_count[0] == 2:
                raise RuntimeError("simulated SAP error")

        with patch("subir_anexos.adjuntar_archivo", side_effect=fake_adjuntar):
            stats = orquestador(
                MagicMock(), "ISA",
                [self.archivo_a, self.archivo_b],
                archivo_activos=self.archivo_activos,
            )

        # 4 intentos, 3 OK, 1 fallido
        self.assertEqual(stats["total_intentos"], 4)
        self.assertEqual(stats["exitosos"], 3)
        self.assertEqual(stats["fallidos"], 1)
        self.assertEqual(len(stats["detalles_fallos"]), 1)
        # Cada fallo es (activo, sub, archivo_str, error_str)
        activo, sub, archivo, error = stats["detalles_fallos"][0]
        self.assertIn("simulated SAP error", error)

    def test_calls_progress_callback_for_each_attempt(self):
        progress_calls = []

        def cb(intento, total, desc):
            progress_calls.append((intento, total, desc))

        with patch("subir_anexos.adjuntar_archivo"):
            orquestador(
                MagicMock(), "ISA",
                [self.archivo_a, self.archivo_b],
                archivo_activos=self.archivo_activos,
                progress_callback=cb,
            )

        # 2 activos × 2 archivos = 4 callbacks
        self.assertEqual(len(progress_calls), 4)
        # El primer intento va con total = 4
        self.assertEqual(progress_calls[0][:2], (1, 4))
        self.assertEqual(progress_calls[-1][:2], (4, 4))

    def test_progress_callback_failure_does_not_break_flow(self):
        """Si el callback lanza, el orquestador captura silenciosamente
        para no romper el flujo SAP por un bug en la GUI."""
        def cb_que_falla(intento, total, desc):
            raise RuntimeError("callback bug")

        with patch("subir_anexos.adjuntar_archivo"):
            stats = orquestador(
                MagicMock(), "ISA", [self.archivo_a],
                archivo_activos=self.archivo_activos,
                progress_callback=cb_que_falla,
            )

        # Aunque el callback falla, los attachments sí se ejecutan.
        self.assertEqual(stats["exitosos"], 2)

    def test_activos_override_takes_priority_over_files(self):
        """Con `activos=` explícito NO se lee ningún archivo de salida/:
        se usan los pares provistos (ej. el .xlsx que el usuario subió)."""
        activos = [(9001, 0), (9002, 3), (9003, 0)]
        with patch(
            "subir_anexos.get_archivo_activos_mas_reciente",
            side_effect=AssertionError("no debe leer salida/"),
        ), patch(
            "subir_anexos.leer_activos_del_excel",
            side_effect=AssertionError("no debe leer archivo"),
        ), patch("subir_anexos.adjuntar_archivo") as mock_adj:
            stats = orquestador(
                MagicMock(), "ISA", [self.archivo_a],
                activos=activos,
            )

        # 3 activos × 1 archivo = 3 attachments.
        self.assertEqual(mock_adj.call_count, 3)
        self.assertEqual(stats["total_intentos"], 3)
        self.assertEqual(stats["exitosos"], 3)

    def test_raises_when_activos_override_empty(self):
        with self.assertRaisesRegex(ValueError, "vac"):
            orquestador(MagicMock(), "ISA", [self.archivo_a], activos=[])


# ---------------------------------------------------------------------------
# validar_y_leer_activos_usuario — .xlsx provisto por el usuario
# ---------------------------------------------------------------------------


class ValidarYLeerActivosUsuarioTest(unittest.TestCase):
    """Valida la lectura/validación estricta del .xlsx que el usuario sube
    con activos ya creados/existentes (Activo Fijo, Subnúmero)."""

    def setUp(self):
        self.tmpdir = Path(tempfile.mkdtemp(prefix="test_usr_"))

    def tearDown(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _crear_xlsx(self, filas, nombre="activos.xlsx", hojas_extra=0):
        """Crea un .xlsx con `filas` (lista de tuplas) en la hoja activa."""
        path = self.tmpdir / nombre
        wb = Workbook()
        ws = wb.active
        for fila in filas:
            ws.append(fila)
        for i in range(hojas_extra):
            wb.create_sheet(f"extra{i}")
        wb.save(path)
        return path

    def test_valid_file_returns_pairs(self):
        path = self._crear_xlsx([
            ("Activo Fijo", "Subnúmero"),
            (8048124, 0),
            (8048125, 1),
        ])
        self.assertEqual(
            validar_y_leer_activos_usuario(path),
            [(8048124, 0), (8048125, 1)],
        )

    def test_rejects_non_xlsx_extension(self):
        # La extensión se valida ANTES de abrir; el archivo ni existe.
        with self.assertRaisesRegex(ValueError, r"\.xlsx"):
            validar_y_leer_activos_usuario(self.tmpdir / "activos.xls")

    def test_rejects_multiple_sheets(self):
        path = self._crear_xlsx(
            [("Activo Fijo", "Subnúmero"), (1, 0)], hojas_extra=1,
        )
        with self.assertRaisesRegex(ValueError, "una sola hoja"):
            validar_y_leer_activos_usuario(path)

    def test_rejects_more_than_two_columns(self):
        path = self._crear_xlsx([
            ("Activo Fijo", "Subnúmero", "Extra"),
            (8048124, 0, "x"),
        ])
        with self.assertRaisesRegex(ValueError, "2 columnas"):
            validar_y_leer_activos_usuario(path)

    def test_rejects_missing_header(self):
        # Sin encabezado: la fila 1 ya son datos numéricos.
        path = self._crear_xlsx([(8048124, 0), (8048125, 1)])
        with self.assertRaisesRegex(ValueError, "encabezado"):
            validar_y_leer_activos_usuario(path)

    def test_rejects_non_numeric_data(self):
        path = self._crear_xlsx([
            ("Activo Fijo", "Subnúmero"),
            (8048124, 0),
            ("no-numero", 1),
        ])
        with self.assertRaisesRegex(ValueError, "Fila 3"):
            validar_y_leer_activos_usuario(path)

    def test_rejects_only_header_no_data(self):
        path = self._crear_xlsx([("Activo Fijo", "Subnúmero")])
        with self.assertRaisesRegex(ValueError, "debajo del encabezado"):
            validar_y_leer_activos_usuario(path)

    def test_accepts_numbers_as_text(self):
        path = self._crear_xlsx([
            ("Activo Fijo", "Subnúmero"),
            ("8048124", "0"),
        ])
        self.assertEqual(
            validar_y_leer_activos_usuario(path), [(8048124, 0)],
        )

    def test_accepts_integer_floats(self):
        path = self._crear_xlsx([
            ("Activo Fijo", "Subnúmero"),
            (8048124.0, 0.0),
        ])
        self.assertEqual(
            validar_y_leer_activos_usuario(path), [(8048124, 0)],
        )

    def test_ignores_trailing_empty_rows(self):
        path = self._crear_xlsx([
            ("Activo Fijo", "Subnúmero"),
            (8048124, 0),
            (None, None),
        ])
        self.assertEqual(
            validar_y_leer_activos_usuario(path), [(8048124, 0)],
        )


# ---------------------------------------------------------------------------
# main() CLI
# ---------------------------------------------------------------------------


class MainEntryPointTest(unittest.TestCase):
    def test_returns_2_when_wrong_arg_count(self):
        self.assertEqual(subir_anexos.main([]), 2)
        self.assertEqual(subir_anexos.main(["ISA"]), 2)

    def test_returns_1_when_file_missing(self):
        self.assertEqual(
            subir_anexos.main(["ISA", "/no/existe.pdf"]), 1
        )

    def test_returns_1_when_invalid_sociedad(self):
        with tempfile.NamedTemporaryFile(suffix=".pdf") as tf:
            self.assertEqual(subir_anexos.main(["XYZ", tf.name]), 1)


if __name__ == "__main__":
    unittest.main()
