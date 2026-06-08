"""Pruebas unitarias para extraer_activos_creados.py.

Replica el estilo de test_sox_report.py / test_sap_upload.py: usa un
`MockSAPSession` para verificar la secuencia exacta de llamadas
findById(...).method() sin necesidad de un SAP real.
"""

import shutil
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import MagicMock, patch

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

import extraer_activos_creados  # noqa: E402
from extraer_activos_creados import (  # noqa: E402
    ACTIVOS_FIJOS_HEADERS,
    ACTIVOS_FIJOS_SHEET_NAME,
    BTN_CONFIRMAR_WND1,
    BTN_EXPORTAR_TBAR0,
    BTN_EXPORTAR_TBAR1,
    CAMPO_CREATOR,
    CAMPO_DY_FILENAME,
    CAMPO_DY_PATH,
    CELDA_PRIMER_REGISTRO,
    HEADER_MENSAJE_LOG,
    LOGS_SHEET_NAME,
    NOMBRE_EXTENSION,
    NOMBRE_PREFIJO,
    PATRON_ACTIVO_LOG,
    T_CODE_SM35P,
    _nombre_archivo_extraccion,
    abrir_primer_registro,
    abrir_sm35p,
    exportar_log,
    extraer_activos_creados as orquestador,
    filtrar_por_usuario,
    get_sap_session,
    procesar_logs,
    validar_usuario_sap,
)


# ---------------------------------------------------------------------------
# Mock de sesión SAP (replica el de test_sap_upload.py / test_sox_report.py)
# ---------------------------------------------------------------------------


class MockSAPSession:
    def __init__(self):
        self._elements: dict = {}
        self.actions: list = []

    def findById(self, sap_id):
        if sap_id not in self._elements:
            self._elements[sap_id] = _MockElement(self, sap_id)
        return self._elements[sap_id]


class _MockElement:
    def __init__(self, session, sap_id):
        self._session = session
        self._sap_id = sap_id
        self.text = ""
        self.caretPosition = 0

    def press(self):
        self._session.actions.append((self._sap_id, "press"))

    def setFocus(self):
        self._session.actions.append((self._sap_id, "setFocus"))

    def maximize(self):
        self._session.actions.append((self._sap_id, "maximize"))

    def sendVKey(self, key):
        self._session.actions.append((self._sap_id, "sendVKey", key))

    def __setattr__(self, name, value):
        if name in ("_session", "_sap_id"):
            super().__setattr__(name, value)
            return
        super().__setattr__(name, value)
        if name in ("text", "caretPosition"):
            self._session.actions.append(
                (self._sap_id, f"set_{name}", value)
            )


# ---------------------------------------------------------------------------
# validar_usuario_sap
# ---------------------------------------------------------------------------


class ValidarUsuarioSapTest(unittest.TestCase):
    def test_accepts_numeric_id(self):
        self.assertEqual(validar_usuario_sap("1017209574"), "1017209574")

    def test_accepts_alphanumeric_id_preserving_casing(self):
        """IDs como INTC37089 deben pasar tal cual (no se fuerza casing)."""
        self.assertEqual(validar_usuario_sap("INTC37089"), "INTC37089")

    def test_strips_surrounding_whitespace(self):
        self.assertEqual(validar_usuario_sap("  USR123  "), "USR123")

    def test_rejects_empty_string(self):
        with self.assertRaises(ValueError):
            validar_usuario_sap("")

    def test_rejects_only_whitespace(self):
        with self.assertRaises(ValueError):
            validar_usuario_sap("   ")

    def test_rejects_non_string(self):
        for value in [None, 123, [], {}]:
            with self.subTest(value=value):
                with self.assertRaises(ValueError):
                    validar_usuario_sap(value)


# ---------------------------------------------------------------------------
# get_sap_session
# ---------------------------------------------------------------------------


class GetSapSessionTest(unittest.TestCase):
    def test_raises_when_pywin32_missing(self):
        with patch.dict("sys.modules", {"win32com": None, "win32com.client": None}):
            with self.assertRaisesRegex(RuntimeError, "pywin32"):
                get_sap_session()

    def test_returns_first_session_when_sap_ok(self):
        fake_session = MagicMock()
        fake_application = MagicMock()
        fake_application.Children.Count = 1
        fake_connection = MagicMock()
        fake_connection.Children.Count = 1
        fake_connection.Children.return_value = fake_session
        fake_application.Children.return_value = fake_connection
        fake_sap_gui = MagicMock()
        fake_sap_gui.GetScriptingEngine = fake_application

        fake_win32com = MagicMock()
        fake_win32com.client.GetObject.return_value = fake_sap_gui

        with patch.dict("sys.modules", {
            "win32com": fake_win32com,
            "win32com.client": fake_win32com.client,
        }):
            result = get_sap_session()

        self.assertIs(result, fake_session)


# ---------------------------------------------------------------------------
# abrir_sm35p
# ---------------------------------------------------------------------------


class AbrirSm35pTest(unittest.TestCase):
    def test_maximizes_and_sends_okcd(self):
        session = MockSAPSession()
        abrir_sm35p(session)

        self.assertIn(("wnd[0]", "maximize"), session.actions)
        self.assertEqual(
            session._elements["wnd[0]/tbar[0]/okcd"].text, T_CODE_SM35P
        )
        self.assertIn(("wnd[0]", "sendVKey", 0), session.actions)

    def test_order_is_maximize_then_okcd_then_enter(self):
        session = MockSAPSession()
        abrir_sm35p(session)

        # Extraer las acciones relevantes en orden.
        idx_max = next(
            i for i, a in enumerate(session.actions) if a == ("wnd[0]", "maximize")
        )
        idx_text = next(
            i for i, a in enumerate(session.actions)
            if a == ("wnd[0]/tbar[0]/okcd", "set_text", T_CODE_SM35P)
        )
        idx_enter = next(
            i for i, a in enumerate(session.actions)
            if a == ("wnd[0]", "sendVKey", 0)
        )
        self.assertLess(idx_max, idx_text)
        self.assertLess(idx_text, idx_enter)


# ---------------------------------------------------------------------------
# filtrar_por_usuario
# ---------------------------------------------------------------------------


class FiltrarPorUsuarioTest(unittest.TestCase):
    def test_sets_creator_with_wildcard_prefix(self):
        session = MockSAPSession()
        filtrar_por_usuario(session, "1017209574")

        self.assertEqual(
            session._elements[CAMPO_CREATOR].text, "*1017209574"
        )

    def test_focuses_creator_and_sends_enter(self):
        session = MockSAPSession()
        filtrar_por_usuario(session, "USR1")

        self.assertIn((CAMPO_CREATOR, "setFocus"), session.actions)
        self.assertIn(("wnd[0]", "sendVKey", 0), session.actions)

    def test_caret_position_matches_filter_length(self):
        """caretPosition se setea al len(*<usuario>) para dejar el cursor
        al final del texto, replicando el recording original."""
        session = MockSAPSession()
        filtrar_por_usuario(session, "1017209574")

        # filtro = "*1017209574" tiene 11 chars
        self.assertEqual(
            session._elements[CAMPO_CREATOR].caretPosition, 11
        )


# ---------------------------------------------------------------------------
# abrir_primer_registro
# ---------------------------------------------------------------------------


class AbrirPrimerRegistroTest(unittest.TestCase):
    def test_focuses_first_row_cell_and_sends_f2(self):
        session = MockSAPSession()
        abrir_primer_registro(session)

        self.assertIn((CELDA_PRIMER_REGISTRO, "setFocus"), session.actions)
        # F2 = sendVKey 2
        self.assertIn(("wnd[0]", "sendVKey", 2), session.actions)

    def test_sets_caret_position_in_cell(self):
        session = MockSAPSession()
        abrir_primer_registro(session)

        self.assertEqual(
            session._elements[CELDA_PRIMER_REGISTRO].caretPosition, 5
        )


# ---------------------------------------------------------------------------
# exportar_log
# ---------------------------------------------------------------------------


class ExportarLogTest(unittest.TestCase):
    def test_presses_export_toolbar_chain(self):
        session = MockSAPSession()
        exportar_log(session, r"C:\salida", "ActivosCreados_USR_x.xlsx")

        self.assertIn((BTN_EXPORTAR_TBAR0, "press"), session.actions)
        self.assertIn((BTN_EXPORTAR_TBAR1, "press"), session.actions)
        self.assertIn((BTN_CONFIRMAR_WND1, "press"), session.actions)

    def test_skips_f4_picker_and_uses_direct_dy_path(self):
        """A diferencia del recording, exportar_log NO usa F4 + picker
        (wnd[2]). Inyecta DY_PATH y DY_FILENAME directamente en wnd[1] y
        confirma con btn[11]. Esto da control total del path de salida."""
        session = MockSAPSession()
        exportar_log(session, r"C:\salida", "ActivosCreados_USR_x.xlsx")

        # No debe haber F4 en wnd[1] ni press en wnd[2]/btn[11]
        self.assertNotIn(("wnd[1]", "sendVKey", 4), session.actions)
        self.assertNotIn(("wnd[2]/tbar[0]/btn[11]", "press"), session.actions)

    def test_fills_dy_path_and_dy_filename(self):
        session = MockSAPSession()
        exportar_log(session, r"C:\salida", "ActivosCreados_USR_x.xlsx")

        self.assertEqual(
            session._elements[CAMPO_DY_PATH].text, r"C:\salida"
        )
        self.assertEqual(
            session._elements[CAMPO_DY_FILENAME].text, "ActivosCreados_USR_x.xlsx"
        )
        self.assertEqual(
            session._elements[CAMPO_DY_FILENAME].caretPosition,
            len("ActivosCreados_USR_x.xlsx"),
        )

    def test_orden_secuencial(self):
        """btn[86] → btn[43] → set DY_PATH/DY_FILENAME → btn[11]."""
        session = MockSAPSession()
        exportar_log(session, r"C:\out", "x.xlsx")

        def idx(action_tuple):
            return session.actions.index(action_tuple)

        self.assertLess(idx((BTN_EXPORTAR_TBAR0, "press")),
                        idx((BTN_EXPORTAR_TBAR1, "press")))
        self.assertLess(idx((BTN_EXPORTAR_TBAR1, "press")),
                        idx((CAMPO_DY_PATH, "set_text", r"C:\out")))
        self.assertLess(idx((CAMPO_DY_PATH, "set_text", r"C:\out")),
                        idx((BTN_CONFIRMAR_WND1, "press")))


# ---------------------------------------------------------------------------
# extraer_activos_creados (orquestador)
# ---------------------------------------------------------------------------


class NombreArchivoExtraccionTest(unittest.TestCase):
    def test_pattern_is_prefix_usuario_timestamp_ext(self):
        nombre = _nombre_archivo_extraccion("1017209574")
        self.assertTrue(nombre.startswith(f"{NOMBRE_PREFIJO}_1017209574_"))
        self.assertTrue(nombre.endswith(NOMBRE_EXTENSION))

    def test_timestamp_format_is_yyyymmdd_hhmmss(self):
        nombre = _nombre_archivo_extraccion("USR1")
        # Patrón: ActivosCreados_USR1_YYYYMMDD_HHMMSS.xlsx
        import re
        self.assertRegex(
            nombre,
            r"^ActivosCreados_USR1_\d{8}_\d{6}\.xlsx$",
        )


class ExtraerActivosCreadosTest(unittest.TestCase):
    def test_calls_all_5_steps_in_order(self):
        """4 etapas SAP + 1 post-procesamiento (procesar_logs)."""
        session = MockSAPSession()
        call_order = []

        def make_recorder(name):
            return lambda *args, **kwargs: call_order.append(name)

        with patch.multiple(
            "extraer_activos_creados",
            abrir_sm35p=make_recorder("abrir"),
            filtrar_por_usuario=make_recorder("filtrar"),
            abrir_primer_registro=make_recorder("primer_registro"),
            exportar_log=make_recorder("exportar"),
            procesar_logs=make_recorder("procesar"),
        ):
            orquestador(session, "1017209574")

        self.assertEqual(
            call_order,
            ["abrir", "filtrar", "primer_registro", "exportar", "procesar"],
        )

    def test_normalizes_usuario_before_passing_to_filtrar(self):
        """El usuario se pasa con strip() pero sin transformación de casing."""
        session = MockSAPSession()
        with patch("extraer_activos_creados.abrir_sm35p"), \
             patch("extraer_activos_creados.filtrar_por_usuario") as mock_filtrar, \
             patch("extraer_activos_creados.abrir_primer_registro"), \
             patch("extraer_activos_creados.exportar_log"), \
             patch("extraer_activos_creados.procesar_logs"):
            orquestador(session, "  INTC37089  ")

        mock_filtrar.assert_called_once_with(session, "INTC37089")

    def test_raises_for_invalid_usuario(self):
        session = MockSAPSession()
        with self.assertRaises(ValueError):
            orquestador(session, "")

    def test_returns_carpeta_and_nombre(self):
        """Devuelve la tupla (carpeta, nombre) usadas. Por default,
        carpeta = SALIDA_DIR y nombre = ActivosCreados_USR_TS.xlsx."""
        session = MockSAPSession()
        with patch("extraer_activos_creados.abrir_sm35p"), \
             patch("extraer_activos_creados.filtrar_por_usuario"), \
             patch("extraer_activos_creados.abrir_primer_registro"), \
             patch("extraer_activos_creados.exportar_log"), \
             patch("extraer_activos_creados.procesar_logs"):
            carpeta, nombre = orquestador(session, "1017209574")

        self.assertTrue(carpeta.endswith("salida"))
        self.assertRegex(
            nombre,
            r"^ActivosCreados_1017209574_\d{8}_\d{6}\.xlsx$",
        )

    def test_passes_carpeta_and_nombre_to_exportar_log(self):
        """exportar_log recibe la carpeta y el nombre del archivo final."""
        session = MockSAPSession()
        with patch("extraer_activos_creados.abrir_sm35p"), \
             patch("extraer_activos_creados.filtrar_por_usuario"), \
             patch("extraer_activos_creados.abrir_primer_registro"), \
             patch("extraer_activos_creados.exportar_log") as mock_exportar, \
             patch("extraer_activos_creados.procesar_logs"):
            orquestador(
                session, "USR1",
                carpeta_destino="/tmp/out", nombre_archivo="custom.xlsx",
            )

        mock_exportar.assert_called_once_with(
            session, "/tmp/out", "custom.xlsx"
        )

    def test_passes_archivo_path_to_procesar_logs(self):
        """procesar_logs recibe el Path absoluto al .xlsx generado por SAP."""
        session = MockSAPSession()
        with patch("extraer_activos_creados.abrir_sm35p"), \
             patch("extraer_activos_creados.filtrar_por_usuario"), \
             patch("extraer_activos_creados.abrir_primer_registro"), \
             patch("extraer_activos_creados.exportar_log"), \
             patch("extraer_activos_creados.procesar_logs") as mock_procesar:
            orquestador(
                session, "USR1",
                carpeta_destino="/tmp/out", nombre_archivo="custom.xlsx",
            )

        mock_procesar.assert_called_once_with(Path("/tmp/out/custom.xlsx"))


# ---------------------------------------------------------------------------
# PATRON_ACTIVO_LOG (regex de parseo de mensajes)
# ---------------------------------------------------------------------------


class PatronActivoLogTest(unittest.TestCase):
    """Verifica que el regex matche el formato real observado en SAP y sus
    variantes razonables."""

    def test_matches_canonical_format(self):
        """Formato observado: 'El act.fj. 8048124 0 se ha creado'."""
        m = PATRON_ACTIVO_LOG.search("El act.fj. 8048124 0 se ha creado")
        self.assertIsNotNone(m)
        self.assertEqual(m.group(1), "8048124")
        self.assertEqual(m.group(2), "0")

    def test_matches_with_space_between_act_and_fj(self):
        """Tolera variante con espacio: 'act. fj. 12345 3'."""
        m = PATRON_ACTIVO_LOG.search("Se creó el act. fj. 12345 3 con éxito")
        self.assertIsNotNone(m)
        self.assertEqual(m.group(1), "12345")
        self.assertEqual(m.group(2), "3")

    def test_matches_uppercase(self):
        """Case-insensitive: 'Act.fj.', 'ACT.FJ.', etc."""
        for variante in [
            "Act.fj. 99 0 ...",
            "ACT.FJ. 99 0 ...",
            "Act. Fj. 99 0 ...",
        ]:
            with self.subTest(variante=variante):
                m = PATRON_ACTIVO_LOG.search(variante)
                self.assertIsNotNone(m, variante)

    def test_does_not_match_without_subnumber(self):
        """Si falta el segundo número (subnúmero), el regex no matchea."""
        self.assertIsNone(
            PATRON_ACTIVO_LOG.search("act.fj. 12345 no_es_numero")
        )

    def test_does_not_match_without_act_fj(self):
        self.assertIsNone(PATRON_ACTIVO_LOG.search("activo 12345 0"))

    def test_finditer_multiple_matches_in_same_line(self):
        """Si una línea menciona dos activos, ambos se extraen."""
        texto = "act.fj. 100 0 / act.fj. 200 1"
        matches = list(PATRON_ACTIVO_LOG.finditer(texto))
        self.assertEqual(len(matches), 2)
        self.assertEqual(matches[0].group(1), "100")
        self.assertEqual(matches[1].group(1), "200")


# ---------------------------------------------------------------------------
# procesar_logs (post-procesamiento)
# ---------------------------------------------------------------------------


class ProcesarLogsTest(unittest.TestCase):
    """Renombra `Sheet1` → `Logs`, parsea mensajes, crea hoja `Activos
    Fijos` con pares (activo, subnúmero) deduplicados."""

    def setUp(self):
        self.tmpdir = Path(tempfile.mkdtemp(prefix="test_procesar_"))

    def tearDown(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _crear_archivo_sap(self, mensajes: list[str], sheet_title="Sheet1") -> Path:
        """Crea un .xlsx que simula el output de SAP (headers + mensajes
        en columna B)."""
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_title
        # Headers iguales al archivo real
        headers = [
            "Hora de log", HEADER_MENSAJE_LOG, "Cód.transacción",
            "Cont.índice", "Módulo", "Dynpros de JD", "Cont.índice",
            "Clase JD", "ID de JD", "Nº de JD",
        ]
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(1, col_idx, h)
        # Filas con mensajes en col B
        for row_idx, mensaje in enumerate(mensajes, start=2):
            ws.cell(row_idx, 2, mensaje)
        ruta = self.tmpdir / "test.xlsx"
        wb.save(ruta)
        return ruta

    def test_renames_active_sheet_to_logs(self):
        archivo = self._crear_archivo_sap(["foo"], sheet_title="Sheet1")
        procesar_logs(archivo)

        wb = load_workbook(archivo)
        self.assertIn(LOGS_SHEET_NAME, wb.sheetnames)
        self.assertNotIn("Sheet1", wb.sheetnames)

    def test_creates_activos_fijos_sheet(self):
        archivo = self._crear_archivo_sap(["El act.fj. 100 0 se ha creado"])
        procesar_logs(archivo)

        wb = load_workbook(archivo)
        self.assertIn(ACTIVOS_FIJOS_SHEET_NAME, wb.sheetnames)

    def test_extracts_single_activo_correctly(self):
        archivo = self._crear_archivo_sap(
            ["El act.fj. 8048124 0 se ha creado"]
        )
        procesar_logs(archivo)

        wb = load_workbook(archivo)
        ws = wb[ACTIVOS_FIJOS_SHEET_NAME]
        # Fila 1 = headers, fila 2 = primer dato
        self.assertEqual(ws.cell(2, 1).value, 8048124)
        self.assertEqual(ws.cell(2, 2).value, 0)

    def test_headers_are_bold_and_correct(self):
        archivo = self._crear_archivo_sap(["foo"])
        procesar_logs(archivo)

        wb = load_workbook(archivo)
        ws = wb[ACTIVOS_FIJOS_SHEET_NAME]
        for col_idx, expected in enumerate(ACTIVOS_FIJOS_HEADERS, start=1):
            self.assertEqual(ws.cell(1, col_idx).value, expected)
            self.assertTrue(ws.cell(1, col_idx).font.bold)

    def test_deduplicates_repeated_activos(self):
        """Si el mismo (activo, sub) aparece varias veces, en Activos
        Fijos sólo aparece una vez."""
        archivo = self._crear_archivo_sap([
            "El act.fj. 100 0 se ha creado",
            "act.fj. 100 0 modificación",
            "act.fj. 100 0 otro mensaje",
        ])
        stats = procesar_logs(archivo)

        self.assertEqual(stats["menciones_total"], 3)
        self.assertEqual(stats["activos_unicos"], 1)

        wb = load_workbook(archivo)
        ws = wb[ACTIVOS_FIJOS_SHEET_NAME]
        # Sólo una fila de datos (fila 2)
        self.assertEqual(ws.cell(2, 1).value, 100)
        self.assertIsNone(ws.cell(3, 1).value)

    def test_treats_same_activo_different_subnumeros_as_distinct(self):
        """(100, 0) y (100, 1) son activos distintos — ambos aparecen."""
        archivo = self._crear_archivo_sap([
            "El act.fj. 100 0 se ha creado",
            "El act.fj. 100 1 se ha creado",
        ])
        stats = procesar_logs(archivo)

        self.assertEqual(stats["activos_unicos"], 2)

        wb = load_workbook(archivo)
        ws = wb[ACTIVOS_FIJOS_SHEET_NAME]
        valores = [(ws.cell(r, 1).value, ws.cell(r, 2).value)
                   for r in range(2, 4)]
        self.assertEqual(valores, [(100, 0), (100, 1)])

    def test_preserves_order_of_first_appearance(self):
        archivo = self._crear_archivo_sap([
            "act.fj. 300 0 ...",
            "act.fj. 100 0 ...",
            "act.fj. 200 0 ...",
            "act.fj. 100 0 ... (duplicado)",
        ])
        procesar_logs(archivo)

        wb = load_workbook(archivo)
        ws = wb[ACTIVOS_FIJOS_SHEET_NAME]
        valores = [ws.cell(r, 1).value for r in range(2, 5)]
        self.assertEqual(valores, [300, 100, 200])

    def test_skips_rows_without_mensaje(self):
        """Filas con mensaje None, vacío o no-string no rompen el flujo."""
        archivo = self._crear_archivo_sap(["", "act.fj. 100 0 ok"])
        # Limpiar la primera fila de mensaje a None
        wb = load_workbook(archivo)
        wb.active.cell(2, 2, None)
        wb.save(archivo)

        stats = procesar_logs(archivo)
        self.assertEqual(stats["activos_unicos"], 1)

    def test_idempotent_replaces_existing_activos_fijos_sheet(self):
        """Si Activos Fijos ya existe (re-run), se borra y recrea."""
        archivo = self._crear_archivo_sap(["act.fj. 100 0 ..."])
        procesar_logs(archivo)
        # Modificar el archivo y re-procesar
        wb = load_workbook(archivo)
        ws = wb[LOGS_SHEET_NAME]
        ws.cell(3, 2, "act.fj. 999 5 nuevo")
        wb.save(archivo)

        stats = procesar_logs(archivo)

        wb = load_workbook(archivo)
        ws = wb[ACTIVOS_FIJOS_SHEET_NAME]
        # Debe haber 2 únicos (100, 0) y (999, 5), no acumulados duplicados
        valores = [(ws.cell(r, 1).value, ws.cell(r, 2).value)
                   for r in range(2, 5) if ws.cell(r, 1).value is not None]
        self.assertEqual(len(valores), 2)

    def test_raises_when_file_missing(self):
        with self.assertRaises(FileNotFoundError):
            procesar_logs(self.tmpdir / "no_existe.xlsx")

    def test_handles_no_matches_gracefully(self):
        """Sin mensajes que matcheen, la hoja Activos Fijos queda sólo
        con headers (sin filas de datos)."""
        archivo = self._crear_archivo_sap([
            "La transacción ha sido procesada con éxito.",
            "Estadística de proceso.",
        ])
        stats = procesar_logs(archivo)

        self.assertEqual(stats["menciones_total"], 0)
        self.assertEqual(stats["activos_unicos"], 0)

        wb = load_workbook(archivo)
        ws = wb[ACTIVOS_FIJOS_SHEET_NAME]
        # Solo headers, sin datos
        self.assertIsNone(ws.cell(2, 1).value)


# ---------------------------------------------------------------------------
# main() entry point
# ---------------------------------------------------------------------------


class MainEntryPointTest(unittest.TestCase):
    def test_returns_2_when_wrong_argument_count(self):
        self.assertEqual(extraer_activos_creados.main([]), 2)
        self.assertEqual(extraer_activos_creados.main(["a", "b"]), 2)

    def test_returns_1_on_validation_error(self):
        self.assertEqual(extraer_activos_creados.main(["   "]), 1)

    def test_returns_1_when_sap_unavailable(self):
        with patch("extraer_activos_creados.get_sap_session",
                   side_effect=RuntimeError("no SAP")):
            self.assertEqual(extraer_activos_creados.main(["USR1"]), 1)

    def test_returns_0_on_happy_path(self):
        with patch("extraer_activos_creados.get_sap_session",
                   return_value=MagicMock()), \
             patch("extraer_activos_creados.extraer_activos_creados",
                   return_value=("/tmp/salida", "x.xlsx")):
            self.assertEqual(extraer_activos_creados.main(["USR1"]), 0)


if __name__ == "__main__":
    unittest.main()
