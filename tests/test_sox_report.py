"""Pruebas unitarias para sox_report.py.

Las funciones que dialogan con SAP GUI Scripting se prueban con
`MockSAPSession` (copiada del estilo de test_sap_upload.py) para verificar
la secuencia exacta de llamadas findById/.method.
"""

import shutil
import sys
import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from unittest.mock import MagicMock, patch

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

import sox_report  # noqa: E402
from sox_report import (  # noqa: E402
    CAMPO_FECHA_DESDE,
    CAMPO_FECHA_HASTA,
    CAMPO_SOCIEDAD,
    CREADOS_FILTRO_VALOR,
    CREADOS_HEADERS,
    CREADOS_SHEET_NAME,
    DOCS_GRID_SHELL,
    PATRON_AF,
    SOX_NODE_KEY,
    STANDARD_FILE_PREFIX,
    STANDARD_SHEET_NAME,
    TREE_SHELL,
    VALID_SOCIEDADES,
    _clasificar_ppe_intg,
    abrir_transaccion_sox,
    exportar_a_excel,
    generar_hoja_creados,
    generar_reporte_sox,
    generar_xlsx_poblacion,
    get_sap_session,
    ingresar_parametros,
    validar_caracter_fecha,
    validar_fecha,
    validar_rango_fechas,
    validar_sociedad,
)


# ---------------------------------------------------------------------------
# Mock de sesión SAP (replica el de test_sap_upload.py)
# ---------------------------------------------------------------------------


class MockSAPSession:
    def __init__(self):
        self._elements: dict = {}
        self.actions: list = []

    def findById(self, sap_id):
        if sap_id not in self._elements:
            self._elements[sap_id] = _MockElement(self, sap_id)
        return self._elements[sap_id]


class _MockElement:
    def __init__(self, session, sap_id):
        self._session = session
        self._sap_id = sap_id
        self.text = ""
        self.caretPosition = 0

    def press(self):
        self._session.actions.append((self._sap_id, "press"))

    def maximize(self):
        self._session.actions.append((self._sap_id, "maximize"))

    def setFocus(self):
        self._session.actions.append((self._sap_id, "setFocus"))

    def close(self):
        self._session.actions.append((self._sap_id, "close"))

    def sendVKey(self, key):
        self._session.actions.append((self._sap_id, "sendVKey", key))

    def doubleClickNode(self, key):
        self._session.actions.append((self._sap_id, "doubleClickNode", key))

    def pressToolbarContextButton(self, key):
        self._session.actions.append(
            (self._sap_id, "pressToolbarContextButton", key)
        )

    def selectContextMenuItem(self, key):
        self._session.actions.append(
            (self._sap_id, "selectContextMenuItem", key)
        )


# ---------------------------------------------------------------------------
# Validaciones
# ---------------------------------------------------------------------------


class ValidarSociedadTest(unittest.TestCase):
    def test_accepts_all_valid_sociedades(self):
        for soc in VALID_SOCIEDADES:
            self.assertEqual(validar_sociedad(soc), soc)

    def test_normalizes_to_uppercase(self):
        self.assertEqual(validar_sociedad("isa"), "ISA")
        self.assertEqual(validar_sociedad("  tran  "), "TRAN")

    def test_rejects_invalid_value(self):
        with self.assertRaises(ValueError) as ctx:
            validar_sociedad("XYZ")
        self.assertIn("XYZ", str(ctx.exception))

    def test_rejects_empty_string(self):
        with self.assertRaises(ValueError):
            validar_sociedad("")

    def test_rejects_only_whitespace(self):
        with self.assertRaises(ValueError):
            validar_sociedad("   ")

    def test_rejects_non_string(self):
        with self.assertRaises(ValueError):
            validar_sociedad(None)  # type: ignore[arg-type]


class ValidarFechaTest(unittest.TestCase):
    def test_accepts_valid_date(self):
        result = validar_fecha("01.05.2026")
        self.assertEqual(result, datetime(2026, 5, 1))

    def test_accepts_with_surrounding_whitespace(self):
        self.assertEqual(validar_fecha("  31.12.2026 "), datetime(2026, 12, 31))

    def test_rejects_wrong_format(self):
        with self.assertRaises(ValueError):
            validar_fecha("2026-05-01")
        with self.assertRaises(ValueError):
            validar_fecha("01/05/2026")

    def test_rejects_invalid_day(self):
        with self.assertRaises(ValueError):
            validar_fecha("32.01.2026")

    def test_rejects_invalid_month(self):
        with self.assertRaises(ValueError):
            validar_fecha("01.13.2026")

    def test_rejects_empty(self):
        with self.assertRaises(ValueError):
            validar_fecha("")

    def test_rejects_alphabetic(self):
        with self.assertRaises(ValueError):
            validar_fecha("ab.cd.efgh")


class ValidarRangoFechasTest(unittest.TestCase):
    def test_accepts_desde_lower_than_hasta(self):
        d, h = validar_rango_fechas("01.05.2026", "31.05.2026")
        self.assertEqual(d, datetime(2026, 5, 1))
        self.assertEqual(h, datetime(2026, 5, 31))

    def test_accepts_equal_dates(self):
        d, h = validar_rango_fechas("15.05.2026", "15.05.2026")
        self.assertEqual(d, h)

    def test_rejects_hasta_before_desde(self):
        with self.assertRaisesRegex(ValueError, "mayor o igual"):
            validar_rango_fechas("31.05.2026", "01.05.2026")

    def test_propagates_format_error_from_desde(self):
        with self.assertRaises(ValueError):
            validar_rango_fechas("bad", "31.05.2026")

    def test_propagates_format_error_from_hasta(self):
        with self.assertRaises(ValueError):
            validar_rango_fechas("01.05.2026", "bad")


class ValidarCaracterFechaTest(unittest.TestCase):
    def test_accepts_digits_and_dots(self):
        self.assertTrue(validar_caracter_fecha("01.05.2026"))
        self.assertTrue(validar_caracter_fecha("123"))
        self.assertTrue(validar_caracter_fecha(""))
        self.assertTrue(validar_caracter_fecha("...."))

    def test_rejects_letters(self):
        self.assertFalse(validar_caracter_fecha("01a"))
        self.assertFalse(validar_caracter_fecha("hola"))

    def test_rejects_special_characters(self):
        self.assertFalse(validar_caracter_fecha("01-05"))
        self.assertFalse(validar_caracter_fecha("01/05"))
        self.assertFalse(validar_caracter_fecha("01 05"))
        self.assertFalse(validar_caracter_fecha("01,05"))

    def test_rejects_more_than_10_characters(self):
        self.assertFalse(validar_caracter_fecha("01.05.20266"))
        self.assertFalse(validar_caracter_fecha("12345678901"))


# ---------------------------------------------------------------------------
# get_sap_session
# ---------------------------------------------------------------------------


class GetSapSessionTest(unittest.TestCase):
    def test_raises_when_pywin32_missing(self):
        with patch.dict(sys.modules, {"win32com": None, "win32com.client": None}):
            with self.assertRaises(RuntimeError) as ctx:
                get_sap_session()
        self.assertIn("pywin32", str(ctx.exception))

    def test_returns_session_on_success(self):
        session = MagicMock(name="session")
        connection = MagicMock()
        connection.Children.Count = 1
        connection.Children.return_value = session
        engine = MagicMock()
        engine.Children.Count = 1
        engine.Children.return_value = connection
        sap_gui_auto = MagicMock()
        sap_gui_auto.GetScriptingEngine = engine
        fake_win32 = MagicMock()
        fake_win32.client.GetObject.return_value = sap_gui_auto

        with patch.dict(sys.modules, {
            "win32com": fake_win32,
            "win32com.client": fake_win32.client,
        }):
            result = get_sap_session()
        self.assertIs(result, session)


# ---------------------------------------------------------------------------
# Pasos del flujo SOX
# ---------------------------------------------------------------------------


class AbrirTransaccionSoxTest(unittest.TestCase):
    def setUp(self):
        # Asegurar que T_CODE_SOX está en None (modo árbol) para los tests
        # del comportamiento por defecto. Tests específicos del modo T-code
        # patchean este valor.
        self._t_code_original = sox_report.T_CODE_SOX
        sox_report.T_CODE_SOX = None

    def tearDown(self):
        sox_report.T_CODE_SOX = self._t_code_original

    def test_maximizes_and_double_clicks_node_when_no_tcode(self):
        session = MockSAPSession()
        abrir_transaccion_sox(session)

        self.assertIn(("wnd[0]", "maximize"), session.actions)
        self.assertIn(
            (TREE_SHELL, "doubleClickNode", SOX_NODE_KEY), session.actions
        )

    def test_uses_okcd_when_tcode_configured(self):
        """Cuando T_CODE_SOX está configurado, navega por la casilla de
        comandos en vez de tocar el árbol — más robusto."""
        sox_report.T_CODE_SOX = "ZSOX_REPORT"
        session = MockSAPSession()
        abrir_transaccion_sox(session)

        self.assertEqual(
            session._elements["wnd[0]/tbar[0]/okcd"].text, "ZSOX_REPORT"
        )
        self.assertIn(("wnd[0]", "sendVKey", 0), session.actions)
        # No se debe haber tocado el árbol
        self.assertFalse(
            any(a[0] == TREE_SHELL for a in session.actions),
            "No debería tocar el árbol cuando T_CODE_SOX está configurado",
        )

    def test_error_when_node_missing_includes_tree_diagnostic(self):
        """Cuando el doubleClickNode falla, el mensaje debe incluir
        pistas para resolver el problema y los nodos disponibles."""
        session = MockSAPSession()
        tree = session.findById(TREE_SHELL)
        tree.doubleClickNode = MagicMock(side_effect=Exception("node missing"))
        # Simular que el árbol expone GetAllNodeKeys/GetNodeTextByKey
        tree.GetAllNodeKeys = MagicMock(return_value=["F00001", "F00002"])
        tree.GetNodeTextByKey = MagicMock(side_effect=lambda k: f"Texto-{k}")

        with self.assertRaises(RuntimeError) as ctx:
            abrir_transaccion_sox(session)
        mensaje = str(ctx.exception)
        # Pista sobre T_CODE_SOX
        self.assertIn("T_CODE_SOX", mensaje)
        # Lista de nodos disponibles en el árbol
        self.assertIn("F00001", mensaje)
        self.assertIn("Texto-F00001", mensaje)
        self.assertIn("F00002", mensaje)

    def test_error_when_node_missing_handles_tree_enumeration_failure(self):
        """Si GetAllNodeKeys también falla, el diagnóstico no debe explotar."""
        session = MockSAPSession()
        tree = session.findById(TREE_SHELL)
        tree.doubleClickNode = MagicMock(side_effect=Exception("node missing"))
        tree.GetAllNodeKeys = MagicMock(side_effect=Exception("API no disponible"))

        with self.assertRaises(RuntimeError) as ctx:
            abrir_transaccion_sox(session)
        mensaje = str(ctx.exception)
        self.assertIn("T_CODE_SOX", mensaje)
        # El bloque de nodos debe estar presente aunque vacío/con error
        self.assertIn("no se pudo enumerar el árbol", mensaje)


class IngresarParametrosTest(unittest.TestCase):
    """Verifica el nuevo flujo del Script2sox.vbs: sociedad por texto, fechas
    vía calendario F4 (focusDate + selectionInterval en formato yyyymmdd)."""

    def test_sets_sociedad_as_text(self):
        session = MockSAPSession()
        ingresar_parametros(session, "ISA", "01.05.2026", "31.05.2026")

        self.assertEqual(session._elements[CAMPO_SOCIEDAD].text, "ISA")

    def test_opens_calendar_for_each_date_field(self):
        session = MockSAPSession()
        ingresar_parametros(session, "ISA", "01.05.2026", "31.05.2026")

        # Foco y caretPosition=0 en el campo Desde antes de F4
        self.assertIn((CAMPO_FECHA_DESDE, "setFocus"), session.actions)
        self.assertEqual(session._elements[CAMPO_FECHA_DESDE].caretPosition, 0)
        # Foco y caretPosition=0 en el campo Hasta antes de F4
        self.assertIn((CAMPO_FECHA_HASTA, "setFocus"), session.actions)
        self.assertEqual(session._elements[CAMPO_FECHA_HASTA].caretPosition, 0)
        # F4 (sendVKey 4) se envió dos veces: una para cada fecha
        f4_calls = [a for a in session.actions if a == ("wnd[0]", "sendVKey", 4)]
        self.assertEqual(len(f4_calls), 2)

    def test_sets_calendar_focus_and_selection_in_yyyymmdd_format(self):
        session = MockSAPSession()
        ingresar_parametros(session, "ISA", "01.05.2026", "31.05.2026")

        from sox_report import CALENDAR_SHELL
        calendario = session._elements[CALENDAR_SHELL]
        # El selectionInterval queda con la última fecha procesada (Hasta).
        # focusDate y selectionInterval se asignan dos veces (una por
        # cada fecha) — el valor final refleja la fecha hasta.
        self.assertEqual(calendario.focusDate, "20260531")
        self.assertEqual(calendario.selectionInterval, "20260531,20260531")

    def test_raises_when_date_format_invalid(self):
        session = MockSAPSession()
        with self.assertRaises(ValueError):
            ingresar_parametros(session, "ISA", "no-fecha", "31.05.2026")

    def test_presses_f8_to_execute(self):
        session = MockSAPSession()
        ingresar_parametros(session, "ISA", "01.05.2026", "31.05.2026")

        self.assertIn(("wnd[0]/tbar[1]/btn[8]", "press"), session.actions)


class ExportarAExcelTest(unittest.TestCase):
    """El método default es 'alv_grid' (&MB_EXPORT > &XXL sobre el grid),
    confirmado por `resources/Script2sox.vbs` para AR15. El modo 'pc_list'
    (%PC) sigue disponible para listas SAP clásicas, pero AR15 NO es una."""

    def setUp(self):
        self._method_original = sox_report.EXPORT_METHOD

    def tearDown(self):
        sox_report.EXPORT_METHOD = self._method_original

    def test_default_method_uses_pc_command(self):
        """En modo pc_list (default), escribe %PC en okcd y envía Enter."""
        sox_report.EXPORT_METHOD = "pc_list"
        session = MockSAPSession()
        exportar_a_excel(session, r"C:\salida", "SOX_ISA.xlsx")

        self.assertEqual(
            session._elements["wnd[0]/tbar[0]/okcd"].text, "%PC"
        )
        self.assertIn(("wnd[0]", "sendVKey", 0), session.actions)

    def test_pc_list_skips_format_step_when_save_dialog_already_open(self):
        """Si DY_PATH ya está disponible en wnd[1] (algunas versiones SAP
        abren save-as directamente), no debe enviar Enter intermedio."""
        sox_report.EXPORT_METHOD = "pc_list"
        session = MockSAPSession()
        exportar_a_excel(session, r"C:\salida", "SOX_ISA.xlsx")

        # MockSAPSession siempre encuentra elementos, así que la rama
        # save_dialog_listo=True se ejecuta. No debe haber sendVKey 0 sobre
        # wnd[1] (solo el sendVKey 0 sobre wnd[0] tras %PC).
        enters_a_wnd1 = [
            a for a in session.actions if a == ("wnd[1]", "sendVKey", 0)
        ]
        self.assertEqual(len(enters_a_wnd1), 0)

    def test_pc_list_sends_enter_when_format_dialog_intercepts(self):
        """Si DY_PATH no aparece en wnd[1] (hay un format dialog primero),
        debe enviar Enter (sendVKey 0) a wnd[1] como OK universal."""
        sox_report.EXPORT_METHOD = "pc_list"
        session = MockSAPSession()
        original_find = session.findById
        dy_path_already_failed = [False]

        def find_with_dy_path_initially_missing(sap_id):
            # La primera lookup de DY_PATH falla (sin save-as todavía).
            # Las siguientes lookups (después del Enter) son OK.
            if sap_id == "wnd[1]/usr/ctxtDY_PATH" and not dy_path_already_failed[0]:
                dy_path_already_failed[0] = True
                raise Exception("save-as todavía no aparece")
            return original_find(sap_id)

        session.findById = find_with_dy_path_initially_missing
        exportar_a_excel(session, r"C:\salida", "SOX_ISA.xlsx")

        # Verificar que se envió Enter a wnd[1]
        self.assertIn(("wnd[1]", "sendVKey", 0), session.actions)

    def test_pc_list_fills_save_dialog(self):
        sox_report.EXPORT_METHOD = "pc_list"
        session = MockSAPSession()
        exportar_a_excel(session, r"C:\salida", "SOX_ISA_x.xlsx")

        self.assertEqual(
            session._elements["wnd[1]/usr/ctxtDY_PATH"].text, r"C:\salida"
        )
        self.assertEqual(
            session._elements["wnd[1]/usr/ctxtDY_FILENAME"].text, "SOX_ISA_x.xlsx"
        )
        self.assertEqual(
            session._elements["wnd[1]/usr/ctxtDY_FILENAME"].caretPosition,
            len("SOX_ISA_x.xlsx"),
        )

    def test_alv_grid_invokes_export_xxl_menu(self):
        sox_report.EXPORT_METHOD = "alv_grid"
        session = MockSAPSession()
        exportar_a_excel(session, r"C:\salida", "SOX_ISA.xlsx")

        self.assertIn(
            (DOCS_GRID_SHELL, "pressToolbarContextButton", "&MB_EXPORT"),
            session.actions,
        )
        self.assertIn(
            (DOCS_GRID_SHELL, "selectContextMenuItem", "&XXL"),
            session.actions,
        )

    def test_alv_grid_fills_save_dialog(self):
        sox_report.EXPORT_METHOD = "alv_grid"
        session = MockSAPSession()
        exportar_a_excel(session, r"C:\salida", "x.xlsx")

        self.assertEqual(
            session._elements["wnd[1]/usr/ctxtDY_PATH"].text, r"C:\salida"
        )
        self.assertEqual(
            session._elements["wnd[1]/usr/ctxtDY_FILENAME"].text, "x.xlsx"
        )

    def test_alv_grid_presses_btn11_to_confirm_save(self):
        """El diálogo de save abierto por &XXL en AR15 confirma con btn[11]
        (no btn[0]). btn[0] no existe en ese diálogo — fue la causa del
        error 'The control could not be found by id' antes del fix."""
        sox_report.EXPORT_METHOD = "alv_grid"
        session = MockSAPSession()
        exportar_a_excel(session, r"C:\salida", "x.xlsx")

        self.assertIn(
            ("wnd[1]/tbar[0]/btn[11]", "press"), session.actions
        )
        self.assertNotIn(
            ("wnd[1]/tbar[0]/btn[0]", "press"), session.actions
        )

    def test_pc_list_presses_btn0_to_confirm_save(self):
        """El diálogo de save abierto por %PC sí usa btn[0]."""
        sox_report.EXPORT_METHOD = "pc_list"
        session = MockSAPSession()
        exportar_a_excel(session, r"C:\salida", "x.xlsx")

        self.assertIn(
            ("wnd[1]/tbar[0]/btn[0]", "press"), session.actions
        )

    def test_none_method_skips_export(self):
        sox_report.EXPORT_METHOD = None
        session = MockSAPSession()
        exportar_a_excel(session, r"C:\salida", "x.xlsx")

        # No debe haber tocado okcd, ni el grid, ni el save dialog
        self.assertFalse(session.actions)

    def test_raises_on_invalid_method(self):
        sox_report.EXPORT_METHOD = "metodo_inventado"
        session = MockSAPSession()
        with self.assertRaisesRegex(ValueError, "EXPORT_METHOD inválido"):
            exportar_a_excel(session, r"C:\salida", "x.xlsx")


# ---------------------------------------------------------------------------
# generar_reporte_sox (orquestador)
# ---------------------------------------------------------------------------


class StepErrorContextTest(unittest.TestCase):
    """Verifica que cuando una operación SAP falla durante los pasos del flujo
    SOX, la excepción re-lanzada contiene contexto suficiente para identificar
    la línea exacta que falló (clave porque las excepciones COM del SAP
    Frontend Server traen descripción vacía).
    """

    def setUp(self):
        # Algunos tests verifican el fallback al árbol — forzamos T_CODE_SOX
        # a None para que ese camino se ejecute. Los tests que verifican el
        # modo T-code patchean explícitamente.
        self._t_code_original = sox_report.T_CODE_SOX
        sox_report.T_CODE_SOX = None

    def tearDown(self):
        sox_report.T_CODE_SOX = self._t_code_original

    def test_abrir_transaccion_raises_with_context_when_maximize_fails(self):
        session = MockSAPSession()
        wnd = session.findById("wnd[0]")
        wnd.maximize = MagicMock(side_effect=Exception("COM error"))

        with self.assertRaisesRegex(RuntimeError, "Maximizar"):
            abrir_transaccion_sox(session)

    def test_abrir_transaccion_raises_with_context_when_tree_not_found(self):
        session = MockSAPSession()
        original = session.findById

        def find_with_error(sap_id):
            if sap_id == TREE_SHELL:
                raise Exception("tree not found")
            return original(sap_id)

        session.findById = find_with_error

        with self.assertRaises(RuntimeError) as ctx:
            abrir_transaccion_sox(session)
        # Mensaje incluye la ruta del árbol y pista para el usuario
        self.assertIn(TREE_SHELL, str(ctx.exception))
        self.assertIn("Easy Access", str(ctx.exception))

    def test_abrir_transaccion_raises_with_context_when_node_not_found(self):
        session = MockSAPSession()
        tree = session.findById(TREE_SHELL)
        tree.doubleClickNode = MagicMock(side_effect=Exception("node missing"))

        with self.assertRaises(RuntimeError) as ctx:
            abrir_transaccion_sox(session)
        msg = str(ctx.exception)
        self.assertIn(SOX_NODE_KEY, msg)
        # Incluye una pista para el usuario
        self.assertIn("menú", msg.lower())

    def test_ingresar_parametros_raises_when_sociedad_field_missing(self):
        session = MockSAPSession()
        original = session.findById

        def find_with_error(sap_id):
            if sap_id == CAMPO_SOCIEDAD:
                raise Exception("field missing")
            return original(sap_id)

        session.findById = find_with_error

        with self.assertRaisesRegex(RuntimeError, "Sociedad"):
            ingresar_parametros(session, "ISA", "01.05.2026", "31.05.2026")

    def test_ingresar_parametros_raises_when_f8_button_missing(self):
        session = MockSAPSession()
        original = session.findById

        def find_with_error(sap_id):
            if sap_id == "wnd[0]/tbar[1]/btn[8]":
                raise Exception("button missing")
            return original(sap_id)

        session.findById = find_with_error

        with self.assertRaisesRegex(RuntimeError, "Ejecutar"):
            ingresar_parametros(session, "ISA", "01.05.2026", "31.05.2026")

    def test_exportar_raises_when_grid_not_found_in_alv_mode(self):
        original_method = sox_report.EXPORT_METHOD
        sox_report.EXPORT_METHOD = "alv_grid"
        try:
            session = MockSAPSession()
            original = session.findById

            def find_with_error(sap_id):
                if sap_id == DOCS_GRID_SHELL:
                    raise Exception("grid missing")
                return original(sap_id)

            session.findById = find_with_error

            with self.assertRaisesRegex(RuntimeError, "grid de resultados"):
                exportar_a_excel(session, r"C:\salida", "x.xlsx")
        finally:
            sox_report.EXPORT_METHOD = original_method


class GenerarReporteSoxTest(unittest.TestCase):
    def test_calls_all_steps_in_order(self):
        session = MockSAPSession()
        call_order = []

        def make_recorder(name, return_value=None):
            def fn(*args, **kwargs):
                call_order.append(name)
                return return_value
            return fn

        fake_poblacion = Path("/tmp/Población_ISA_31.05.2026.xlsx")
        with patch.multiple(
            "sox_report",
            abrir_transaccion_sox=make_recorder("abrir"),
            ingresar_parametros=make_recorder("ingresar"),
            exportar_a_excel=make_recorder("exportar"),
            generar_xlsx_poblacion=make_recorder("poblacion", fake_poblacion),
            generar_hoja_creados=make_recorder("creados"),
        ):
            generar_reporte_sox(
                session, "ISA", "01.05.2026", "31.05.2026",
                carpeta_destino="/tmp", nombre_archivo="x.xlsx",
            )

        self.assertEqual(
            call_order,
            ["abrir", "ingresar", "exportar", "poblacion", "creados"],
        )

    def test_normalizes_sociedad_before_passing(self):
        session = MockSAPSession()
        fake_poblacion = Path("/tmp/Población_ISA_31.05.2026.xlsx")
        with patch("sox_report.ingresar_parametros") as mock_ing, \
             patch("sox_report.abrir_transaccion_sox"), \
             patch("sox_report.exportar_a_excel"), \
             patch(
                 "sox_report.generar_xlsx_poblacion",
                 return_value=fake_poblacion,
             ), \
             patch("sox_report.generar_hoja_creados"):
            generar_reporte_sox(
                session, "isa", "01.05.2026", "31.05.2026",
                carpeta_destino="/tmp", nombre_archivo="x.xlsx",
            )

        mock_ing.assert_called_once_with(
            session, "ISA", "01.05.2026", "31.05.2026"
        )

    def test_passes_normalized_sociedad_and_fecha_hasta_to_poblacion(self):
        """`generar_xlsx_poblacion` recibe la sociedad normalizada (uppercase)
        y la fecha hasta tal cual la ingresó el usuario (validada)."""
        session = MockSAPSession()
        fake_poblacion = Path("/tmp/Población_ISA_31.05.2026.xlsx")
        with patch("sox_report.abrir_transaccion_sox"), \
             patch("sox_report.ingresar_parametros"), \
             patch("sox_report.exportar_a_excel"), \
             patch(
                 "sox_report.generar_xlsx_poblacion",
                 return_value=fake_poblacion,
             ) as mock_pob, \
             patch("sox_report.generar_hoja_creados"):
            generar_reporte_sox(
                session, "isa", "01.05.2026", "31.05.2026",
                carpeta_destino="/tmp", nombre_archivo="SOX_x.xlsx",
            )

        # generar_xlsx_poblacion(archivo_sap, carpeta_destino, sociedad, fecha_hasta)
        args, _ = mock_pob.call_args
        archivo_sap, carpeta, sociedad, fecha = args
        self.assertEqual(archivo_sap, Path("/tmp/SOX_x.xlsx"))
        self.assertEqual(carpeta, Path("/tmp"))
        self.assertEqual(sociedad, "ISA")
        self.assertEqual(fecha, "31.05.2026")

    def test_passes_poblacion_path_to_generar_hoja_creados(self):
        """`generar_hoja_creados` recibe el Path al archivo Población que
        produjo el paso anterior."""
        session = MockSAPSession()
        fake_poblacion = Path("/anywhere/Población_ISA_31.05.2026.xlsx")
        with patch("sox_report.abrir_transaccion_sox"), \
             patch("sox_report.ingresar_parametros"), \
             patch("sox_report.exportar_a_excel"), \
             patch(
                 "sox_report.generar_xlsx_poblacion",
                 return_value=fake_poblacion,
             ), \
             patch("sox_report.generar_hoja_creados") as mock_creados:
            generar_reporte_sox(
                session, "ISA", "01.05.2026", "31.05.2026",
            )

        mock_creados.assert_called_once_with(fake_poblacion)

    def test_raises_for_invalid_sociedad(self):
        session = MockSAPSession()
        with self.assertRaises(ValueError):
            generar_reporte_sox(
                session, "XYZ", "01.05.2026", "31.05.2026",
            )

    def test_raises_for_invalid_date_range(self):
        session = MockSAPSession()
        with self.assertRaises(ValueError):
            generar_reporte_sox(
                session, "ISA", "31.05.2026", "01.05.2026",
            )

    def test_returns_poblacion_filename_with_sociedad_and_fecha_hasta(self):
        """El deliverable final es Población_*, no SOX_* — el handler GUI
        muestra ese nombre al usuario."""
        session = MockSAPSession()
        fake_poblacion = Path("/anywhere/salida/Población_ISA_31.05.2026.xlsx")
        with patch("sox_report.abrir_transaccion_sox"), \
             patch("sox_report.ingresar_parametros"), \
             patch("sox_report.exportar_a_excel"), \
             patch(
                 "sox_report.generar_xlsx_poblacion",
                 return_value=fake_poblacion,
             ), \
             patch("sox_report.generar_hoja_creados"):
            carpeta, nombre = generar_reporte_sox(
                session, "ISA", "01.05.2026", "31.05.2026",
            )

        self.assertEqual(nombre, "Población_ISA_31.05.2026.xlsx")
        self.assertEqual(carpeta, "/anywhere/salida")

    def test_export_method_none_skips_poblacion_and_returns_sap_intermediate(self):
        """Si EXPORT_METHOD=None no hay archivo SAP del cual leer; se omite
        tanto el Población_* como la hoja Creados, y se devuelve el nombre
        intermedio."""
        original_method = sox_report.EXPORT_METHOD
        sox_report.EXPORT_METHOD = None
        try:
            session = MockSAPSession()
            with patch("sox_report.abrir_transaccion_sox"), \
                 patch("sox_report.ingresar_parametros"), \
                 patch("sox_report.exportar_a_excel"), \
                 patch("sox_report.generar_xlsx_poblacion") as mock_pob, \
                 patch("sox_report.generar_hoja_creados") as mock_creados:
                carpeta, nombre = generar_reporte_sox(
                    session, "ISA", "01.05.2026", "31.05.2026",
                    carpeta_destino="/tmp", nombre_archivo="SOX_x.xlsx",
                )

            mock_pob.assert_not_called()
            mock_creados.assert_not_called()
            self.assertEqual(nombre, "SOX_x.xlsx")
            self.assertEqual(carpeta, "/tmp")
        finally:
            sox_report.EXPORT_METHOD = original_method


# ---------------------------------------------------------------------------
# generar_xlsx_poblacion (paso post-SAP)
# ---------------------------------------------------------------------------


class GenerarXlsxPoblacionTest(unittest.TestCase):
    """Lee el .xlsx que SAP exportó y produce `Población_{SOC}_{FECHA}.xlsx`
    con el contenido en la hoja `Original_SAP`."""

    def setUp(self):
        self.tmpdir = Path(tempfile.mkdtemp(prefix="test_poblacion_"))

    def tearDown(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _crear_archivo_sap(self, filename, rows):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        for row in rows:
            ws.append(row)
        ruta = self.tmpdir / filename
        wb.save(ruta)
        return ruta

    def test_creates_file_with_standard_name(self):
        sap = self._crear_archivo_sap("SOX_ISA_x.xlsx", [["a", "b"]])
        resultado = generar_xlsx_poblacion(sap, self.tmpdir, "ISA", "31.03.2026")

        self.assertEqual(resultado.name, "Población_ISA_31.03.2026.xlsx")
        self.assertTrue(resultado.exists())

    def test_filename_uses_constants(self):
        """El nombre se construye con las constantes — si se cambian, el
        nombre cambia consistentemente."""
        sap = self._crear_archivo_sap("SOX.xlsx", [["a"]])
        resultado = generar_xlsx_poblacion(sap, self.tmpdir, "ITCH", "15.06.2027")

        esperado = f"{STANDARD_FILE_PREFIX}_ITCH_15.06.2027.xlsx"
        self.assertEqual(resultado.name, esperado)

    def test_sheet_is_named_original_sap(self):
        sap = self._crear_archivo_sap("SOX.xlsx", [["x"]])
        resultado = generar_xlsx_poblacion(sap, self.tmpdir, "ISA", "31.03.2026")

        wb = load_workbook(resultado)
        self.assertEqual(wb.sheetnames, [STANDARD_SHEET_NAME])

    def test_content_matches_source(self):
        filas = [
            ["Fecha", "Usuario", "Valor"],
            ["2026-03-02", "INTC37089", "*** creado ***"],
            ["2026-03-03", "INTC37090", "x"],
        ]
        sap = self._crear_archivo_sap("SOX.xlsx", filas)
        resultado = generar_xlsx_poblacion(sap, self.tmpdir, "ISA", "31.03.2026")

        wb = load_workbook(resultado)
        ws = wb.active
        leidas = [list(row) for row in ws.iter_rows(values_only=True)]
        self.assertEqual(leidas, filas)

    def test_preserves_datetime_and_numeric_types(self):
        """El reporte SAP tiene columnas datetime/time/numeric — deben
        preservarse, no convertirse a string."""
        filas = [
            ["Fecha", "Hora", "Monto"],
            [datetime(2026, 3, 2, 0, 0), datetime(2026, 3, 2, 13, 0, 49).time(), 12345.67],
        ]
        sap = self._crear_archivo_sap("SOX.xlsx", filas)
        resultado = generar_xlsx_poblacion(sap, self.tmpdir, "ISA", "31.03.2026")

        wb = load_workbook(resultado)
        ws = wb.active
        row2 = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        self.assertEqual(row2[0], datetime(2026, 3, 2, 0, 0))
        self.assertEqual(row2[2], 12345.67)

    def test_preserves_number_format_per_cell(self):
        """SAP usa number_format específico para Fecha (`mm-dd-yy`) y Hora
        (`[$-F400]h:mm:ss\\ AM/PM`). Sin copiar number_format, openpyxl
        renderiza datetime/time con formato default (ISO + 24h) y el usuario
        ve `2026-03-02 0:00:00` y `13:00:49` en vez de `2/03/2026` y
        `1:00:49 p. m.` como en el original SAP."""
        # Crear source con number_format específicos por celda
        wb = Workbook()
        ws = wb.active
        ws.cell(1, 1, "Fecha")
        ws.cell(1, 2, "Hora")
        ws.cell(2, 1, datetime(2026, 3, 2)).number_format = "mm-dd-yy"
        ws.cell(2, 2, datetime(2026, 3, 2, 13, 0, 49).time()).number_format = (
            "[$-F400]h:mm:ss\\ AM/PM"
        )
        sap = self.tmpdir / "SOX_format.xlsx"
        wb.save(sap)

        resultado = generar_xlsx_poblacion(sap, self.tmpdir, "ISA", "31.03.2026")

        wb_new = load_workbook(resultado)
        ws_new = wb_new.active
        self.assertEqual(ws_new.cell(2, 1).number_format, "mm-dd-yy")
        self.assertEqual(
            ws_new.cell(2, 2).number_format, "[$-F400]h:mm:ss\\ AM/PM"
        )
        # Las celdas sin formato específico mantienen "General"
        self.assertEqual(ws_new.cell(1, 1).number_format, "General")

    def test_raises_when_source_file_missing(self):
        with self.assertRaises(FileNotFoundError) as ctx:
            generar_xlsx_poblacion(
                self.tmpdir / "no_existe.xlsx",
                self.tmpdir,
                "ISA",
                "31.03.2026",
            )
        self.assertIn("reporte SAP", str(ctx.exception))

    def test_raises_value_error_when_source_is_not_xlsx(self):
        """Algunas versiones de SAP exportan MHTML con extensión .xlsx —
        openpyxl falla, debemos dar un error accionable."""
        fake = self.tmpdir / "fake.xlsx"
        fake.write_text("este no es un xlsx real", encoding="utf-8")

        with self.assertRaises(ValueError) as ctx:
            generar_xlsx_poblacion(fake, self.tmpdir, "ISA", "31.03.2026")
        self.assertIn("MHTML", str(ctx.exception))

    def test_creates_destination_folder_if_missing(self):
        sap = self._crear_archivo_sap("SOX.xlsx", [["a"]])
        dest = self.tmpdir / "subdir" / "nested"
        resultado = generar_xlsx_poblacion(sap, dest, "ISA", "31.03.2026")

        self.assertTrue(resultado.exists())
        self.assertEqual(resultado.parent, dest)

    def test_fecha_with_whitespace_is_normalized(self):
        """fecha_hasta se re-formatea con validar_fecha→strftime, así que
        whitespace alrededor se normaliza en el nombre del archivo."""
        sap = self._crear_archivo_sap("SOX.xlsx", [["a"]])
        resultado = generar_xlsx_poblacion(
            sap, self.tmpdir, "ISA", "  31.03.2026  "
        )
        self.assertEqual(resultado.name, "Población_ISA_31.03.2026.xlsx")

    def test_raises_for_invalid_fecha_hasta(self):
        sap = self._crear_archivo_sap("SOX.xlsx", [["a"]])
        with self.assertRaises(ValueError):
            generar_xlsx_poblacion(sap, self.tmpdir, "ISA", "no-es-fecha")


# ---------------------------------------------------------------------------
# _clasificar_ppe_intg + PATRON_AF (helpers de generar_hoja_creados)
# ---------------------------------------------------------------------------


class ClasificarPpeIntgTest(unittest.TestCase):
    """Verifica el mapeo: '19'→Intangible, '20'/'14'→Activo Construcción,
    cualquier otro → PPE. Equivalente a la fórmula Excel IFS de la columna L."""

    def test_19_is_intangible(self):
        self.assertEqual(_clasificar_ppe_intg("19"), "Intangible")

    def test_20_is_activo_construccion(self):
        self.assertEqual(_clasificar_ppe_intg("20"), "Activo Construcción")

    def test_14_is_activo_construccion(self):
        self.assertEqual(_clasificar_ppe_intg("14"), "Activo Construcción")

    def test_other_prefixes_default_to_ppe(self):
        for prefijo in ["80", "12", "33", "01", "99", "", "abc"]:
            with self.subTest(prefijo=prefijo):
                self.assertEqual(_clasificar_ppe_intg(prefijo), "PPE")


class PatronAfRegexTest(unittest.TestCase):
    """El parseo de la columna D depende de este regex; un cambio aquí
    propaga al resto del flujo."""

    def test_parses_standard_format(self):
        m = PATRON_AF.match("AF 8047759-0 Buje 500 kV-RV")
        self.assertIsNotNone(m)
        self.assertEqual(m.group(1), "8047759")
        self.assertEqual(m.group(2), "0")
        self.assertEqual(m.group(3), "Buje 500 kV-RV")

    def test_parses_with_multiple_spaces_after_af(self):
        """`\\s+` acepta uno o más espacios entre AF y el código."""
        m = PATRON_AF.match("AF   2000421-0 Activo Test")
        self.assertIsNotNone(m)
        self.assertEqual(m.group(1), "2000421")

    def test_parses_denomination_with_special_chars(self):
        m = PATRON_AF.match("AF 1234-5 Banco de Baterías 125 Vcc #1")
        self.assertIsNotNone(m)
        self.assertEqual(m.group(3), "Banco de Baterías 125 Vcc #1")

    def test_parses_denomination_with_dashes(self):
        """La denominación puede contener guiones (no confundir con el `-`
        separador código-subnúmero)."""
        m = PATRON_AF.match("AF 8047759-0 Buje 500 kV-RV")
        self.assertEqual(m.group(3), "Buje 500 kV-RV")

    def test_does_not_match_non_af_prefix(self):
        self.assertIsNone(PATRON_AF.match("XX 8047759-0 Test"))
        self.assertIsNone(PATRON_AF.match("Activo 8047759-0 Test"))

    def test_does_not_match_when_codigo_is_not_numeric(self):
        self.assertIsNone(PATRON_AF.match("AF abc-0 Test"))


# ---------------------------------------------------------------------------
# generar_hoja_creados (etapa post-Población)
# ---------------------------------------------------------------------------


class GenerarHojaCreadosTest(unittest.TestCase):
    """Lee `Original_SAP` del workbook, filtra '*** creado ***', parsea
    col D y escribe la hoja `Creados` con observaciones + headers + datos."""

    def setUp(self):
        self.tmpdir = Path(tempfile.mkdtemp(prefix="test_creados_"))

    def tearDown(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)

    def _crear_poblacion(self, rows_original_sap, file_name="Población_x.xlsx"):
        """Crea un workbook con la hoja `Original_SAP` y las filas dadas.
        La primera fila es el header (8 columnas), las siguientes son datos.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = STANDARD_SHEET_NAME
        # Header
        for col_idx, header in enumerate([
            "Fecha", "Hora", "Usuario", "Identificación de objeto editada",
            "Valor de objeto ampliado", "Denominación de atributo",
            "Valor editado nuevo", "Valor editado antiguo",
        ], start=1):
            ws.cell(1, col_idx, header)
        # Datos
        for row_offset, row in enumerate(rows_original_sap, start=2):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row_offset, col_idx, value)
        # number_format de Fecha y Hora en la primera fila de datos (para
        # que generar_hoja_creados lo herede en la hoja Creados).
        if rows_original_sap:
            ws.cell(2, 1).number_format = "mm-dd-yy"
            ws.cell(2, 2).number_format = "[$-F400]h:mm:ss\\ AM/PM"

        ruta = self.tmpdir / file_name
        wb.save(ruta)
        return ruta

    def test_filter_keeps_only_creado_rows(self):
        rows = [
            [datetime(2026, 3, 2), datetime(2026, 3, 2, 13, 0).time(),
             "USR1", "AF 1000-0 Test1", "", "Atributo", CREADOS_FILTRO_VALOR, ""],
            [datetime(2026, 3, 2), datetime(2026, 3, 2, 14, 0).time(),
             "USR2", "AF 2000-0 Test2", "", "Atributo", "*** modificado ***", ""],
            [datetime(2026, 3, 2), datetime(2026, 3, 2, 15, 0).time(),
             "USR3", "AF 3000-0 Test3", "", "Atributo", CREADOS_FILTRO_VALOR, ""],
        ]
        poblacion = self._crear_poblacion(rows)
        stats = generar_hoja_creados(poblacion)

        self.assertEqual(stats["filas_filtradas"], 2)
        self.assertEqual(stats["filas_escritas"], 2)
        self.assertEqual(stats["filas_descartadas"], 0)

    def test_creates_creados_sheet_alongside_original_sap(self):
        rows = [
            [datetime(2026, 3, 2), datetime(2026, 3, 2, 13, 0).time(),
             "USR1", "AF 1000-0 Test", "", "", CREADOS_FILTRO_VALOR, ""],
        ]
        poblacion = self._crear_poblacion(rows)
        generar_hoja_creados(poblacion)

        wb = load_workbook(poblacion)
        self.assertIn(STANDARD_SHEET_NAME, wb.sheetnames)
        self.assertIn(CREADOS_SHEET_NAME, wb.sheetnames)

    def test_observations_block_in_rows_1_to_8(self):
        poblacion = self._crear_poblacion([])
        generar_hoja_creados(poblacion)

        wb = load_workbook(poblacion)
        ws = wb[CREADOS_SHEET_NAME]
        self.assertEqual(ws.cell(1, 1).value, "Observaciones")
        self.assertEqual(ws.cell(3, 1).value, "1.")
        self.assertIn("Activo fijo", ws.cell(3, 2).value)
        self.assertEqual(ws.cell(7, 1).value, "[a]")
        self.assertIn("-------------[a]-------------", ws.cell(8, 11).value)

    def test_headers_in_row_10_are_bold(self):
        poblacion = self._crear_poblacion([])
        generar_hoja_creados(poblacion)

        wb = load_workbook(poblacion)
        ws = wb[CREADOS_SHEET_NAME]
        for col_idx, header in enumerate(CREADOS_HEADERS, start=1):
            self.assertEqual(ws.cell(10, col_idx).value, header)
            self.assertTrue(
                ws.cell(10, col_idx).font.bold,
                f"Header de col {col_idx} ({header}) debería estar en bold",
            )

    def test_data_starts_at_row_11(self):
        rows = [
            [datetime(2026, 3, 2), datetime(2026, 3, 2, 13, 0).time(),
             "USR1", "AF 1234-0 Test", "valor_amp", "atributo",
             CREADOS_FILTRO_VALOR, "valor_ant"],
        ]
        poblacion = self._crear_poblacion(rows)
        generar_hoja_creados(poblacion)

        wb = load_workbook(poblacion)
        ws = wb[CREADOS_SHEET_NAME]
        # Fila 11 debe ser la primera (y única) fila de datos
        self.assertEqual(ws.cell(11, 3).value, "USR1")
        self.assertEqual(ws.cell(11, 4).value, 1234)  # Activo Fijo (int)
        self.assertEqual(ws.cell(11, 5).value, 0)     # Subnúmero (int)
        self.assertEqual(ws.cell(11, 6).value, "Test")  # Denominación
        # Fila 12 vacía
        self.assertIsNone(ws.cell(12, 3).value)

    def test_parses_codigo_subnumero_denominacion(self):
        rows = [
            [datetime(2026, 3, 2), datetime(2026, 3, 2, 13, 0).time(),
             "USR", "AF 8047759-0 Buje 500 kV-RV", "", "",
             CREADOS_FILTRO_VALOR, ""],
        ]
        poblacion = self._crear_poblacion(rows)
        generar_hoja_creados(poblacion)

        wb = load_workbook(poblacion)
        ws = wb[CREADOS_SHEET_NAME]
        self.assertEqual(ws.cell(11, 4).value, 8047759)  # código int
        self.assertEqual(ws.cell(11, 5).value, 0)        # subnúmero int
        self.assertEqual(ws.cell(11, 6).value, "Buje 500 kV-RV")

    def test_column_k_is_text_format_with_2_chars(self):
        """K debe ser TEXTO (number_format '@') de 2 caracteres, no número."""
        rows = [
            [datetime(2026, 3, 2), datetime(2026, 3, 2, 13, 0).time(),
             "USR", "AF 1900000-0 Test", "", "", CREADOS_FILTRO_VALOR, ""],
        ]
        poblacion = self._crear_poblacion(rows)
        generar_hoja_creados(poblacion)

        wb = load_workbook(poblacion)
        ws = wb[CREADOS_SHEET_NAME]
        k_cell = ws.cell(11, 11)
        self.assertEqual(k_cell.value, "19")
        self.assertIsInstance(k_cell.value, str)
        self.assertEqual(k_cell.number_format, "@")

    def test_column_l_classifies_19_as_intangible(self):
        rows = [
            [datetime(2026, 3, 2), datetime(2026, 3, 2, 13, 0).time(),
             "USR", "AF 1900000-0 Software", "", "",
             CREADOS_FILTRO_VALOR, ""],
        ]
        poblacion = self._crear_poblacion(rows)
        generar_hoja_creados(poblacion)

        wb = load_workbook(poblacion)
        ws = wb[CREADOS_SHEET_NAME]
        self.assertEqual(ws.cell(11, 11).value, "19")
        self.assertEqual(ws.cell(11, 12).value, "Intangible")

    def test_column_l_classifies_20_and_14_as_activo_construccion(self):
        rows = [
            [datetime(2026, 3, 2), datetime(2026, 3, 2, 13, 0).time(),
             "USR", "AF 2000421-0 Obra A", "", "", CREADOS_FILTRO_VALOR, ""],
            [datetime(2026, 3, 2), datetime(2026, 3, 2, 13, 0).time(),
             "USR", "AF 1400999-1 Obra B", "", "", CREADOS_FILTRO_VALOR, ""],
        ]
        poblacion = self._crear_poblacion(rows)
        generar_hoja_creados(poblacion)

        wb = load_workbook(poblacion)
        ws = wb[CREADOS_SHEET_NAME]
        self.assertEqual(ws.cell(11, 12).value, "Activo Construcción")
        self.assertEqual(ws.cell(12, 12).value, "Activo Construcción")

    def test_column_l_default_is_ppe(self):
        rows = [
            [datetime(2026, 3, 2), datetime(2026, 3, 2, 13, 0).time(),
             "USR", "AF 8047759-0 Equipo", "", "",
             CREADOS_FILTRO_VALOR, ""],
        ]
        poblacion = self._crear_poblacion(rows)
        generar_hoja_creados(poblacion)

        wb = load_workbook(poblacion)
        ws = wb[CREADOS_SHEET_NAME]
        self.assertEqual(ws.cell(11, 12).value, "PPE")

    def test_preserves_fecha_hora_number_format(self):
        rows = [
            [datetime(2026, 3, 2), datetime(2026, 3, 2, 13, 0).time(),
             "USR", "AF 1234-0 Test", "", "", CREADOS_FILTRO_VALOR, ""],
        ]
        poblacion = self._crear_poblacion(rows)
        generar_hoja_creados(poblacion)

        wb = load_workbook(poblacion)
        ws = wb[CREADOS_SHEET_NAME]
        self.assertEqual(ws.cell(11, 1).number_format, "mm-dd-yy")
        self.assertEqual(
            ws.cell(11, 2).number_format, "[$-F400]h:mm:ss\\ AM/PM"
        )

    def test_skips_rows_that_dont_match_regex_and_counts_them(self):
        """Filas que pasan el filtro pero col D no matchea el regex se
        omiten del output y se cuentan en `filas_descartadas`."""
        rows = [
            [datetime(2026, 3, 2), datetime(2026, 3, 2, 13, 0).time(),
             "USR1", "AF 1234-0 Válido", "", "", CREADOS_FILTRO_VALOR, ""],
            [datetime(2026, 3, 2), datetime(2026, 3, 2, 13, 0).time(),
             "USR2", "FORMATO INVALIDO", "", "", CREADOS_FILTRO_VALOR, ""],
            [datetime(2026, 3, 2), datetime(2026, 3, 2, 13, 0).time(),
             "USR3", "AF sin-código-0 X", "", "", CREADOS_FILTRO_VALOR, ""],
        ]
        poblacion = self._crear_poblacion(rows)
        stats = generar_hoja_creados(poblacion)

        self.assertEqual(stats["filas_filtradas"], 3)
        self.assertEqual(stats["filas_descartadas"], 2)
        self.assertEqual(stats["filas_escritas"], 1)

        wb = load_workbook(poblacion)
        ws = wb[CREADOS_SHEET_NAME]
        # Solo la fila válida llegó (fila 11)
        self.assertEqual(ws.cell(11, 3).value, "USR1")
        self.assertIsNone(ws.cell(12, 3).value)

    def test_skips_rows_where_col_d_is_not_text(self):
        """Si col D no es string (None, número, etc.) se descarta."""
        rows = [
            [datetime(2026, 3, 2), datetime(2026, 3, 2, 13, 0).time(),
             "USR1", None, "", "", CREADOS_FILTRO_VALOR, ""],
            [datetime(2026, 3, 2), datetime(2026, 3, 2, 13, 0).time(),
             "USR2", 123, "", "", CREADOS_FILTRO_VALOR, ""],
        ]
        poblacion = self._crear_poblacion(rows)
        stats = generar_hoja_creados(poblacion)

        self.assertEqual(stats["filas_filtradas"], 2)
        self.assertEqual(stats["filas_descartadas"], 2)
        self.assertEqual(stats["filas_escritas"], 0)

    def test_filter_is_exact_match_case_sensitive(self):
        """Variantes que no son exactamente `*** creado ***` no pasan."""
        rows = [
            [datetime(2026, 3, 2), datetime(2026, 3, 2, 13, 0).time(),
             "U", "AF 1-0 T", "", "", "*** Creado ***", ""],  # Cap C
            [datetime(2026, 3, 2), datetime(2026, 3, 2, 13, 0).time(),
             "U", "AF 1-0 T", "", "", " *** creado ***", ""],  # leading space
            [datetime(2026, 3, 2), datetime(2026, 3, 2, 13, 0).time(),
             "U", "AF 1-0 T", "", "", "creado", ""],
            [datetime(2026, 3, 2), datetime(2026, 3, 2, 13, 0).time(),
             "U", "AF 1-0 T", "", "", CREADOS_FILTRO_VALOR, ""],  # match
        ]
        poblacion = self._crear_poblacion(rows)
        stats = generar_hoja_creados(poblacion)
        self.assertEqual(stats["filas_filtradas"], 1)

    def test_replaces_existing_creados_sheet(self):
        """Si la hoja Creados ya existe (corrida previa), se borra y recrea."""
        rows = [
            [datetime(2026, 3, 2), datetime(2026, 3, 2, 13, 0).time(),
             "USR1", "AF 1000-0 Primero", "", "",
             CREADOS_FILTRO_VALOR, ""],
        ]
        poblacion = self._crear_poblacion(rows)
        # Primera corrida
        generar_hoja_creados(poblacion)
        # Modificar fuente y re-correr
        wb = load_workbook(poblacion)
        ws_src = wb[STANDARD_SHEET_NAME]
        ws_src.cell(2, 4, "AF 9999-0 Segundo")
        wb.save(poblacion)
        # Segunda corrida
        generar_hoja_creados(poblacion)

        wb = load_workbook(poblacion)
        ws = wb[CREADOS_SHEET_NAME]
        # Solo debe quedar el "Segundo" (no duplicarse con el "Primero")
        self.assertEqual(ws.cell(11, 6).value, "Segundo")
        self.assertIsNone(ws.cell(12, 6).value)

    def test_raises_value_error_when_original_sap_sheet_missing(self):
        # Workbook sin Original_SAP
        wb = Workbook()
        wb.active.title = "OtraHoja"
        ruta = self.tmpdir / "sin_original.xlsx"
        wb.save(ruta)

        with self.assertRaises(ValueError) as ctx:
            generar_hoja_creados(ruta)
        self.assertIn(STANDARD_SHEET_NAME, str(ctx.exception))

    def test_raises_file_not_found_when_file_missing(self):
        with self.assertRaises(FileNotFoundError):
            generar_hoja_creados(self.tmpdir / "no_existe.xlsx")

    def test_empty_source_produces_only_observations_and_headers(self):
        """Sin filas en Original_SAP, Creados igual tiene observaciones +
        headers; stats reporta 0 filas en todos los conteos."""
        poblacion = self._crear_poblacion([])
        stats = generar_hoja_creados(poblacion)

        self.assertEqual(stats["filas_filtradas"], 0)
        self.assertEqual(stats["filas_escritas"], 0)

        wb = load_workbook(poblacion)
        ws = wb[CREADOS_SHEET_NAME]
        self.assertEqual(ws.cell(1, 1).value, "Observaciones")
        self.assertEqual(ws.cell(10, 1).value, "Fecha")
        self.assertIsNone(ws.cell(11, 1).value)


# ---------------------------------------------------------------------------
# main() entry point
# ---------------------------------------------------------------------------


class MainEntryPointTest(unittest.TestCase):
    def test_returns_2_when_wrong_argument_count(self):
        self.assertEqual(sox_report.main(["ISA"]), 2)
        self.assertEqual(sox_report.main([]), 2)
        self.assertEqual(sox_report.main(["a", "b", "c", "d"]), 2)

    def test_returns_1_when_invalid_sociedad(self):
        self.assertEqual(
            sox_report.main(["XYZ", "01.05.2026", "31.05.2026"]), 1
        )

    def test_returns_1_when_invalid_date_range(self):
        self.assertEqual(
            sox_report.main(["ISA", "31.05.2026", "01.05.2026"]), 1
        )

    def test_returns_1_when_sap_session_fails(self):
        with patch(
            "sox_report.get_sap_session",
            side_effect=RuntimeError("SAP no abierto"),
        ):
            self.assertEqual(
                sox_report.main(["ISA", "01.05.2026", "31.05.2026"]), 1
            )

    def test_returns_0_on_happy_path(self):
        with patch("sox_report.get_sap_session", return_value=MagicMock()), \
             patch(
                 "sox_report.generar_reporte_sox",
                 return_value=("/tmp", "SOX_ISA.xlsx"),
             ) as mock_flow:
            self.assertEqual(
                sox_report.main(["ISA", "01.05.2026", "31.05.2026"]), 0
            )
            mock_flow.assert_called_once()


if __name__ == "__main__":
    unittest.main()
