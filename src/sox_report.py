"""
sox_report.py — Generación del Reporte SOX vía SAP GUI Scripting.

Replica los pasos grabados en `resources/Script2sox.vbs` (versión actual,
con T-code y calendario F4) — reemplaza la grabación inicial fragil con
nodos F00xxx del árbol que estaba en `resources/Scriptsox.vbs`.

Flujo:
1. Maximizar ventana y abrir la transacción SAP `AR15` vía okcd.
2. Llenar Sociedad (P_BUKRS) — texto directo.
3. Llenar Fecha Desde (S_DATUM-LOW) y Fecha Hasta (S_DATUM-HIGH) usando
   el calendario emergente (sendVKey 4 → focusDate + selectionInterval
   en formato yyyymmdd).
4. Ejecutar el reporte (F8).
5. (Opcional) Exportar a Excel vía menú contextual del grid
   (&MB_EXPORT → &XXL) y guardar en `salida/`.
   IMPORTANTE: el grid de AR15 usa otro shell ID que el del recording
   original. Si la exportación falla, hay que re-grabar ese paso y
   actualizar `DOCS_GRID_SHELL`.

REQUISITOS DE EJECUCIÓN
=======================
Sistema operativo: Windows con SAP GUI for Windows abierto y sesión
iniciada. Mismos requisitos que `sap_upload.py` (ver su docstring).

USO
===
    python src/sox_report.py SOCIEDAD DESDE HASTA

    Ejemplo:
        python src/sox_report.py ISA 01.05.2026 31.05.2026

También se puede invocar desde la GUI vía el botón "Control SOX" de
`main.py`.
"""

from __future__ import annotations

import re
import sys
import time
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from paths import SALIDA_DIR  # noqa: E402  (paths centralizado dev/bundled)

# Nombre estándar del archivo final entregable. Patrón:
#   {STANDARD_FILE_PREFIX}_{SOCIEDAD}_{FECHA_HASTA}.xlsx
# Ej: Población_ISA_31.03.2026.xlsx
STANDARD_FILE_PREFIX = "Población"
STANDARD_SHEET_NAME = "Original_SAP"

# Segunda hoja del Población — generada filtrando filas "*** creado ***" de
# Original_SAP y parseando la columna D (formato "AF <code>-<sub> <denom>").
CREADOS_SHEET_NAME = "Creados"
CREADOS_FILTRO_VALOR = "*** creado ***"

# Regex que parsea la columna D ("Identificación de objeto editada") de SAP.
# Formato: "AF" + uno o más espacios + código numérico + "-" + subnúmero
# numérico + uno o más espacios + denominación (texto libre con espacios).
# Ej: "AF 8047759-0 Buje 500 kV-RV" → ("8047759", "0", "Buje 500 kV-RV").
PATRON_AF = re.compile(r"^AF\s+(\d+)-(\d+)\s+(.+)$")

# Headers de la hoja Creados, en orden. Se escriben en la fila 10.
CREADOS_HEADERS = (
    "Fecha", "Hora", "Usuario", "Activo Fijo", "Subnúmero",
    "Identificación de objeto editada", "Valor de objeto ampliado",
    "Denominación de atributo", "Valor editado nuevo",
    "Valor editado antiguo", "Extrae", "PPE o Intangible",
)

# Tercera hoja del Población: evidencia visual del proceso (5 screenshots
# embedded). Se genera al final del flujo, después de Creados.
IPE_SHEET_NAME = "IPE"
# Filenames + descripciones de las 5 evidencias, en orden de aparición.
# Los filenames se usan tanto para guardar en el tempdir como para localizar
# y embebir en la hoja IPE. Si una captura falla (soft-fail), el filename
# no existirá y `generar_hoja_ipe` lo reportará como "no disponible".
IPE_SCREENSHOTS_INFO = (
    ("01_parametros_ingresados.png",
     "1. Pantalla de Modificaciones con sociedad y fechas ingresadas (antes de F8)."),
    ("02_primer_registro.png",
     "2. Primer registro de la tabla Modificaciones registros maestros AF."),
    ("03_ultimo_registro.png",
     "3. Último registro de la tabla (scroll al final)."),
    ("04_status_bar_bytes.png",
     "4. Status bar SAP con el conteo de bytes exportados."),
    ("05_propiedades_archivo.png",
     "5. Propiedades del archivo SAP descargado (bytes deben coincidir con #4)."),
)
# Ancho máximo (px) al que se escala cada screenshot embedded en IPE. Las
# capturas SAP suelen ser 1920+ px de ancho y aumentan mucho el tamaño del
# .xlsx; escalarlas mantiene el archivo manejable sin perder legibilidad.
IPE_IMAGE_MAX_WIDTH = 1200

# Títulos de ventanas de la app Tkinter que se deben minimizar antes de
# capturar pantalla, para que las screenshots IPE muestren SAP limpio sin
# la UI de "Creación Activos SAP" encima. El título debe coincidir con
# `root.title(...)` en `src/main.py`. Si en el futuro se renombra la
# ventana, actualizar aquí también.
TITULOS_VENTANA_APP = ("Creación Activos SAP",)

# Bloque de observaciones (filas 1-9) que va encima de los datos en la hoja
# Creados. Lista de (fila, columna, texto) — celdas no listadas quedan vacías.
CREADOS_OBSERVACIONES = (
    (1, 1, "Observaciones"),
    (3, 1, "1."),
    (3, 2, "En la Columna D se separa su codigo de Activo fijo, su subnúmero y "
           "su nombre con el fin de realizar la busqueda de los activos fijo "
           "por su codigo."),
    (4, 1, "2."),
    (4, 2, "Los activos fijos de PPE se pueden identificar con sus números "
           "iniciales, los cuales comienzan diferente al número 19. Con el "
           "número 19 comienzan los activos intangibles, identificados en las "
           "columnas [a]."),
    (5, 1, "3."),
    (5, 2, "También se identifican los activos en construcción en las columnas "
           "añadidas [a]"),
    (6, 1, "4."),
    (6, 2, "De la columna \"Valor editado nuevo\" se toman los activos con "
           "concepto \"Creado\""),
    (7, 1, "[a]"),
    (7, 2, "Se añade la columna K en la cual se extrae los dos primeros codigos "
           "de cada activo, para en la columna L la cual también se insertó "
           "para identificar que tipo de activo fijo es."),
    (8, 11, "-------------[a]-------------"),
)

# ---------------------------------------------------------------------------
# CONFIGURACIÓN
# ---------------------------------------------------------------------------

# Sociedades válidas (mismas opciones que el combo del formulario).
VALID_SOCIEDADES = (
    "TRAN", "ISA", "ITCH", "CEYA", "CABA", "RPAE", "CTMP", "REPD", "ISAP",
    "XM",
)

# Formato esperado en los campos de fecha del formulario (y de SAP).
DATE_FORMAT_USER = "%d.%m.%Y"

# IDs SAP capturados de resources/Scriptsox.vbs.
TREE_SHELL = (
    "wnd[0]/usr/cntlIMAGE_CONTAINER/shellcont/shell/shellcont[0]/shell"
)
SOX_NODE_KEY = "F00039"

# T-code SAP de la transacción del reporte SOX. Forma ROBUSTA de abrir la
# transacción (escribir el código en okcd y Enter) — no depende del árbol
# del menú, que tiene IDs (F00xxx) inestables entre usuarios y sesiones.
#
# Confirmado por `resources/Script2sox.vbs`: la T-code es AR15 (transacción
# estándar SAP de "Origen de altas de inmovilizado" / "Asset History").
# Si la T-code real de tu instalación es otra (variante Z*), ajustar aquí.
#
# Si se deja en None, el script hace fallback al árbol con SOX_NODE_KEY.
#
# Prefijo "/n": fuerza a SAP a CANCELAR la transacción actual y abrir AR15
# FRESCA desde cualquier pantalla. Es imprescindible para el multiselect: al
# terminar el reporte de una sociedad, SAP queda en la pantalla de resultados
# de AR15; sin "/n", escribir "AR15" crudo en la casilla de comandos desde
# ahí no reinicia la transacción y la siguiente sociedad falla. Mismo patrón
# que `subir_anexos.T_CODE_AS02 = "/nas02"`.
T_CODE_SOX: str | None = "/nAR15"

# Botón "Atrás" (F3) estándar de la barra de herramientas de SAP. Presionarlo
# devuelve una pantalla hacia atrás. Extraído de `resources/ScriptanexoREP.vbs`
# (líneas finales: dos `press` seguidos) — desde la pantalla de resultados de
# AR15, dos "Atrás" devuelven a SAP a la pantalla inicial, dejándolo listo
# para la siguiente sociedad del multiselect.
BTN_ATRAS_SAP = "wnd[0]/tbar[0]/btn[3]"

# Shell del calendario emergente que aparece al presionar F4 sobre un
# campo de fecha de SAP. Vía: setFocus + sendVKey(4) sobre el campo →
# foco/selección sobre este shell.
CALENDAR_SHELL = "wnd[1]/usr/cntlCONTAINER/shellcont/shell"

# Formato yyyymmdd que espera el calendario SAP para focusDate y
# selectionInterval (distinto del dd.mm.aaaa del formulario).
DATE_FORMAT_SAP_CALENDAR = "%Y%m%d"

CAMPO_SOCIEDAD = "wnd[0]/usr/ctxtP_BUKRS"
CAMPO_FECHA_DESDE = "wnd[0]/usr/ctxtS_DATUM-LOW"
CAMPO_FECHA_HASTA = "wnd[0]/usr/ctxtS_DATUM-HIGH"

# Shell del grid de resultados de AR15. Confirmado por
# `resources/Script2sox.vbs` (recording vigente). Si en otra instalación
# el ID es diferente, re-grabar y actualizar este valor.
DOCS_GRID_SHELL = (
    "wnd[0]/usr/subDISPLAY:SAPLBANK_OBJ_CHDOC:0210/"
    "cntlCC_CHANGE_DOCUMENTS_SURVAY/shellcont/shell/shellcont[1]/shell"
)

# Botón del diálogo "Save List in File" que abre &XXL en el ALV grid de
# AR15. El recording usa btn[11] (Generar/Reemplazar); el btn[0] del
# diálogo común de %PC no existe en este diálogo, por eso era el origen
# del error "The control could not be found by id" antes del fix.
ALV_SAVE_DIALOG_OK_BTN = "btn[11]"

# Método de exportación a archivo:
#   "alv_grid" → &MB_EXPORT > &XXL sobre el grid. Default — es lo que usa
#                AR15 (ALV grid, no lista clásica).
#   "pc_list"  → usa la T-code %PC (System > List > Save > File). Sólo
#                aplica a listas SAP clásicas (AR15 NO lo es).
#   None       → no exporta. Deja el reporte abierto en SAP y el usuario
#                guarda manualmente.
EXPORT_METHOD: str | None = "alv_grid"


# ---------------------------------------------------------------------------
# LOGGING
# ---------------------------------------------------------------------------

def _log(mensaje: str) -> None:
    ts = time.strftime("%H:%M:%S")
    print(f"[{ts}] {mensaje}", flush=True)


def _archivo_esta_bloqueado(archivo: Path) -> bool:
    """Detecta si un archivo está abierto en exclusiva por otro proceso
    (típicamente Excel). Intenta abrirlo con modo `r+b` (lectura+escritura
    binaria), que en Windows requiere acceso exclusivo que Excel bloquea.

    Si el archivo no existe, devuelve False (no está bloqueado, está libre
    para crear). Si existe y se puede abrir en r+b sin error → False.
    Si lanza PermissionError u OSError → True.
    """
    if not archivo.exists():
        return False
    try:
        with open(archivo, "r+b"):
            pass
        return False
    except (PermissionError, OSError):
        return True


def _guardar_workbook_seguro(wb, archivo: Path) -> None:
    """Wrapper de `wb.save(archivo)` que captura `PermissionError` y re-lanza
    con un mensaje accionable. Sin esto, el usuario veía el `[Errno 13]
    Permission denied` crudo de openpyxl y no era obvio que la causa era
    el archivo abierto en Excel.
    """
    try:
        wb.save(archivo)
    except PermissionError as exc:
        raise PermissionError(
            f"No se pudo guardar {archivo.name}: el archivo está abierto "
            f"en Excel (u otro proceso lo bloqueó). Por favor ciérralo y "
            f"vuelve a generar el reporte SOX.\n\n"
            f"Ruta: {archivo}"
        ) from exc


def _nombre_archivo_poblacion(sociedad_norm: str, fecha_hasta: str) -> str:
    """Construye el nombre estándar del archivo Población final.
    Compartido por `generar_xlsx_poblacion` (para nombrarlo al crearlo) y
    `generar_reporte_sox` (para chequeo temprano de bloqueo)."""
    fecha_norm = validar_fecha(fecha_hasta, etiqueta="fecha hasta").strftime(
        DATE_FORMAT_USER
    )
    return f"{STANDARD_FILE_PREFIX}_{sociedad_norm}_{fecha_norm}.xlsx"


def _ejecutar(descripcion: str, fn, *args, **kwargs):
    """Ejecuta `fn(*args, **kwargs)` loguenado la operación. Si falla,
    re-lanza con un mensaje descriptivo que dice exactamente qué intentaba
    hacer — esto es clave porque las excepciones COM de SAP (`SAP Frontend
    Server`) suelen venir con descripción vacía.
    """
    _log(f"  → {descripcion}")
    try:
        return fn(*args, **kwargs)
    except Exception as exc:
        raise RuntimeError(
            f"Falló: {descripcion}\n"
            f"Detalle técnico SAP: {exc!r}"
        ) from exc


# ---------------------------------------------------------------------------
# VALIDACIONES
# ---------------------------------------------------------------------------

def validar_sociedad(sociedad: str) -> str:
    """Verifica que la sociedad esté en VALID_SOCIEDADES.

    Devuelve la sociedad normalizada (uppercase + strip). Lanza ValueError
    si no es válida o está vacía.
    """
    if not isinstance(sociedad, str) or not sociedad.strip():
        raise ValueError("Debes seleccionar una sociedad.")
    norm = sociedad.strip().upper()
    if norm not in VALID_SOCIEDADES:
        raise ValueError(
            f"Sociedad inválida: '{sociedad}'. "
            f"Opciones válidas: {', '.join(VALID_SOCIEDADES)}."
        )
    return norm


def validar_fecha(fecha_str: str, etiqueta: str = "fecha") -> datetime:
    """Valida y parsea una fecha en formato dd.mm.aaaa.

    Args:
        fecha_str: cadena a parsear.
        etiqueta: nombre del campo (para mensajes de error).
    """
    if not isinstance(fecha_str, str) or not fecha_str.strip():
        raise ValueError(f"La {etiqueta} está vacía.")
    try:
        return datetime.strptime(fecha_str.strip(), DATE_FORMAT_USER)
    except ValueError as exc:
        raise ValueError(
            f"La {etiqueta} '{fecha_str}' no tiene el formato esperado dd.mm.aaaa."
        ) from exc


def validar_rango_fechas(desde: str, hasta: str) -> tuple[datetime, datetime]:
    """Valida ambas fechas y que `hasta >= desde`."""
    f_desde = validar_fecha(desde, etiqueta="fecha desde")
    f_hasta = validar_fecha(hasta, etiqueta="fecha hasta")
    if f_hasta < f_desde:
        raise ValueError(
            f"La fecha hasta ({hasta}) debe ser mayor o igual a la fecha desde ({desde})."
        )
    return f_desde, f_hasta


def validar_caracter_fecha(propuesto: str) -> bool:
    """Validación per-keystroke: solo dígitos y puntos, máx 10 caracteres.

    Se usa como `validatecommand` de los Entry de fecha para impedir que el
    usuario escriba letras u otros caracteres extraños.
    """
    if len(propuesto) > 10:
        return False
    return all(c.isdigit() or c == "." for c in propuesto)


# ---------------------------------------------------------------------------
# CONEXIÓN A SAP
# ---------------------------------------------------------------------------

def get_sap_session():
    """Conecta al SAP GUI Scripting Engine. Igual lógica que sap_upload."""
    try:
        import win32com.client  # type: ignore
    except ImportError as exc:
        raise RuntimeError(
            "Falta la dependencia pywin32. Instalar con: pip install pywin32"
        ) from exc

    try:
        sap_gui_auto = win32com.client.GetObject("SAPGUI")
    except Exception as exc:
        raise RuntimeError(
            "No se pudo conectar a SAP GUI. Verifica:\n"
            "  - SAP GUI for Windows está abierto y con sesión iniciada.\n"
            "  - SAP GUI Scripting habilitado en Options del cliente.\n"
            "  - sapgui/user_scripting = TRUE en el servidor SAP."
        ) from exc

    application = sap_gui_auto.GetScriptingEngine
    if application.Children.Count == 0:
        raise RuntimeError("No hay conexiones SAP activas en este SAP GUI.")
    connection = application.Children(0)
    if connection.Children.Count == 0:
        raise RuntimeError(
            "No hay sesiones activas en la conexión SAP. "
            "Inicia sesión en el sistema SAP antes de correr este script."
        )
    return connection.Children(0)


# ---------------------------------------------------------------------------
# PASOS DEL FLUJO SOX
# ---------------------------------------------------------------------------

def _intentar_listar_nodos_arbol(tree) -> str:
    """Intenta enumerar los nodos visibles del árbol SAP para diagnóstico.

    Útil cuando un `doubleClickNode(...)` falla porque el ID grabado ya no
    aplica. Devuelve una cadena con los pares (key → texto) listos para
    incluir en el mensaje de error.
    """
    try:
        keys = tree.GetAllNodeKeys()
    except Exception as exc:
        return f"  (no se pudo enumerar el árbol: {exc!r})"

    keys_list = list(keys) if keys else []
    if not keys_list:
        return "  (árbol vacío o sin nodos visibles)"

    lineas = []
    for key in keys_list[:50]:
        try:
            texto = tree.GetNodeTextByKey(key)
            lineas.append(f"  {key} → {texto}")
        except Exception:
            lineas.append(f"  {key} → (no se pudo leer el texto del nodo)")
    if len(keys_list) > 50:
        lineas.append(f"  ... ({len(keys_list) - 50} nodos más)")
    return "\n".join(lineas)


def abrir_transaccion_sox(session) -> None:
    """Maximiza la ventana y abre la transacción del reporte SOX.

    Si `T_CODE_SOX` está configurado, navega vía la casilla de comandos
    (okcd) — esto es ROBUSTO entre usuarios y sesiones. Si no, intenta el
    fallback de doble-clic en el árbol con `SOX_NODE_KEY` (como el
    recording original), pero los IDs F00xxx del árbol son inestables.
    """
    _log("Paso 1/4: Abriendo transacción SOX...")

    wnd = _ejecutar(
        "Localizar ventana principal wnd[0]",
        session.findById, "wnd[0]",
    )
    _ejecutar("Maximizar ventana principal", wnd.maximize)

    # Camino preferido: T-code en la casilla de comandos.
    if T_CODE_SOX:
        _log(f"  Modo T-code (recomendado): usando '{T_CODE_SOX}'")
        okcd = _ejecutar(
            "Localizar casilla de comandos (wnd[0]/tbar[0]/okcd)",
            session.findById, "wnd[0]/tbar[0]/okcd",
        )
        _ejecutar(
            f"Escribir T-code '{T_CODE_SOX}' en okcd",
            lambda: setattr(okcd, "text", T_CODE_SOX),
        )
        _ejecutar("Enviar Enter (sendVKey 0)", wnd.sendVKey, 0)
        return

    # Fallback: navegación del árbol como en el recording.
    _log(
        f"  Modo árbol (fallback): doble-clic en nodo {SOX_NODE_KEY!r} — "
        f"frágil entre usuarios. Considera configurar T_CODE_SOX."
    )

    try:
        tree = _ejecutar(
            f"Localizar árbol del menú SAP ({TREE_SHELL})",
            session.findById, TREE_SHELL,
        )
    except RuntimeError as exc:
        raise RuntimeError(
            f"{exc}\n\n"
            f"PISTA: el árbol del menú no se encuentra. Verifica:\n"
            f"  • Estar logueado en SAP y en la pantalla SAP Easy Access.\n"
            f"  • Que el menú de roles del usuario sea visible (no minimizado).\n"
            f"  • Que la ruta del árbol coincida con tu instalación."
        ) from exc

    try:
        _ejecutar(
            f"Doble clic en el nodo {SOX_NODE_KEY!r} del árbol",
            tree.doubleClickNode, SOX_NODE_KEY,
        )
    except RuntimeError as exc:
        # Diagnóstico extra: listar los nodos disponibles para que el usuario
        # identifique cuál es el correcto en SU árbol.
        nodos_disponibles = _intentar_listar_nodos_arbol(tree)
        raise RuntimeError(
            f"{exc}\n\n"
            f"DIAGNÓSTICO: los IDs del árbol SAP (F00xxx) son posiciones\n"
            f"secuenciales asignadas cuando se renderiza el menú del usuario\n"
            f"que grabó el script. Cambian entre usuarios y sesiones.\n\n"
            f"SOLUCIÓN ROBUSTA: configurar la constante T_CODE_SOX al inicio\n"
            f"de src/sox_report.py con la T-code real de la transacción\n"
            f"(ej. T_CODE_SOX = 'ZTRX_SOX').\n\n"
            f"Para descubrir la T-code:\n"
            f"  1. En SAP, abre la transacción manualmente (como lo hacías).\n"
            f"  2. Ve a 'Sistema → Estado' o mira la barra de título.\n"
            f"  3. El campo 'Transacción' muestra la T-code (ej. ZSOX_REPORT).\n\n"
            f"Nodos visibles en tu árbol actual:\n{nodos_disponibles}"
        ) from exc


def volver_a_pantalla_inicial(session, veces: int = 2) -> None:
    """Devuelve SAP a la pantalla inicial presionando "Atrás" (F3) `veces`.

    Replica el final de `resources/ScriptanexoREP.vbs` (dos `press` sobre
    `wnd[0]/tbar[0]/btn[3]`). Necesario entre sociedades del multiselect: tras
    exportar el reporte, SAP queda en la pantalla de resultados de AR15; dos
    "Atrás" lo devuelven a la pantalla inicial para poder abrir AR15 de nuevo.

    Cada press es **best-effort** (se loguea pero NO se relanza si falla): si
    SAP ya estaba una pantalla más arriba, o el botón no está disponible en
    ese estado, no queremos abortar el flujo por eso — la apertura de la
    siguiente sociedad usa igual `/nAR15`, que reinicia desde cualquier lado.

    IMPORTANTE: `veces` es fijo en 2 (como el recording), NO "hasta el tope".
    Presionar Atrás de más en la pantalla inicial de SAP dispara el popup
    "¿Desea salir del sistema?"; por eso solo se llama desde el estado de
    resultados, donde 2 Atrás = pantalla inicial.
    """
    _log(f"Volviendo a la pantalla inicial ({veces}× Atrás F3)...")
    for i in range(1, veces + 1):
        try:
            boton = session.findById(BTN_ATRAS_SAP)
            boton.press()
            _log(f"  Atrás {i}/{veces} OK")
        except Exception as exc:  # best-effort: no abortar el flujo
            _log(f"  Atrás {i}/{veces} no disponible (ignorado): {exc!r}")
            break


def _seleccionar_fecha_calendario(
    session, campo_id: str, fecha_str: str, etiqueta: str
) -> None:
    """Selecciona una fecha en SAP usando el calendario emergente F4.

    Replica `resources/Script2sox.vbs`: foco en el campo → sendVKey(4) abre
    el calendario en wnd[1] → setea focusDate y selectionInterval con la
    fecha en formato yyyymmdd.

    Args:
        session: sesión SAP GUI.
        campo_id: ID del campo de fecha (CAMPO_FECHA_DESDE / CAMPO_FECHA_HASTA).
        fecha_str: fecha en formato dd.mm.aaaa (se convierte internamente).
        etiqueta: nombre legible del campo para los logs (ej. "Desde").
    """
    fecha_sap = validar_fecha(fecha_str, etiqueta=etiqueta).strftime(
        DATE_FORMAT_SAP_CALENDAR
    )

    campo = _ejecutar(
        f"Localizar campo Fecha {etiqueta} ({campo_id})",
        session.findById, campo_id,
    )
    _ejecutar(f"Foco en campo Fecha {etiqueta}", campo.setFocus)
    _ejecutar(
        f"Posicionar cursor en Fecha {etiqueta} (caretPosition=0)",
        lambda: setattr(campo, "caretPosition", 0),
    )

    wnd = _ejecutar(
        "Localizar ventana principal wnd[0]",
        session.findById, "wnd[0]",
    )
    _ejecutar(
        f"Abrir calendario emergente F4 para Fecha {etiqueta}",
        wnd.sendVKey, 4,
    )

    calendario = _ejecutar(
        f"Localizar calendario emergente ({CALENDAR_SHELL})",
        session.findById, CALENDAR_SHELL,
    )
    _ejecutar(
        f"Enfocar fecha {fecha_sap} en calendario",
        lambda: setattr(calendario, "focusDate", fecha_sap),
    )
    _ejecutar(
        f"Seleccionar intervalo {fecha_sap},{fecha_sap}",
        lambda: setattr(
            calendario, "selectionInterval", f"{fecha_sap},{fecha_sap}"
        ),
    )


def ingresar_parametros(
    session, sociedad: str, fecha_desde: str, fecha_hasta: str
) -> None:
    """Llena P_BUKRS (texto directo) y Fecha Desde/Hasta vía calendario F4.
    NO ejecuta el reporte (F8) — esa parte está en `ejecutar_reporte` para
    permitir capturar un screenshot de los parámetros antes de la ejecución.
    """
    _log(
        f"Paso 2/4: Ingresando sociedad='{sociedad}', "
        f"desde='{fecha_desde}', hasta='{fecha_hasta}'..."
    )

    sociedad_field = _ejecutar(
        f"Localizar campo Sociedad ({CAMPO_SOCIEDAD})",
        session.findById, CAMPO_SOCIEDAD,
    )
    _ejecutar(
        f"Asignar Sociedad = '{sociedad}'",
        lambda: setattr(sociedad_field, "text", sociedad),
    )

    _seleccionar_fecha_calendario(
        session, CAMPO_FECHA_DESDE, fecha_desde, "Desde"
    )
    _seleccionar_fecha_calendario(
        session, CAMPO_FECHA_HASTA, fecha_hasta, "Hasta"
    )


def ejecutar_reporte(session) -> None:
    """Pulsa F8 (`wnd[0]/tbar[1]/btn[8]`) para ejecutar el reporte después
    de que `ingresar_parametros` haya llenado el formulario. Split aparte
    para permitir capturar una evidencia (screenshot) del estado del
    formulario antes de la ejecución."""
    _log("Paso 3/4: Ejecutando reporte (F8)...")
    boton_f8 = _ejecutar(
        "Localizar botón Ejecutar (F8 = wnd[0]/tbar[1]/btn[8])",
        session.findById, "wnd[0]/tbar[1]/btn[8]",
    )
    _ejecutar("Pulsar Ejecutar (F8)", boton_f8.press)


def _rellenar_save_dialog(
    session,
    carpeta_destino: str,
    nombre_archivo: str,
    boton_ok_id: str = "btn[0]",
) -> None:
    """Llena el diálogo de SAP "Save File" (DY_PATH + DY_FILENAME) y pulsa
    el botón de confirmación.

    El ID del botón varía según el diálogo:
      - %PC abre el diálogo "Save List in File" con OK en btn[0].
      - &XXL del ALV grid abre un diálogo similar pero con OK en btn[11]
        (Generar/Reemplazar); btn[0] no existe en ese diálogo.
    """
    path_field = _ejecutar(
        "Localizar campo de ruta (wnd[1]/usr/ctxtDY_PATH)",
        session.findById, "wnd[1]/usr/ctxtDY_PATH",
    )
    _ejecutar(
        f"Asignar ruta = '{carpeta_destino}'",
        lambda: setattr(path_field, "text", carpeta_destino),
    )

    nombre_field = _ejecutar(
        "Localizar campo de nombre (wnd[1]/usr/ctxtDY_FILENAME)",
        session.findById, "wnd[1]/usr/ctxtDY_FILENAME",
    )
    _ejecutar(
        f"Asignar nombre = '{nombre_archivo}'",
        lambda: setattr(nombre_field, "text", nombre_archivo),
    )
    _ejecutar(
        "Posicionar cursor al final del nombre",
        lambda: setattr(nombre_field, "caretPosition", len(nombre_archivo)),
    )

    boton_id_completo = f"wnd[1]/tbar[0]/{boton_ok_id}"
    boton_ok = _ejecutar(
        f"Localizar botón OK del diálogo ({boton_id_completo})",
        session.findById, boton_id_completo,
    )
    _ejecutar(f"Pulsar OK ({boton_ok_id}) para guardar", boton_ok.press)


def _exportar_via_pc_list(
    session, carpeta_destino: str, nombre_archivo: str
) -> None:
    """Exporta usando %PC (System > List > Save > File). Funciona para
    listas SAP clásicas como AR15.

    La estructura del diálogo que abre %PC varía entre versiones SAP:
      - Versión A: muestra primero un diálogo de selección de formato
        con OK en `tbar[0]/btn[0]`, después abre el save-as.
      - Versión B: muestra un popup distinto sin ese botón estándar.
      - Versión C: abre directamente el save-as (sin paso de formato).

    Estrategia robusta:
      1. Tras %PC, probar si ya hay un save-as en wnd[1] (buscar DY_PATH).
      2. Si no, enviar Enter (sendVKey 0) a wnd[1] — actúa como "OK"
         universal en cualquier diálogo modal sin depender del ID del
         botón. Después intentamos de nuevo el save-as.
    """
    _log("  Modo PC list: usando %PC en okcd...")
    wnd = _ejecutar(
        "Localizar ventana principal wnd[0]",
        session.findById, "wnd[0]",
    )
    okcd = _ejecutar(
        "Localizar casilla de comandos (wnd[0]/tbar[0]/okcd)",
        session.findById, "wnd[0]/tbar[0]/okcd",
    )
    _ejecutar(
        "Escribir '%PC' en okcd (System > List > Save > File)",
        lambda: setattr(okcd, "text", "%PC"),
    )
    _ejecutar("Enviar Enter (sendVKey 0)", wnd.sendVKey, 0)

    # ¿Save-as ya está abierto? Si DY_PATH existe en wnd[1], saltamos el
    # paso de "confirmar formato".
    save_dialog_listo = False
    try:
        session.findById("wnd[1]/usr/ctxtDY_PATH")
        save_dialog_listo = True
        _log("  → Save-as detectado directamente en wnd[1] (sin paso de formato)")
    except Exception:
        _log("  → wnd[1] no es el save-as todavía; enviando Enter para avanzar")

    if not save_dialog_listo:
        # Mandar Enter al diálogo de formato (cualquiera que sea su
        # estructura). Enter actúa como OK por default.
        wnd1 = _ejecutar(
            "Localizar diálogo intermedio (wnd[1])",
            session.findById, "wnd[1]",
        )
        _ejecutar("Confirmar formato con Enter (sendVKey 0)", wnd1.sendVKey, 0)

    # Diálogo de guardar archivo (debe estar abierto ahora)
    _rellenar_save_dialog(session, carpeta_destino, nombre_archivo)
    _log(f"Archivo guardado en: {carpeta_destino}\\{nombre_archivo}")


def _exportar_via_alv_grid(
    session, carpeta_destino: str, nombre_archivo: str
) -> None:
    """Exporta usando el menú contextual del ALV grid (&MB_EXPORT > &XXL).
    Replica `resources/Script2sox.vbs` para AR15 — usa DOCS_GRID_SHELL y
    cierra el diálogo de save con `ALV_SAVE_DIALOG_OK_BTN` (btn[11])."""
    _log("  Modo ALV grid: usando &MB_EXPORT + &XXL...")
    grid = _ejecutar(
        f"Localizar grid de resultados ({DOCS_GRID_SHELL})",
        session.findById, DOCS_GRID_SHELL,
    )
    _ejecutar(
        "Abrir menú de exportación (&MB_EXPORT)",
        grid.pressToolbarContextButton, "&MB_EXPORT",
    )
    _ejecutar(
        "Seleccionar exportación a Excel (&XXL)",
        grid.selectContextMenuItem, "&XXL",
    )
    _rellenar_save_dialog(
        session,
        carpeta_destino,
        nombre_archivo,
        boton_ok_id=ALV_SAVE_DIALOG_OK_BTN,
    )
    _log(f"Archivo guardado en: {carpeta_destino}\\{nombre_archivo}")


def exportar_a_excel(
    session, carpeta_destino: str, nombre_archivo: str
) -> None:
    """Exporta el reporte al archivo dado usando el método configurado en
    `EXPORT_METHOD`.

    - "pc_list" (default): usa %PC, recomendado para AR15.
    - "alv_grid": usa &MB_EXPORT > &XXL, requiere DOCS_GRID_SHELL válido.
    - None: no exporta; el reporte queda visible en SAP para guardar
      manualmente.
    """
    _log("Paso 4/4: Exportando reporte...")

    if EXPORT_METHOD is None:
        _log(
            "  EXPORT_METHOD=None → exportación omitida. "
            "Guarda el reporte manualmente desde SAP."
        )
        return

    if EXPORT_METHOD == "pc_list":
        _exportar_via_pc_list(session, carpeta_destino, nombre_archivo)
        return

    if EXPORT_METHOD == "alv_grid":
        _exportar_via_alv_grid(session, carpeta_destino, nombre_archivo)
        return

    raise ValueError(
        f"EXPORT_METHOD inválido: '{EXPORT_METHOD}'. "
        f"Usa 'pc_list', 'alv_grid' o None."
    )


def generar_xlsx_poblacion(
    archivo_sap: Path,
    carpeta_destino: Path,
    sociedad: str,
    fecha_hasta: str,
) -> Path:
    """Produce el .xlsx final con nombre estándar a partir del reporte SAP.

    Lee el contenido del .xlsx que SAP exportó (típicamente `SOX_*.xlsx`) y
    lo copia a un nuevo archivo con nombre estándar:

        Población_{SOCIEDAD}_{FECHA_HASTA}.xlsx

    Ej: `Población_ISA_31.03.2026.xlsx`.

    El nuevo archivo tiene una única hoja llamada `Original_SAP` con el
    contenido del reporte original (sólo valores, sin fórmulas).

    Args:
        archivo_sap: ruta al .xlsx generado por SAP.
        carpeta_destino: carpeta donde guardar el nuevo archivo.
        sociedad: código normalizado (uppercase) de sociedad.
        fecha_hasta: fecha hasta en formato dd.mm.aaaa (ya validada).

    Returns:
        Path al archivo Población_* creado.

    Raises:
        FileNotFoundError: si `archivo_sap` no existe.
        ValueError: si openpyxl no puede leer el archivo (algunas versiones
            de SAP exportan MHTML con extensión .xlsx — habría que
            re-grabar la exportación con formato XLSX real).
    """
    if not archivo_sap.exists():
        raise FileNotFoundError(
            f"No se encontró el reporte SAP en {archivo_sap}.\n"
            f"Verifica que la exportación SAP haya guardado el archivo "
            f"correctamente antes de generar el .xlsx estándar."
        )

    _log(f"Leyendo reporte SAP: {archivo_sap.name}")
    try:
        wb_sap = load_workbook(archivo_sap, data_only=True)
    except Exception as exc:
        raise ValueError(
            f"No se pudo abrir {archivo_sap.name} como .xlsx. "
            f"Algunas versiones de SAP exportan MHTML con extensión .xlsx; "
            f"si es ese tu caso, re-grabar la exportación con formato XLSX real. "
            f"Detalle técnico: {exc!r}"
        ) from exc

    ws_sap = wb_sap.active

    wb_new = Workbook()
    ws_new = wb_new.active
    ws_new.title = STANDARD_SHEET_NAME

    # Copiar celda por celda preservando `number_format` — sin esto, las
    # columnas Fecha y Hora se ven con el formato default de openpyxl
    # (ISO + 24h) en vez del formato corto + AM/PM que usa SAP.
    rows_copiadas = 0
    for row in ws_sap.iter_rows():
        for cell in row:
            new_cell = ws_new.cell(
                row=cell.row,
                column=cell.column,
                value=cell.value,
            )
            new_cell.number_format = cell.number_format
        rows_copiadas += 1

    nombre_archivo = _nombre_archivo_poblacion(sociedad, fecha_hasta)
    carpeta_destino.mkdir(parents=True, exist_ok=True)
    ruta_destino = carpeta_destino / nombre_archivo
    _guardar_workbook_seguro(wb_new, ruta_destino)

    _log(
        f"Generado: {ruta_destino.name} ({rows_copiadas} filas, "
        f"hoja '{STANDARD_SHEET_NAME}')"
    )
    return ruta_destino


# ---------------------------------------------------------------------------
# Helpers de captura de pantalla (evidencias para hoja IPE)
# ---------------------------------------------------------------------------
#
# Todas las capturas son SOFT-FAIL: si PIL no está disponible, si SAP no
# expone los atributos esperados del grid, o si el diálogo Propiedades de
# Windows no abre, se loguea y se continúa. El flujo SOX nunca se rompe
# por una evidencia fallida — el archivo IPE simplemente reportará qué
# capturas faltaron.

# Lista módulo-level de HWNDs minimizados durante el flujo de capturas
# IPE, para que `_restaurar_ventanas_app` los pueda volver a mostrar al
# final. Se popula en `_minimizar_ventanas_app` y se vacía en
# `_restaurar_ventanas_app`. Compartida por toda la corrida de
# `generar_reporte_sox`.
_VENTANAS_MINIMIZADAS_PARA_CAPTURA: list[int] = []


def _minimizar_ventanas_app() -> None:
    """Minimiza las ventanas de la app Tkinter (títulos en
    `TITULOS_VENTANA_APP`) para que las capturas IPE no incluyan la UI.

    Trackea los HWNDs minimizados en `_VENTANAS_MINIMIZADAS_PARA_CAPTURA`
    para que `_restaurar_ventanas_app` pueda revertir al final. Idempotente:
    si una ventana ya está minimizada (IsIconic), no la toca (preserva la
    intención del usuario).

    Soft-fail si pywin32 no está disponible (no-op en macOS/Linux).
    """
    try:
        import win32gui  # type: ignore
        import win32con  # type: ignore
    except ImportError:
        return

    def callback(hwnd, _):
        try:
            if not win32gui.IsWindowVisible(hwnd):
                return
            if win32gui.IsIconic(hwnd):
                # Ya minimizada por el usuario; no la tocamos.
                return
            titulo = win32gui.GetWindowText(hwnd)
            if titulo in TITULOS_VENTANA_APP:
                win32gui.ShowWindow(hwnd, win32con.SW_MINIMIZE)
                _VENTANAS_MINIMIZADAS_PARA_CAPTURA.append(hwnd)
                _log(f"  Ventana minimizada para capturas: {titulo!r}")
        except Exception:
            pass

    try:
        win32gui.EnumWindows(callback, None)
    except Exception as exc:
        _log(f"  Error enumerando ventanas para minimizar: {exc!r}")


def _restaurar_ventanas_app() -> None:
    """Restaura las ventanas que `_minimizar_ventanas_app` minimizó
    durante esta corrida. Soft-fail si pywin32 no está."""
    if not _VENTANAS_MINIMIZADAS_PARA_CAPTURA:
        return
    try:
        import win32gui  # type: ignore
        import win32con  # type: ignore
    except ImportError:
        _VENTANAS_MINIMIZADAS_PARA_CAPTURA.clear()
        return

    for hwnd in _VENTANAS_MINIMIZADAS_PARA_CAPTURA:
        try:
            win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
        except Exception:
            pass
    _log(
        f"  Restauradas {len(_VENTANAS_MINIMIZADAS_PARA_CAPTURA)} ventana(s) "
        f"de la app tras el flujo de capturas."
    )
    _VENTANAS_MINIMIZADAS_PARA_CAPTURA.clear()


def _verificar_capturas_disponibles() -> bool:
    """Diagnóstico ejecutado al inicio del flujo SOX. Verifica que PIL +
    `ImageGrab.grab()` estén disponibles y funcionales para capturar
    pantalla; loguea el resultado para que el usuario sepa de antemano
    si las capturas de la hoja IPE estarán vacías y por qué.

    No retorna False de forma bloqueante — el flujo continúa igual; sólo
    sirve para diagnóstico temprano.
    """
    try:
        from PIL import __version__ as pil_version
    except ImportError:
        _log(
            "⚠ Pillow NO está instalado en este entorno Python. "
            "La hoja IPE del Población quedará con 'Captura no disponible' "
            "en cada slot. Solución: `pip install -r requirements.txt` "
            "(o `pip install Pillow`)."
        )
        return False

    try:
        from PIL import ImageGrab
    except ImportError as exc:
        _log(
            f"⚠ Pillow {pil_version} instalado pero PIL.ImageGrab no se "
            f"pudo importar — capturas IPE no funcionarán. Detalle: {exc!r}"
        )
        return False

    try:
        # Captura de prueba mínima (1 pixel) para validar que ImageGrab
        # funciona en este sistema (no es remote desktop sin permisos,
        # session 0 isolation, etc.).
        ImageGrab.grab(bbox=(0, 0, 1, 1))
        _log(
            f"OK — Pillow {pil_version} + ImageGrab.grab() funcional. "
            f"Las capturas de la hoja IPE deberían funcionar."
        )
        return True
    except Exception as exc:
        _log(
            f"⚠ Pillow {pil_version} importado pero ImageGrab.grab(1x1) "
            f"falló — capturas IPE estarán vacías. Detalle: "
            f"{type(exc).__name__}: {exc}"
        )
        return False


def _capturar_pantalla(output_path: Path) -> bool:
    """Captura toda la pantalla primaria (incluyendo barra de tareas) y la
    guarda como PNG en `output_path`. Returns True si éxito.

    Si falla, loguea el tipo de excepción + las primeras líneas del
    traceback para facilitar diagnóstico (vs sólo `repr(exc)` que a veces
    no da pistas suficientes).
    """
    try:
        from PIL import ImageGrab
    except ImportError:
        _log(
            f"  Captura omitida (Pillow no instalado): {output_path.name}. "
            f"Solución: `pip install -r requirements.txt`."
        )
        return False
    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        img = ImageGrab.grab()
        img.save(output_path, "PNG")
        _log(
            f"  Captura guardada: {output_path.name} "
            f"({img.width}x{img.height})"
        )
        return True
    except Exception as exc:
        import traceback
        _log(
            f"  Error capturando pantalla {output_path.name}: "
            f"{type(exc).__name__}: {exc}"
        )
        # Loguear el traceback (limitado) para diagnosticar el origen real.
        for line in traceback.format_exc().splitlines()[-5:]:
            _log(f"    {line}")
        return False


def _esperar_archivo_listo(
    archivo: Path, timeout_seg: float = 10.0, poll_seg: float = 0.5
) -> bool:
    """Espera a que `archivo` exista y su tamaño se estabilice (no esté
    siendo escrito todavía). Retorna True si el archivo está listo dentro
    del timeout, False en caso contrario.

    Necesario porque SAP a veces tarda 1-3s en cerrar el handle del archivo
    después de exportar, y abrir el diálogo Propiedades sobre un archivo
    aún siendo escrito produce "Las propiedades para este archivo no están
    disponibles" en Windows.
    """
    inicio = time.time()
    tamano_previo = -1
    while time.time() - inicio < timeout_seg:
        if archivo.exists():
            tamano = archivo.stat().st_size
            if tamano > 0 and tamano == tamano_previo:
                return True
            tamano_previo = tamano
        time.sleep(poll_seg)
    return False


def _capturar_propiedades_archivo(archivo: Path, output_path: Path) -> bool:
    """Abre el diálogo Propiedades del archivo en Windows vía
    `ShellExecuteExW` (API nativa, no COM), captura screenshot, y cierra
    el diálogo con Escape.

    Solo funciona en Windows. No-op en otros sistemas (ctypes.wintypes no
    existe).

    Usamos `ShellExecuteExW` en vez de `Shell.Application.NameSpace +
    ParseName + InvokeVerb('Properties')` porque ese camino COM es
    inestable cuando el thread tiene un apartment COM compartido con SAP
    GUI Scripting — termina mostrando "Las propiedades para este archivo
    no están disponibles" (caía al namespace del Escritorio).
    """
    try:
        import ctypes
        from ctypes import wintypes
    except ImportError:
        _log(
            f"  Captura propiedades omitida (ctypes.wintypes no "
            f"disponible — no es Windows): {output_path.name}"
        )
        return False

    if not archivo.exists():
        _log(
            f"  Captura propiedades omitida (archivo no existe): "
            f"{archivo.name}"
        )
        return False

    # SAP cierra el handle del archivo de forma asíncrona; esperar a que
    # esté listo evita el error "propiedades no disponibles".
    if not _esperar_archivo_listo(archivo):
        _log(
            f"  Archivo {archivo.name} no se estabilizó en 10s; "
            f"intentando Propiedades igual."
        )

    try:
        # SHELLEXECUTEINFOW + ShellExecuteExW: API canónica de Windows
        # para invocar verbos del shell. fMask = INVOKEIDLIST hace que
        # Properties abra el diálogo del shell estándar (no el de la
        # aplicación que abre el archivo).
        SEE_MASK_INVOKEIDLIST = 0x0000000C
        SW_SHOWNORMAL = 1

        class SHELLEXECUTEINFOW(ctypes.Structure):
            _fields_ = [
                ("cbSize", wintypes.DWORD),
                ("fMask", wintypes.ULONG),
                ("hwnd", wintypes.HWND),
                ("lpVerb", wintypes.LPCWSTR),
                ("lpFile", wintypes.LPCWSTR),
                ("lpParameters", wintypes.LPCWSTR),
                ("lpDirectory", wintypes.LPCWSTR),
                ("nShow", ctypes.c_int),
                ("hInstApp", wintypes.HINSTANCE),
                ("lpIDList", ctypes.c_void_p),
                ("lpClass", wintypes.LPCWSTR),
                ("hkeyClass", wintypes.HKEY),
                ("dwHotKey", wintypes.DWORD),
                ("hIcon", wintypes.HANDLE),
                ("hProcess", wintypes.HANDLE),
            ]

        sei = SHELLEXECUTEINFOW()
        sei.cbSize = ctypes.sizeof(sei)
        sei.fMask = SEE_MASK_INVOKEIDLIST
        sei.lpVerb = "properties"
        # `.resolve()` para garantizar path absoluto — paths relativos
        # confunden al shell y pueden caer al Escritorio.
        sei.lpFile = str(archivo.resolve())
        sei.lpDirectory = str(archivo.parent.resolve())
        sei.nShow = SW_SHOWNORMAL

        ok_invoke = ctypes.windll.shell32.ShellExecuteExW(ctypes.byref(sei))
        if not ok_invoke:
            err = ctypes.get_last_error() or 0
            _log(
                f"  ShellExecuteExW (verb=properties) falló (Win32 error "
                f"{err}) para {archivo.name}"
            )
            return False

        # Dar tiempo a Windows para renderizar el diálogo de Propiedades.
        time.sleep(2.5)
        capturado = _capturar_pantalla(output_path)
        # Cerrar el diálogo con ESC (VK_ESCAPE = 0x1B).
        # keybd_event: bScan=0, dwFlags=0 (key down), then dwFlags=2 (key up).
        ctypes.windll.user32.keybd_event(0x1B, 0, 0, 0)
        ctypes.windll.user32.keybd_event(0x1B, 0, 2, 0)
        time.sleep(0.3)  # dar tiempo a que se cierre antes de continuar
        return capturado
    except Exception as exc:
        _log(f"  Error capturando propiedades de {archivo.name}: {exc!r}")
        return False


def _scroll_grid_a_primero(session) -> bool:
    """Posiciona el grid de AR15 en su primer registro (fila 0). Soft-fail."""
    try:
        grid = session.findById(DOCS_GRID_SHELL)
        grid.firstVisibleRow = 0
        grid.currentCellRow = 0
        grid.selectedRows = "0"
        time.sleep(0.5)
        return True
    except Exception as exc:
        _log(f"  Scroll al primer registro falló: {exc!r}")
        return False


def _scroll_grid_a_ultimo(session) -> bool:
    """Posiciona el grid de AR15 en su último registro. Soft-fail.

    Usa `RowCount` para encontrar el último índice, ajusta `firstVisibleRow`
    para mostrar las últimas filas, y selecciona la última.
    """
    try:
        grid = session.findById(DOCS_GRID_SHELL)
        total = grid.RowCount
        if total <= 0:
            _log("  Grid sin filas, no hay último registro para mostrar")
            return False
        ultimo = total - 1
        # Mostrar las últimas ~20 filas para que el último sea visible con contexto.
        grid.firstVisibleRow = max(0, ultimo - 20)
        grid.currentCellRow = ultimo
        grid.selectedRows = str(ultimo)
        time.sleep(0.5)
        return True
    except Exception as exc:
        _log(f"  Scroll al último registro falló: {exc!r}")
        return False


def generar_hoja_ipe(
    archivo_poblacion: Path,
    screenshots_dir: Path,
) -> dict[str, int]:
    """Añade (o reemplaza) la hoja `IPE` al workbook Población con los
    screenshots embedded.

    Lee los archivos PNG listados en `IPE_SCREENSHOTS_INFO` desde
    `screenshots_dir` y los embebe en una nueva hoja `IPE` precedidos por
    sus descripciones. Soft-fail: si un screenshot no existe, se escribe
    una nota en su lugar y se reporta en el dict de retorno.

    Args:
        archivo_poblacion: workbook destino (ya debe tener Original_SAP y
            Creados).
        screenshots_dir: carpeta temporal donde están los .png.

    Returns:
        Dict con stats: `embedded` (cuántos se embebieron OK),
        `missing` (cuántos faltaron), `missing_names` (lista de filenames).
    """
    from openpyxl.drawing.image import Image as XlsxImage

    _log(f"Generando hoja '{IPE_SHEET_NAME}' en {archivo_poblacion.name}...")
    wb = load_workbook(archivo_poblacion)
    if IPE_SHEET_NAME in wb.sheetnames:
        del wb[IPE_SHEET_NAME]
    ws = wb.create_sheet(IPE_SHEET_NAME)

    bold_lg = Font(bold=True, size=14)
    bold = Font(bold=True)
    ws.cell(1, 1, "IPE — Evidencias del proceso de generación").font = bold_lg

    embedded = 0
    missing_names: list[str] = []
    fila = 3

    for filename, descripcion in IPE_SCREENSHOTS_INFO:
        img_path = screenshots_dir / filename
        ws.cell(fila, 1, descripcion).font = bold
        fila += 1

        if not img_path.exists():
            ws.cell(fila, 1, "  ⚠ Captura no disponible (falló en el momento de tomarla)")
            missing_names.append(filename)
            fila += 3
            continue

        try:
            img = XlsxImage(str(img_path))
            # Escalar manteniendo aspect ratio si excede el ancho máximo.
            if img.width > IPE_IMAGE_MAX_WIDTH:
                ratio = IPE_IMAGE_MAX_WIDTH / img.width
                img.height = int(img.height * ratio)
                img.width = IPE_IMAGE_MAX_WIDTH
            anchor = f"A{fila}"
            ws.add_image(img, anchor)
            # Reservar filas debajo de la imagen (aprox 18px por fila).
            filas_imagen = max(1, img.height // 18)
            fila += filas_imagen + 2
            embedded += 1
        except Exception as exc:
            ws.cell(fila, 1, f"  ⚠ Error embebiendo imagen: {exc!r}")
            missing_names.append(filename)
            fila += 3

    _guardar_workbook_seguro(wb, archivo_poblacion)

    _log(
        f"Hoja '{IPE_SHEET_NAME}' generada: {embedded} screenshots embebidos, "
        f"{len(missing_names)} faltantes."
    )
    return {
        "embedded": embedded,
        "missing": len(missing_names),
        "missing_names": missing_names,
    }


def generar_hoja_creados(archivo_poblacion: Path) -> dict[str, int]:
    """Añade (o reemplaza) la hoja `Creados` al workbook Población.

    Lee la hoja `Original_SAP`, filtra las filas donde la columna G es
    exactamente `*** creado ***`, parsea la columna D con `PATRON_AF` para
    extraer (código, subnúmero, denominación), y escribe el resultado en
    una hoja `Creados` con la estructura:

      - Filas 1-9: bloque de observaciones (textos explicativos).
      - Fila 10: headers en negrita.
      - Filas 11+: datos. Columnas A-B preservan `number_format` de Fecha
        y Hora del original (para que se vean iguales). Columna K (primeros
        2 dígitos del código) se escribe como texto (`number_format = "@"`)
        para preservar ceros a la izquierda. Columna L = clasificación.

    Si la hoja `Creados` ya existe, se borra y recrea. El archivo se
    guarda sobre sí mismo.

    Args:
        archivo_poblacion: ruta al .xlsx Población que tiene Original_SAP.

    Returns:
        Dict con stats:
          - filas_filtradas: cuántas filas matchearon el filtro.
          - filas_escritas: cuántas filas se escribieron en Creados
            (= filtradas - descartadas por regex inválida).
          - filas_descartadas: cuántas pasaron filtro pero la col D no
            matcheó `PATRON_AF` (se omiten y se loguean).

    Raises:
        FileNotFoundError: si el archivo no existe.
        ValueError: si el workbook no tiene la hoja `Original_SAP`.
    """
    if not archivo_poblacion.exists():
        raise FileNotFoundError(
            f"No existe el archivo Población: {archivo_poblacion}"
        )

    _log(f"Generando hoja '{CREADOS_SHEET_NAME}' en {archivo_poblacion.name}...")
    wb = load_workbook(archivo_poblacion)

    if STANDARD_SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"El workbook {archivo_poblacion.name} no tiene la hoja "
            f"'{STANDARD_SHEET_NAME}'. ¿Es un archivo Población_* válido?"
        )

    ws_src = wb[STANDARD_SHEET_NAME]

    # Tomar el number_format de Fecha y Hora de la primera fila de datos
    # del original para replicarlo en Creados (mantiene la apariencia).
    if ws_src.max_row >= 2:
        fecha_nf = ws_src.cell(2, 1).number_format
        hora_nf = ws_src.cell(2, 2).number_format
    else:
        fecha_nf = "mm-dd-yy"
        hora_nf = "[$-F400]h:mm:ss\\ AM/PM"

    # Filtrar + parsear en una sola pasada para no duplicar memoria.
    filas_filtradas = 0
    filas_descartadas = 0
    filas_para_escribir: list[tuple] = []

    for row in ws_src.iter_rows(min_row=2, values_only=True):
        # Skip filas con menos de 8 columnas (datos truncados / vacíos).
        if len(row) < 8:
            continue

        col_g = row[6]
        if col_g != CREADOS_FILTRO_VALOR:
            continue
        filas_filtradas += 1

        col_d = row[3]
        if not isinstance(col_d, str):
            filas_descartadas += 1
            _log(
                f"  Fila descartada (col D no es texto): {col_d!r}"
            )
            continue

        match = PATRON_AF.match(col_d)
        if not match:
            filas_descartadas += 1
            _log(
                f"  Fila descartada (col D no matchea regex): {col_d!r}"
            )
            continue

        codigo = int(match.group(1))
        subnumero = int(match.group(2))
        denominacion = match.group(3)

        filas_para_escribir.append((
            row[0],         # A: Fecha
            row[1],         # B: Hora
            row[2],         # C: Usuario
            codigo,         # D: Activo Fijo (int)
            subnumero,      # E: Subnúmero (int)
            denominacion,   # F: Denominación
            row[4],         # G: Valor de objeto ampliado
            row[5],         # H: Denominación de atributo
            row[6],         # I: Valor editado nuevo (siempre "*** creado ***")
            row[7],         # J: Valor editado antiguo
        ))

    # Recrear la hoja desde cero (idempotente).
    if CREADOS_SHEET_NAME in wb.sheetnames:
        del wb[CREADOS_SHEET_NAME]
    ws_dst = wb.create_sheet(CREADOS_SHEET_NAME)

    # Bloque de observaciones (filas 1-9).
    for fila, columna, texto in CREADOS_OBSERVACIONES:
        ws_dst.cell(fila, columna, texto)

    # Headers en fila 10 (negrita).
    bold = Font(bold=True)
    for col_idx, header in enumerate(CREADOS_HEADERS, start=1):
        cell = ws_dst.cell(10, col_idx, header)
        cell.font = bold

    # Datos desde fila 11.
    for offset, datos in enumerate(filas_para_escribir):
        fila_excel = 11 + offset
        ws_dst.cell(fila_excel, 1, datos[0]).number_format = fecha_nf
        ws_dst.cell(fila_excel, 2, datos[1]).number_format = hora_nf
        ws_dst.cell(fila_excel, 3, datos[2])
        ws_dst.cell(fila_excel, 4, datos[3])
        ws_dst.cell(fila_excel, 5, datos[4])
        ws_dst.cell(fila_excel, 6, datos[5])
        ws_dst.cell(fila_excel, 7, datos[6])
        ws_dst.cell(fila_excel, 8, datos[7])
        ws_dst.cell(fila_excel, 9, datos[8])
        ws_dst.cell(fila_excel, 10, datos[9])

        # IMPORTANTE: las fórmulas se ESCRIBEN en inglés con `,` como
        # separador porque el estándar OOXML del .xlsx así lo exige. Excel
        # las TRADUCE al locale del usuario al mostrarlas (Excel-ES verá
        # =EXTRAE(...) y =SI(...) en la barra de fórmulas). Escribirlas en
        # español directamente en el XML hace que Excel reporte el archivo
        # como dañado ("Hemos encontrado un problema con contenido...").
        #
        # K: =MID(D{n},1,2). MID siempre devuelve texto, así que K queda
        # como string al ser evaluada (preserva ceros a la izquierda). El
        # number_format="@" refuerza el tipo declarativamente. Usar fórmula
        # permite que el usuario modifique D y vea K/L recalcularse.
        cell_k = ws_dst.cell(fila_excel, 11, f"=MID(D{fila_excel},1,2)")
        cell_k.number_format = "@"

        # L: IF anidado clasificando los primeros 2 dígitos. Usamos IF (no
        # IFS) porque IFS es "future function" y necesitaría prefijo
        # `_xlfn.IFS` para ser portable; IF es universal y compatible con
        # cualquier versión/locale de Excel sin trucos.
        #
        # Equivalente lógico:
        #   K="19"        → "Intangible"
        #   K="20" o "14" → "Activo Construcción"
        #   cualquier otro → "PPE"
        ws_dst.cell(
            fila_excel, 12,
            f'=IF(K{fila_excel}="19","Intangible",'
            f'IF(K{fila_excel}="20","Activo Construcción",'
            f'IF(K{fila_excel}="14","Activo Construcción","PPE")))'
        )

    _guardar_workbook_seguro(wb, archivo_poblacion)

    filas_escritas = len(filas_para_escribir)
    _log(
        f"Hoja '{CREADOS_SHEET_NAME}' generada: {filas_escritas} filas escritas "
        f"({filas_filtradas} matchearon filtro, "
        f"{filas_descartadas} descartadas por regex)."
    )
    return {
        "filas_filtradas": filas_filtradas,
        "filas_escritas": filas_escritas,
        "filas_descartadas": filas_descartadas,
    }


def generar_reporte_sox(
    session,
    sociedad: str,
    fecha_desde: str,
    fecha_hasta: str,
    carpeta_destino: str | None = None,
    nombre_archivo: str | None = None,
) -> tuple[str, str]:
    """Ejecuta el flujo SOX completo y devuelve (carpeta, nombre) del
    archivo final con nombre estándar `Población_*`.

    Flujo (7 etapas):
        1. Abrir AR15.
        2. Llenar parámetros (sociedad + fechas).
        3. Capturar screenshot 1 (parámetros).
        4. Ejecutar reporte (F8). Capturar screenshots 2 y 3 (primer/último
           registro de la tabla, con scroll).
        5. Exportar grid a .xlsx (`SOX_*.xlsx`). Capturar screenshot 4
           (status bar con bytes) y 5 (propiedades del archivo en Windows).
        6. Generar `Población_{SOC}_{FECHA_HASTA}.xlsx` con hojas
           `Original_SAP` (copia) y `Creados` (filtro + parseo).
        7. Embeber las 5 evidencias en una hoja `IPE` dentro del Población.

    Si `EXPORT_METHOD` es `None`, omite las etapas 6 y 7 (no hay archivo
    SAP) y devuelve el nombre intermedio. Los screenshots se capturan en
    un tempdir que se limpia automáticamente al salir.

    Args:
        session: sesión SAP GUI activa.
        sociedad: código de sociedad (debe estar en VALID_SOCIEDADES).
        fecha_desde: fecha inicial en formato dd.mm.aaaa.
        fecha_hasta: fecha final en formato dd.mm.aaaa.
        carpeta_destino: ruta donde guardar los .xlsx (default: salida/).
        nombre_archivo: nombre del intermedio SAP (default: SOX_{soc}_{ts}.xlsx).

    Returns:
        (carpeta, nombre) del archivo `Población_*` final. Si EXPORT_METHOD
        es None, devuelve (carpeta, nombre del intermedio SAP).
    """
    import tempfile

    sociedad_norm = validar_sociedad(sociedad)
    validar_rango_fechas(fecha_desde, fecha_hasta)

    if carpeta_destino is None:
        SALIDA_DIR.mkdir(parents=True, exist_ok=True)
        carpeta_destino = str(SALIDA_DIR)
    if nombre_archivo is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"SOX_{sociedad_norm}_{ts}.xlsx"

    # Chequeo temprano: si el Población predicho ya existe Y está abierto
    # en Excel, abortar AHORA para no correr el flujo SAP completo (minutos)
    # y fallar al final con `PermissionError` al intentar escribirlo.
    archivo_poblacion_predicho = (
        Path(carpeta_destino) / _nombre_archivo_poblacion(sociedad_norm, fecha_hasta)
    )
    if _archivo_esta_bloqueado(archivo_poblacion_predicho):
        raise PermissionError(
            f"El archivo {archivo_poblacion_predicho.name} ya existe y está "
            f"abierto en Excel. Por favor ciérralo antes de generar el "
            f"reporte SOX (de lo contrario el flujo SAP correrá inútilmente "
            f"y fallará al guardar el resultado).\n\n"
            f"Ruta: {archivo_poblacion_predicho}"
        )

    inicio = time.monotonic()
    _log("=== Iniciando flujo SOX ===")
    # Diagnóstico temprano: si las capturas no van a funcionar (Pillow no
    # instalado, ImageGrab roto), el usuario lo ve YA en los logs, no al
    # final cuando ve la hoja IPE vacía.
    _verificar_capturas_disponibles()
    abrir_transaccion_sox(session)
    ingresar_parametros(session, sociedad_norm, fecha_desde, fecha_hasta)

    # Tempdir para las 5 capturas; se limpia automáticamente al salir del
    # `with`. Las capturas se embeben en la hoja IPE del Población antes
    # de que tmpdir desaparezca.
    with tempfile.TemporaryDirectory(prefix="sox_evidencias_") as tmp:
        screenshots_dir = Path(tmp)

        # Minimizar la ventana de la app Tkinter antes de capturar para
        # que las screenshots IPE muestren SAP limpio sin la UI de la
        # app encima. Se restaura al final (try/finally) pase lo que pase.
        _minimizar_ventanas_app()
        # Dar tiempo a que la animación de minimizar termine antes de la
        # primera captura.
        time.sleep(0.5)

        try:
            # Screenshot 1: parámetros llenados, ANTES de F8.
            _capturar_pantalla(screenshots_dir / "01_parametros_ingresados.png")

            ejecutar_reporte(session)
            # Dar tiempo a que el grid de AR15 renderice los resultados antes
            # de scrollear y capturar.
            time.sleep(1.5)

            # Screenshot 2: primer registro de la tabla.
            _scroll_grid_a_primero(session)
            _capturar_pantalla(screenshots_dir / "02_primer_registro.png")

            # Screenshot 3: último registro (scroll al final).
            _scroll_grid_a_ultimo(session)
            _capturar_pantalla(screenshots_dir / "03_ultimo_registro.png")
            # Restablecer la posición del cursor antes de exportar (algunas
            # versiones de SAP exportan desde la fila seleccionada).
            _scroll_grid_a_primero(session)

            exportar_a_excel(session, carpeta_destino, nombre_archivo)

            # Screenshot 4: status bar SAP con bytes recién exportados.
            _capturar_pantalla(screenshots_dir / "04_status_bar_bytes.png")

            # Screenshot 5: diálogo Propiedades del archivo SAP en Windows.
            archivo_sap = Path(carpeta_destino) / nombre_archivo
            _capturar_propiedades_archivo(
                archivo_sap,
                screenshots_dir / "05_propiedades_archivo.png",
            )

            if EXPORT_METHOD is None:
                _log(
                    "EXPORT_METHOD=None → omitiendo Población y hoja IPE "
                    "(no hay archivo SAP del cual leer)."
                )
                volver_a_pantalla_inicial(session)
                duracion = time.monotonic() - inicio
                _log(f"=== Flujo SOX finalizado en {duracion:.1f}s ===")
                return carpeta_destino, nombre_archivo

            archivo_poblacion = generar_xlsx_poblacion(
                archivo_sap,
                Path(carpeta_destino),
                sociedad_norm,
                fecha_hasta,
            )

            # Etapas post-procesamiento: añadir hojas Creados e IPE al
            # workbook Población. IPE va al final porque embebe las capturas
            # del tempdir antes de que éste se elimine al salir del `with`.
            generar_hoja_creados(archivo_poblacion)
            generar_hoja_ipe(archivo_poblacion, screenshots_dir)
        finally:
            # Restaurar la ventana de la app aunque haya fallado alguna
            # etapa — si no, el usuario se queda sin GUI visible.
            _restaurar_ventanas_app()

    # Devolver SAP a la pantalla inicial para que la siguiente sociedad del
    # multiselect pueda abrir AR15 de nuevo (extraído de ScriptanexoREP.vbs:
    # dos "Atrás" desde la pantalla de resultados). El caso de fallo a mitad
    # de flujo se cubre con el prefijo `/nAR15` de `abrir_transaccion_sox`.
    volver_a_pantalla_inicial(session)

    duracion = time.monotonic() - inicio
    _log(f"=== Flujo SOX finalizado en {duracion:.1f}s ===")
    return str(archivo_poblacion.parent), archivo_poblacion.name


# ---------------------------------------------------------------------------
# ENTRY POINT
# ---------------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    argv = argv if argv is not None else sys.argv[1:]
    print("=" * 70, flush=True)
    print("Generación de Reporte SOX vía SAP GUI Scripting", flush=True)
    print("=" * 70, flush=True)

    if len(argv) != 3:
        print(
            "Uso: python src/sox_report.py SOCIEDAD DESDE HASTA\n"
            "Ejemplo: python src/sox_report.py ISA 01.05.2026 31.05.2026",
            file=sys.stderr,
        )
        return 2

    sociedad, desde, hasta = argv

    try:
        validar_sociedad(sociedad)
        validar_rango_fechas(desde, hasta)
    except ValueError as exc:
        print(f"ERROR de validación: {exc}", file=sys.stderr, flush=True)
        return 1

    try:
        session = get_sap_session()
    except RuntimeError as exc:
        print(f"ERROR: {exc}", file=sys.stderr, flush=True)
        return 1

    try:
        carpeta, nombre = generar_reporte_sox(session, sociedad, desde, hasta)
    except Exception as exc:
        print(f"\nERROR durante el flujo SOX: {exc}", file=sys.stderr, flush=True)
        return 1

    print(flush=True)
    print("=" * 70, flush=True)
    print(f"Reporte SOX generado: {carpeta}\\{nombre}", flush=True)
    print("=" * 70, flush=True)
    return 0


if __name__ == "__main__":
    sys.exit(main())
