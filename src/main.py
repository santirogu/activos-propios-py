import contextlib
import threading
import time
import traceback
import tkinter as tk
from tkinter import messagebox, ttk
from pathlib import Path
from datetime import datetime

import openpyxl
from tkinter import filedialog
from tkcalendar import DateEntry

import branding
from paths import (
    PROJECT_ROOT,
    SALIDA_DIR,
    asegurar_formato_dinamico,
    formato_dinamico_path,
    validar_entrada_unica,
)

OUTPUT_DIR = SALIDA_DIR
SHEET_NAME = "LSMW "

# Intervalo (ms) con que se re-evalúa el estado del botón "Subir a SAP" para
# habilitarlo/deshabilitarlo según existan o no .txt en salida/.
_POLL_INTERVAL_MS = 1000

# Flag módulo-level: True mientras un worker de subir_a_sap está corriendo.
# Sirve para que el polling NO toque el estado del botón durante la carga
# (el worker tiene control exclusivo en ese momento).
_upload_en_curso = False


def _log(mensaje: str) -> None:
    """Imprime un mensaje con timestamp [HH:MM:SS] y flush=True para que
    aparezca en tiempo real al ejecutar `python src/main.py` desde terminal."""
    ts = time.strftime("%H:%M:%S")
    print(f"[{ts}] {mensaje}", flush=True)


def _show_unexpected_error(title: str, exc: BaseException) -> None:
    """Loguea la excepción completa y muestra un diálogo con el detalle.

    Sirve como red de seguridad para excepciones que ningún `except`
    específico capturó: el usuario verá un error en pantalla en vez de
    quedarse sin retroalimentación.
    """
    tb_text = "".join(
        traceback.format_exception(type(exc), exc, exc.__traceback__)
    )
    _log(f"ERROR — {title}: {exc}")
    print(tb_text, flush=True)
    messagebox.showerror(
        title,
        f"{type(exc).__name__}: {exc}\n\n--- Detalle técnico ---\n{tb_text}",
    )


@contextlib.contextmanager
def _sap_com_apartment():
    """Inicializa el apartamento COM del thread actual y lo libera al salir.

    Windows exige `pythoncom.CoInitialize()` antes de cualquier llamada COM
    desde un thread que no sea el main de la app — sin esto, `GetObject('SAPGUI')`
    en los workers de Subir a SAP / Generar Reporte SOX falla con un error
    genérico ("No se pudo conectar a SAP GUI") aunque SAP esté abierto.

    Loguea cada paso (CoInitialize / CoUninitialize) para diagnosticar
    problemas de conexión desde threads.

    No-op en sistemas sin pythoncom (Mac/Linux).
    """
    thread_id = threading.get_ident()
    try:
        import pythoncom  # type: ignore
    except ImportError:
        _log(f"_sap_com_apartment: pythoncom no disponible (thread={thread_id}, no-op)")
        yield
        return

    _log(f"_sap_com_apartment: llamando CoInitialize() en thread={thread_id}...")
    try:
        pythoncom.CoInitialize()
        _log(f"_sap_com_apartment: CoInitialize OK (thread={thread_id})")
    except Exception as exc:
        _log(f"_sap_com_apartment: CoInitialize FALLÓ — {exc!r}")
        raise

    try:
        yield
    finally:
        try:
            pythoncom.CoUninitialize()
            _log(f"_sap_com_apartment: CoUninitialize OK (thread={thread_id})")
        except Exception as exc:
            _log(f"_sap_com_apartment: CoUninitialize falló (ignorado) — {exc!r}")


def _cerrar_splash() -> None:
    """Cierra el splash de PyInstaller si la app corre bundleada.

    `pyi_splash` es un módulo que PyInstaller INYECTA en el runtime sólo
    cuando el spec incluye un `Splash(...)` (ver GestionActivosFijos.spec).
    En dev mode (`python src/main.py`) el import falla y simplemente no
    hay splash que cerrar — es no-op. En macOS Splash no se soporta, así
    que tampoco existe.

    Se llama justo antes de `root.mainloop()` para que el splash
    desaparezca al mismo tiempo que la ventana principal se hace visible
    (UX limpio: el usuario ve la pantalla de carga → desaparece →
    aparece la app).
    """
    try:
        import pyi_splash  # noqa: I001  (módulo inyectado por PyInstaller)
    except ImportError:
        return
    try:
        if pyi_splash.is_alive():
            pyi_splash.close()
    except Exception:
        # Nunca queremos que un fallo del splash impida el arranque.
        pass


def _install_tk_exception_handler(root: tk.Tk) -> None:
    """Reemplaza el handler default de Tkinter (que solo imprime a stderr)
    por uno que muestra un diálogo. Captura cualquier excepción no manejada
    en callbacks Tkinter — sin esto, los errores son invisibles para el
    usuario que abre la app por doble-clic."""

    def handler(exc_type, exc_value, tb) -> None:
        tb_text = "".join(traceback.format_exception(exc_type, exc_value, tb))
        _log(f"ERROR no manejado en callback Tkinter: {exc_value}")
        print(tb_text, flush=True)
        messagebox.showerror(
            "Error inesperado",
            f"{exc_type.__name__}: {exc_value}\n\n"
            f"--- Detalle técnico ---\n{tb_text}",
        )

    root.report_callback_exception = handler


def _hay_txt_en_salida() -> bool:
    """True si hay al menos un archivo LSMW_*.txt en salida/."""
    return OUTPUT_DIR.exists() and any(OUTPUT_DIR.glob("LSMW_*.txt"))


def _refrescar_estado_boton_subir(button: tk.Widget) -> None:
    """Sincroniza el estado del botón con la presencia de .txt en salida/.

    Si hay un upload en curso, no toca el botón (el worker lo controla).
    """
    if _upload_en_curso:
        return
    button.config(state="normal" if _hay_txt_en_salida() else "disabled")


def export_sheet_to_tsv(
    excel_path: Path,
    sheet_name: str,
    output_dir: Path,
    file_prefix: str = "LSMW",
) -> tuple[Path, int]:
    """Lee `sheet_name` del workbook en `excel_path` y escribe un .txt
    separado por tabulación dentro de `output_dir`. Devuelve (ruta, filas)."""
    if not excel_path.exists():
        raise FileNotFoundError(f"No se encontró el archivo: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"La hoja '{sheet_name.strip()}' no existe en el archivo.")

    ws = wb[sheet_name]
    output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = output_dir / f"{file_prefix}_{ts}.txt"

    rows_written = 0
    with output_path.open("w", encoding="utf-8", newline="") as f:
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v) for v in row]
            f.write("\t".join(cells) + "\n")
            rows_written += 1

    return output_path, rows_written


def extraer_lsmw_a_txt(status_var: tk.StringVar) -> None:
    try:
        _log("Botón 'Extraer información en txt' presionado")
        _log(f"OUTPUT_DIR = {OUTPUT_DIR}")

        # Antes de leer, verificar que en la carpeta «entrada» haya UN solo
        # .xlsm. Con más de uno la elección del archivo es ambigua, así que
        # advertimos al usuario y abortamos para que deje solo el correcto.
        entrada_ok, advertencia = validar_entrada_unica()
        if not entrada_ok:
            _log(f"Conflicto en carpeta entrada/: {advertencia}")
            messagebox.showwarning("Conflicto en la carpeta «entrada»", advertencia)
            status_var.set(
                "Hay más de un Formato Dinámico en la carpeta «entrada». "
                "Deja solo uno."
            )
            return

        excel_path = formato_dinamico_path()
        _log(f"EXCEL_PATH = {excel_path}")

        # Si ya existe(n) .txt previo(s) en salida/, pedir confirmación antes
        # de reemplazar.
        existentes = (
            sorted(OUTPUT_DIR.glob("LSMW_*.txt")) if OUTPUT_DIR.exists() else []
        )
        _log(f"Archivos LSMW_*.txt previos en salida/: {len(existentes)}")
        if existentes:
            reemplazar = messagebox.askyesno(
                "Archivo ya existente",
                f"Ya existe un .txt generado en salida/:\n"
                f"  {existentes[-1].name}\n\n"
                f"¿Deseas reemplazarlo por uno nuevo?",
            )
            if not reemplazar:
                _log("Usuario canceló el reemplazo. Conservando archivo existente.")
                status_var.set(
                    "Operación cancelada. Se conservó el archivo existente."
                )
                return
            for old in existentes:
                try:
                    old.unlink()
                    _log(f"Archivo borrado: {old.name}")
                except OSError as exc:
                    _log(f"Error al borrar {old.name}: {exc}")
                    messagebox.showerror(
                        "Error al borrar archivo",
                        f"No se pudo borrar {old.name}:\n{exc}",
                    )
                    return

        _log("Generando nuevo .txt desde la hoja LSMW...")
        try:
            output_path, rows_written = export_sheet_to_tsv(
                excel_path, SHEET_NAME, OUTPUT_DIR
            )
        except FileNotFoundError as exc:
            _log(f"FileNotFoundError: {exc}")
            messagebox.showerror("Archivo no encontrado", str(exc))
            return
        except ValueError as exc:
            _log(f"ValueError: {exc}")
            messagebox.showerror("Hoja no encontrada", str(exc))
            return
        except Exception as exc:
            _log(f"Excepción durante export_sheet_to_tsv: {exc}")
            messagebox.showerror("Error al exportar", str(exc))
            return

        _log(f"Generado: {output_path.name} ({rows_written} filas)")
        status_var.set(f"Exportado: {output_path.name} ({rows_written} filas)")
        messagebox.showinfo(
            "Extracción completa",
            f"Se generó el archivo:\n{output_path}\n\nFilas exportadas: {rows_written}",
        )
    except Exception as exc:
        # Red de seguridad: cualquier excepción no prevista (acceso a
        # OUTPUT_DIR, errores de Tkinter, etc.) se muestra al usuario con
        # el traceback completo en consola.
        _show_unexpected_error("Error inesperado al extraer", exc)


def subir_a_sap(root: tk.Tk, status_var: tk.StringVar, button: tk.Widget) -> None:
    """Lanza la carga LSMW a SAP en un hilo background.

    Confirma con el usuario, deshabilita el botón mientras corre y va
    actualizando `status_var` desde el hilo principal vía `root.after`.
    """
    confirmar = messagebox.askyesno(
        "Confirmar carga a SAP",
        "Esto tomará el .txt más reciente de salida/ y ejecutará el flujo "
        "LSMW en la sesión SAP abierta.\n\n"
        "Asegúrate de:\n"
        "  • Tener SAP abierto y con sesión iniciada.\n"
        "  • Tener el proyecto LSMW pre-cargado.\n"
        "  • No tocar SAP mientras se ejecuta el script.\n\n"
        "¿Continuar?",
    )
    if not confirmar:
        return

    global _upload_en_curso
    _upload_en_curso = True
    button.config(state="disabled")

    def update_status(text: str) -> None:
        root.after(0, status_var.set, text)

    def show_info(title: str, message: str) -> None:
        root.after(0, lambda: messagebox.showinfo(title, message))

    def show_error(title: str, message: str) -> None:
        root.after(0, lambda: messagebox.showerror(title, message))

    def worker() -> None:
        global _upload_en_curso
        with _sap_com_apartment():
            try:
                try:
                    from sap_upload import (
                        get_latest_txt,
                        get_sap_session,
                        run_lsmw_flow,
                    )
                except ImportError as exc:
                    show_error(
                        "Error de import",
                        f"No se pudo importar sap_upload:\n{exc}",
                    )
                    return

                try:
                    update_status("Buscando .txt más reciente en salida/...")
                    latest = get_latest_txt()

                    update_status("Conectando a la sesión SAP...")
                    session = get_sap_session()

                    update_status("Ejecutando flujo LSMW (no toques SAP)...")
                    run_lsmw_flow(session, str(latest.parent), latest.name)

                    update_status(
                        "Carga completada. Revisa SM35 para el log de la BDC."
                    )
                    show_info(
                        "Carga completada",
                        "Flujo LSMW ejecutado correctamente.\n\n"
                        "Revisa SM35 para ver el log de la sesión BDC.",
                    )
                except Exception as exc:
                    update_status("")
                    show_error("Error en carga SAP", str(exc))
            finally:
                _upload_en_curso = False
                root.after(0, lambda: _refrescar_estado_boton_subir(button))

    threading.Thread(target=worker, daemon=True).start()


# ---------------------------------------------------------------------------
# Control SOX — diálogo de generación de reporte
# ---------------------------------------------------------------------------

def _generar_reporte_sox_handler(
    root: tk.Tk,
    sociedades: list[str],
    fecha_desde: str,
    fecha_hasta: str,
    status_var: tk.StringVar,
    button: tk.Widget,
    btn_atras: tk.Widget,
) -> None:
    """Valida los inputs y lanza el worker que genera el reporte SOX para
    CADA sociedad seleccionada (un reporte por sociedad, no consolidado).

    El form vive en `root` (no en un Toplevel); por eso usamos `root.after`
    para los callbacks thread-safe. Mientras el worker corre, deshabilita
    tanto el botón Generar como el botón Atrás (no queremos que el usuario
    vuelva al menú a mitad de un flujo SAP).

    **Soft-fail por sociedad**: si el flujo falla para una sociedad, se
    registra y se continúa con las demás; al final se muestra un resumen
    `X OK / Y con error`."""
    try:
        from sox_report import validar_sociedad, validar_rango_fechas
    except ImportError as exc:
        messagebox.showerror(
            "Error de import", f"No se pudo importar sox_report:\n{exc}"
        )
        return

    if not sociedades:
        messagebox.showerror(
            "Datos inválidos",
            "Selecciona al menos una sociedad antes de generar el reporte.",
        )
        return

    try:
        sociedades_norm = [validar_sociedad(s) for s in sociedades]
        validar_rango_fechas(fecha_desde, fecha_hasta)
    except ValueError as exc:
        messagebox.showerror("Datos inválidos", str(exc))
        return

    lista_soc = ", ".join(sociedades_norm)
    if not messagebox.askyesno(
        "Confirmar generación del reporte SOX",
        f"Se generará un reporte SOX por CADA sociedad seleccionada "
        f"({len(sociedades_norm)}):\n"
        f"  • Sociedades: {lista_soc}\n"
        f"  • Desde: {fecha_desde}\n"
        f"  • Hasta: {fecha_hasta}\n\n"
        f"Los archivos se guardarán en salida/ (uno por sociedad).\n\n"
        f"Asegúrate de tener SAP abierto y con sesión iniciada.\n\n"
        f"¿Continuar?",
    ):
        return

    button.config(state="disabled")
    btn_atras.config(state="disabled")

    def update_status(text: str) -> None:
        root.after(0, status_var.set, text)

    def show_info(title: str, message: str) -> None:
        root.after(0, lambda: messagebox.showinfo(title, message))

    def show_error(title: str, message: str) -> None:
        root.after(0, lambda: messagebox.showerror(title, message))

    def reenable() -> None:
        def _do():
            button.config(state="normal")
            btn_atras.config(state="normal")
        root.after(0, _do)

    def worker() -> None:
        with _sap_com_apartment():
            try:
                try:
                    from sox_report import generar_reporte_sox, get_sap_session
                except ImportError as exc:
                    show_error(
                        "Error de import",
                        f"No se pudo importar sox_report:\n{exc}",
                    )
                    return

                try:
                    update_status("Conectando a la sesión SAP...")
                    session = get_sap_session()
                except Exception as exc:
                    update_status("")
                    show_error("Error generando reporte SOX", str(exc))
                    return

                total = len(sociedades_norm)
                exitosos: list[str] = []
                fallidos: list[tuple[str, str]] = []
                for idx, soc in enumerate(sociedades_norm, start=1):
                    update_status(
                        f"Generando {idx}/{total}: {soc} "
                        f"({fecha_desde} → {fecha_hasta})..."
                    )
                    try:
                        carpeta, nombre = generar_reporte_sox(
                            session, soc, fecha_desde, fecha_hasta
                        )
                        exitosos.append(nombre)
                        _log(f"SOX OK ({idx}/{total}): {soc} → {nombre}")
                    except Exception as exc:
                        fallidos.append((soc, str(exc)))
                        _log(f"SOX FALLO ({idx}/{total}): {soc} — {exc!r}")

                update_status(
                    f"Finalizado: {len(exitosos)} OK, {len(fallidos)} con error"
                )

                resumen = (
                    f"Generación finalizada.\n\n"
                    f"  • Reportes OK: {len(exitosos)} / {total}\n"
                    f"  • Con error: {len(fallidos)}\n"
                )
                if exitosos:
                    resumen += "\nGenerados en salida/:\n"
                    for nombre in exitosos:
                        resumen += f"  • {nombre}\n"
                if fallidos:
                    resumen += "\nCon error:\n"
                    for soc, err in fallidos:
                        err_corto = err if len(err) <= 180 else err[:180] + "…"
                        resumen += f"  • {soc}: {err_corto}\n"

                if fallidos:
                    show_error("Reporte SOX con errores", resumen)
                else:
                    show_info("Reporte SOX generado", resumen)
            finally:
                reenable()

    threading.Thread(target=worker, daemon=True).start()


def _extraer_activos_creados_handler(
    root: tk.Tk,
    usuario_sap: str,
    button: tk.Widget,
    btn_atras: tk.Widget,
) -> None:
    """Valida el Usuario SAP y lanza el worker que ejecuta el flujo SM35P.

    Replica el patrón de `_generar_reporte_sox_handler`:
      - validación previa del input (`validar_usuario_sap`)
      - confirmación al usuario
      - deshabilita botón Ejecutar Y botón Atrás durante el worker
      - worker corre en thread daemon, comunica vía `root.after(0, ...)`
      - envuelve todo en `_sap_com_apartment()` para CoInitialize en thread
    """
    try:
        from extraer_activos_creados import validar_usuario_sap
    except ImportError as exc:
        messagebox.showerror(
            "Error de import",
            f"No se pudo importar extraer_activos_creados:\n{exc}",
        )
        return

    try:
        usuario_norm = validar_usuario_sap(usuario_sap)
    except ValueError as exc:
        messagebox.showerror("Datos inválidos", str(exc))
        return

    if not messagebox.askyesno(
        "Confirmar extracción de activos creados",
        f"Se ejecutará la transacción SM35P en SAP filtrando los logs "
        f"por el usuario:\n"
        f"  • Usuario SAP: {usuario_norm}\n\n"
        f"Asegúrate de tener SAP abierto y con sesión iniciada.\n\n"
        f"¿Continuar?",
    ):
        return

    button.config(state="disabled")
    btn_atras.config(state="disabled")

    def show_info(title: str, message: str) -> None:
        root.after(0, lambda: messagebox.showinfo(title, message))

    def show_error(title: str, message: str) -> None:
        root.after(0, lambda: messagebox.showerror(title, message))

    def reenable() -> None:
        def _do():
            button.config(state="normal")
            btn_atras.config(state="normal")
        root.after(0, _do)

    def worker() -> None:
        with _sap_com_apartment():
            try:
                try:
                    from extraer_activos_creados import (
                        get_sap_session, extraer_activos_creados,
                    )
                except ImportError as exc:
                    show_error(
                        "Error de import",
                        f"No se pudo importar extraer_activos_creados:\n{exc}",
                    )
                    return

                try:
                    session = get_sap_session()
                    carpeta, nombre = extraer_activos_creados(
                        session, usuario_norm
                    )
                    show_info(
                        "Extracción completada",
                        f"Log extraído correctamente.\n\n"
                        f"Archivo guardado en:\n{carpeta}\\{nombre}",
                    )
                except Exception as exc:
                    show_error(
                        "Error en la extracción", str(exc)
                    )
            finally:
                reenable()

    threading.Thread(target=worker, daemon=True).start()


def _subir_anexos_handler(
    root: tk.Tk,
    sociedad: str,
    archivos: list[Path],
    activos_usuario: list[tuple[int, int]] | None,
    nombre_archivo_usuario: str | None,
    status_var: tk.StringVar,
    button: tk.Widget,
    btn_atras: tk.Widget,
) -> None:
    """Valida la sociedad + archivos y lanza el worker que ejecuta el
    flujo de subida de anexos a cada activo.

    La lista de activos depende de si el usuario cargó un `.xlsx` propio:
      - `activos_usuario` no None → se usan esos activos (override).
      - `activos_usuario` None → se usa el último `ActivosCreados_*.xlsx`
        de `salida/` (comportamiento por defecto de "Extraer Activos
        Creados").

    Patrón consistente con `_extraer_activos_creados_handler`:
      - Validaciones previas (sociedad en la lista, al menos un archivo).
      - Confirmación al usuario.
      - Deshabilita Subir + Atrás durante el worker.
      - Worker en thread daemon, envuelto en `_sap_com_apartment()`.
      - `progress_callback` actualiza `status_var` vía `root.after`.
      - Soft-fail: el módulo loguea + acumula fallos; al final muestra
        resumen "X OK, Y fallos".
    """
    try:
        from subir_anexos import validar_sociedad
    except ImportError as exc:
        messagebox.showerror(
            "Error de import",
            f"No se pudo importar subir_anexos:\n{exc}",
        )
        return

    try:
        sociedad_norm = validar_sociedad(sociedad)
    except ValueError as exc:
        messagebox.showerror("Datos inválidos", str(exc))
        return

    if not archivos:
        messagebox.showerror(
            "Sin archivos",
            "Selecciona al menos un archivo antes de subir anexos.",
        )
        return

    if activos_usuario is not None:
        origen_desc = (
            f"los {len(activos_usuario)} activo(s) del archivo "
            f"'{nombre_archivo_usuario}'"
        )
    else:
        origen_desc = (
            "los activos de la hoja 'Activos Fijos' del último "
            "ActivosCreados_*.xlsx en salida/"
        )

    if not messagebox.askyesno(
        "Confirmar subida de anexos",
        f"Se subirán {len(archivos)} archivo(s) como adjuntos a CADA uno de "
        f"{origen_desc}.\n\n"
        f"  • Sociedad: {sociedad_norm}\n"
        f"  • Archivos: {len(archivos)}\n\n"
        f"Esto puede tomar varios minutos dependiendo del número de "
        f"activos. NO interactúes con SAP durante el proceso.\n\n"
        f"¿Continuar?",
    ):
        return

    button.config(state="disabled")
    btn_atras.config(state="disabled")

    def update_status(text: str) -> None:
        root.after(0, status_var.set, text)

    def show_info(title: str, message: str) -> None:
        root.after(0, lambda: messagebox.showinfo(title, message))

    def show_error(title: str, message: str) -> None:
        root.after(0, lambda: messagebox.showerror(title, message))

    def reenable() -> None:
        def _do():
            button.config(state="normal")
            btn_atras.config(state="normal")
        root.after(0, _do)

    def progress_callback(intento: int, total: int, desc: str) -> None:
        update_status(f"Subiendo {intento}/{total}: {desc}")

    def worker() -> None:
        with _sap_com_apartment():
            try:
                try:
                    from subir_anexos import (
                        get_sap_session, subir_anexos,
                    )
                except ImportError as exc:
                    show_error(
                        "Error de import",
                        f"No se pudo importar subir_anexos:\n{exc}",
                    )
                    return

                try:
                    update_status("Conectando a la sesión SAP...")
                    session = get_sap_session()

                    update_status(
                        f"Preparando subida de {len(archivos)} "
                        f"archivo(s) para sociedad {sociedad_norm}..."
                    )
                    stats = subir_anexos(
                        session,
                        sociedad_norm,
                        archivos,
                        activos=activos_usuario,
                        progress_callback=progress_callback,
                    )

                    update_status(
                        f"Finalizado: {stats['exitosos']} OK, "
                        f"{stats['fallidos']} fallos"
                    )

                    resumen = (
                        f"Subida finalizada.\n\n"
                        f"  • Exitosos: {stats['exitosos']} / "
                        f"{stats['total_intentos']}\n"
                        f"  • Fallidos: {stats['fallidos']}\n"
                    )
                    if stats["detalles_fallos"]:
                        resumen += "\nPrimeros fallos:\n"
                        for activo, sub, archivo, error in (
                            stats["detalles_fallos"][:3]
                        ):
                            nombre = Path(archivo).name
                            # Truncar mensajes muy largos para que el
                            # messagebox quepa en pantalla.
                            err_corto = (
                                error if len(error) <= 220
                                else error[:220] + "…"
                            )
                            resumen += (
                                f"\n  • {activo}-{sub} / {nombre}:\n"
                                f"    {err_corto}\n"
                            )
                        if len(stats["detalles_fallos"]) > 3:
                            resumen += (
                                f"\n  ... y "
                                f"{len(stats['detalles_fallos']) - 3} más "
                                f"(ver consola para detalles)\n"
                            )

                    if stats["fallidos"] == 0:
                        show_info("Subida completada", resumen)
                    else:
                        show_error("Subida con fallos", resumen)
                except Exception as exc:
                    update_status("")
                    show_error("Error en la subida", str(exc))
            finally:
                reenable()

    threading.Thread(target=worker, daemon=True).start()


def _crear_footer_copyright(parent: tk.Misc) -> tk.Label:
    """Footer pequeño con copyright + año actual, centrado abajo.

    Se llama UNA vez sobre `root` desde `main()` y persiste a través del
    switching de sub-vistas: como se empaca con `side="bottom"` ANTES
    de cualquier sub-frame, claimea el strip inferior del root y los
    frames que se montan después (vía `pack(fill="both", expand=True)`)
    ocupan el espacio restante por encima.

    Tamaño y color discretos (Helvetica 8pt + ISA_GRIS_CLARO) para que
    no compita visualmente con el contenido principal. Ver LICENSE en
    la raíz del repo para los términos completos.
    """
    año = datetime.now().year
    footer = tk.Label(
        parent,
        text=f"© {año} El Hub de ISA · Todos los derechos reservados",
        font=("Helvetica", 8),
        fg=branding.ISA_GRIS_CLARO,
        bg=branding.ISA_FONDO,
    )
    footer.pack(side="bottom", pady=(0, 6))
    return footer


def _crear_card_visual(
    parent: tk.Frame,
    *,
    titulo: str,
    descripcion: str,
    btn_texto: str,
    btn_style: str,
    command=None,
    disabled: bool = False,
    width: int | None = None,
    height: int | None = None,
) -> tuple[tk.Frame, branding.RoundedButton]:
    """Construye una "card" del menú principal (estilo maqueta de diseño).

    Estructura vertical (sin icono — decisión de UX):
        ┌────────────────────────┐
        │       Título           │  ← 12pt bold navy
        │  Descripción centrada  │  ← 8pt gris, multi-línea
        │  en gris suave.        │
        │  [  Botón color   ]    │  ← color según btn_style
        └────────────────────────┘

    Borde plano 1 px en `ISA_GRIS_BORDE` (Tk no soporta drop-shadow).

    Si se pasan `width` y `height`, la card se fuerza a esas dimensiones
    exactas con `pack_propagate(False)` — útil para que las 3 cards del
    menú queden simétricas pese a tener descripciones de distinta
    longitud. Si se omiten, la card auto-dimensiona a su contenido.

    NO se hace `.pack()` aquí — el caller controla el layout (top row
    side='left' vs bottom row centered). Esto da la flexibilidad para
    el layout 2+1 (2 cards arriba, Reportes centrada abajo).

    Args:
        parent: contenedor donde vivirá la card.
        titulo: encabezado de la card (ej. "Activos Fijos").
        descripcion: párrafo gris debajo del título. Multi-línea OK.
        btn_texto: texto del botón al pie (ej. "Acceder  →").
        btn_style: estilo de RoundedButton ("primary"/"naranja"/"verde").
        command: callback del botón. Ignorado si `disabled=True`.
        disabled: arranca con el botón en `state="disabled"`.
        width, height: si se pasan, se fuerza el tamaño exacto.

    Returns:
        Tupla `(card, boton)`:
            - card: el Frame para que el caller decida cómo empaquetarlo.
            - boton: el RoundedButton interior, expuesto para tests y
              para reconfig posterior del command/state.
    """
    card = tk.Frame(
        parent,
        bg=branding.ISA_BLANCO,
        highlightthickness=1,
        highlightbackground=branding.ISA_GRIS_BORDE,
        bd=0,
    )
    if width is not None and height is not None:
        card.configure(width=width, height=height)
        card.pack_propagate(False)

    tk.Label(
        card,
        text=titulo,
        font=("Helvetica", 12, "bold"),
        fg=branding.ISA_AZUL,
        bg=branding.ISA_BLANCO,
    ).pack(pady=(16, 6), padx=12)

    tk.Label(
        card,
        text=descripcion,
        font=("Helvetica", 8),
        fg=branding.ISA_GRIS,
        bg=branding.ISA_BLANCO,
        justify="center",
    ).pack(pady=(0, 14), padx=12)

    boton = branding.RoundedButton(
        card,
        text=btn_texto,
        style=btn_style,
        padx=12, pady=6, width=160,
        font=("Helvetica", 10, "bold"),
        command=command,
    )
    if disabled:
        boton.config(state="disabled")
    boton.pack(pady=(0, 16), padx=12)

    return card, boton


def _crear_panel_card(parent: tk.Misc) -> tk.Frame:
    """Frame con bg blanco + borde gris 1 px para envolver el contenido
    interactivo de un sub-formulario (botones, entries, listbox, etc.).

    Replica la estética de las cards del menú principal para que TODA la
    app tenga el mismo lenguaje visual: contenedores blancos con borde
    sutil sobre el fondo blanco general. Los widgets que se pasen como
    hijos heredarán el `bg` del panel automáticamente porque
    `RoundedButton` lee `parent.cget("bg")` para mezclar sus esquinas.

    Se empaqueta con `pady=(4, 8)` y `padx=20` para dar el aire
    consistente al contenedor; el caller solo agrega widgets adentro.
    """
    panel = tk.Frame(
        parent,
        bg=branding.ISA_BLANCO,
        highlightthickness=1,
        highlightbackground=branding.ISA_GRIS_BORDE,
        bd=0,
    )
    panel.pack(pady=(4, 8), padx=20)
    return panel


def _crear_header_form(
    root: tk.Tk,
    frame: tk.Frame,
    parent_frame: tk.Frame,
    titulo: str,
    subtitulo: str | None = None,
) -> tk.Button:
    """Construye el encabezado consistente de cualquier sub-formulario:
    botón "← Atrás" arriba-izquierda + logo centrado + título + subtítulo.

    Devuelve el botón Atrás para que el caller pueda referenciarlo (p. ej.
    para deshabilitarlo durante un worker). El botón ya tiene cableado el
    comando que destruye `frame` y re-empaca `parent_frame`.
    """
    btn_atras = branding.RoundedButton(
        frame, text="← Atrás", style="tertiary",
        font=("Helvetica", 9), padx=8, pady=2,
    )
    btn_atras.pack(anchor="w", padx=10, pady=(10, 0))

    def volver() -> None:
        frame.destroy()
        parent_frame.pack(fill="both", expand=True)

    btn_atras.config(command=volver)

    # Logo (reusa la referencia del root para no recargar la imagen).
    if getattr(root, "_logo_ref", None) is not None:
        tk.Label(
            frame, image=root._logo_ref, bg=branding.ISA_FONDO
        ).pack(pady=(4, 6))

    tk.Label(
        frame,
        text=titulo,
        font=("Helvetica", 14, "bold"),
        fg=branding.ISA_AZUL,
        bg=branding.ISA_FONDO,
    ).pack(pady=(4, 4))

    if subtitulo is not None:
        tk.Label(
            frame,
            text=subtitulo,
            font=("Helvetica", 10),
            fg=branding.ISA_GRIS,
            bg=branding.ISA_FONDO,
        ).pack(pady=(0, 12))

    return btn_atras


def abrir_activos_fijos(root: tk.Tk, frame_menu: tk.Frame) -> tk.Frame:
    """Sub-formulario "Activos Fijos" — accesible desde el menú principal.

    Contiene dos botones:
      - "Extraer información en txt": misma lógica que el botón homónimo
        anterior (función `extraer_lsmw_a_txt`).
      - "Creación de Activo": misma lógica que el viejo "Subir a SAP"
        (función `subir_a_sap`), incluyendo el polling que lo habilita
        sólo cuando hay un LSMW_*.txt en salida/.

    El polling se inicia al entrar a la vista y se cancela al salir
    (back button) para no dejar callbacks programados sobre widgets
    destruidos.

    Devuelve el `Frame` creado para que los tests puedan inspeccionar
    los widgets vía atributos.
    """
    frame_menu.pack_forget()

    frame_activos = tk.Frame(root, bg=branding.ISA_FONDO)
    btn_atras = _crear_header_form(
        root, frame_activos, frame_menu,
        titulo="Activos Fijos",
        subtitulo="Extracción de información y creación de activos en SAP",
    )

    status_var = tk.StringVar(value="")

    # Los 4 botones de acción + el status viven dentro de un panel-card
    # blanco con borde gris — mismo lenguaje visual que las cards del
    # menú principal. Ancho fijo del botón: 260 px.
    panel = _crear_panel_card(frame_activos)
    _ANCHO_BOTON_ACTIVOS = 260

    btn_extraer = branding.RoundedButton(
        panel,
        text="Extraer información en txt",
        style="primary",
        padx=18, pady=8, width=_ANCHO_BOTON_ACTIVOS,
        command=lambda: extraer_lsmw_a_txt(status_var),
    )
    btn_extraer.pack(pady=(14, 8), padx=18)

    btn_creacion = branding.RoundedButton(
        panel,
        text="Creación de Activo",
        style="primary",
        padx=18, pady=8, width=_ANCHO_BOTON_ACTIVOS,
    )
    btn_creacion.config(
        command=lambda: subir_a_sap(root, status_var, btn_creacion),
        state="disabled",
    )
    btn_creacion.pack(pady=(0, 8), padx=18)

    btn_extraer_creados = branding.RoundedButton(
        panel,
        text="Extraer Activos Creados",
        style="primary",
        padx=18, pady=8, width=_ANCHO_BOTON_ACTIVOS,
        command=lambda: abrir_extraer_creados(root, frame_activos),
    )
    btn_extraer_creados.pack(pady=(0, 8), padx=18)

    btn_subir_anexos = branding.RoundedButton(
        panel,
        text="Subir Anexos",
        style="primary",
        padx=18, pady=8, width=_ANCHO_BOTON_ACTIVOS,
        command=lambda: abrir_subir_anexos(root, frame_activos),
    )
    btn_subir_anexos.pack(pady=(0, 14), padx=18)

    tk.Label(
        panel,
        textvariable=status_var,
        font=("Helvetica", 9),
        fg=branding.ISA_VERDE_OK,
        bg=branding.ISA_BLANCO,
        wraplength=400,
    ).pack(pady=(0, 14), padx=18)

    # Polling scoped al frame: se inicia ahora, se cancela cuando el
    # frame se destruye (vía el binding <Destroy>). Esto evita callbacks
    # programados sobre un Button ya destruido cuando el usuario regresa
    # al menú.
    polling_id_holder: list[str | None] = [None]

    def poll_creacion_activo() -> None:
        if not btn_creacion.winfo_exists():
            return
        _refrescar_estado_boton_subir(btn_creacion)
        polling_id_holder[0] = root.after(
            _POLL_INTERVAL_MS, poll_creacion_activo
        )

    def on_frame_destroy(_event):
        # Cancelar polling para no dejar callbacks sueltos.
        if polling_id_holder[0] is not None:
            try:
                root.after_cancel(polling_id_holder[0])
            except Exception:
                pass

    frame_activos.bind("<Destroy>", on_frame_destroy)

    # Estado inicial + arrancar polling.
    _refrescar_estado_boton_subir(btn_creacion)
    polling_id_holder[0] = root.after(
        _POLL_INTERVAL_MS, poll_creacion_activo
    )

    frame_activos.pack(fill="both", expand=True)

    # Exponer atributos clave para los tests.
    frame_activos.status_var = status_var
    frame_activos.btn_extraer = btn_extraer
    frame_activos.btn_creacion = btn_creacion
    frame_activos.btn_extraer_creados = btn_extraer_creados
    frame_activos.btn_subir_anexos = btn_subir_anexos
    frame_activos.btn_atras = btn_atras

    return frame_activos


def abrir_extraer_creados(root: tk.Tk, frame_activos: tk.Frame) -> tk.Frame:
    """Sub-formulario "Extraer Activos Creados" — accesible desde
    Activos Fijos. Consulta los activos creados por un usuario SAP.

    Form: campo "Usuario SAP" + botón "Ejecutar".

    NOTA: la lógica del botón Ejecutar aún no está implementada. Por
    ahora muestra un messagebox "En desarrollo" para que el usuario
    sepa que es un placeholder intencional (no un bug).
    """
    frame_activos.pack_forget()

    frame_extraer = tk.Frame(root, bg=branding.ISA_FONDO)
    btn_atras = _crear_header_form(
        root, frame_extraer, frame_activos,
        titulo="Extraer Activos Creados",
        subtitulo="Consulta de activos creados por usuario SAP",
    )

    # Form (input + botón) dentro de un panel-card blanco con borde
    # gris — mismo lenguaje visual que las cards del menú principal.
    panel = _crear_panel_card(frame_extraer)

    form = tk.Frame(panel, bg=branding.ISA_BLANCO)
    form.pack(pady=(16, 12), padx=18)

    tk.Label(
        form, text="Usuario SAP:", anchor="e", width=12,
        bg=branding.ISA_BLANCO, fg=branding.ISA_AZUL,
        font=("Helvetica", 10),
    ).grid(row=0, column=0, padx=4, pady=8, sticky="e")

    usuario_var = tk.StringVar()
    usuario_entry = tk.Entry(
        form, textvariable=usuario_var, width=22,
        font=("Helvetica", 11),
    )
    usuario_entry.grid(row=0, column=1, padx=4, pady=8, sticky="w")

    btn_ejecutar = branding.RoundedButton(
        panel,
        text="Ejecutar",
        style="primary",
        padx=18, pady=8, width=170,
    )
    btn_ejecutar.config(
        command=lambda: _extraer_activos_creados_handler(
            root, usuario_var.get(), btn_ejecutar, btn_atras,
        )
    )
    btn_ejecutar.pack(pady=(0, 16), padx=18)

    frame_extraer.pack(fill="both", expand=True)

    # Exponer atributos clave para los tests.
    frame_extraer.usuario_var = usuario_var
    frame_extraer.usuario_entry = usuario_entry
    frame_extraer.btn_ejecutar = btn_ejecutar
    frame_extraer.btn_atras = btn_atras

    return frame_extraer


def abrir_subir_anexos(root: tk.Tk, frame_activos: tk.Frame) -> tk.Frame:
    """Sub-formulario "Subir Anexos" — accesible desde Activos Fijos.

    Permite seleccionar 1+ archivos del sistema de archivos y subirlos
    como adjuntos a cada activo fijo de la hoja `Activos Fijos` del
    último `ActivosCreados_*.xlsx` en `salida/` (vía SAP AS02 + GOS).

    Form:
      - `Sociedad` (Combobox readonly, mismas opciones que Control SOX).
      - Botón `Seleccionar archivos` (abre `filedialog.askopenfilenames`).
      - Listbox con los archivos seleccionados (botón `Quitar` los borra).
      - Botón `Subir Anexos a SAP` (cableado a `_subir_anexos_handler`).
      - Status label con progreso durante el worker.
    """
    # Importar VALID_SOCIEDADES lazy para no romper en macOS dev (no
    # tiene pywin32 pero sí openpyxl, así que sox_report sí importa OK).
    from sox_report import VALID_SOCIEDADES

    frame_activos.pack_forget()

    frame_anexos = tk.Frame(root, bg=branding.ISA_FONDO)
    btn_atras = _crear_header_form(
        root, frame_anexos, frame_activos,
        titulo="Subir Anexos",
        subtitulo="Adjunta archivos a los activos creados en SAP",
    )

    # Todo el contenido interactivo (combo + botones + listbox + botón
    # principal + status) vive dentro de un panel-card blanco con borde
    # gris — mismo lenguaje visual que las cards del menú.
    panel = _crear_panel_card(frame_anexos)

    # --- Sociedad ---
    form = tk.Frame(panel, bg=branding.ISA_BLANCO)
    form.pack(pady=(14, 4), padx=18)

    tk.Label(
        form, text="Sociedad:", anchor="e", width=10,
        bg=branding.ISA_BLANCO, fg=branding.ISA_AZUL,
        font=("Helvetica", 10),
    ).grid(row=0, column=0, padx=4, pady=6, sticky="e")

    sociedad_var = tk.StringVar()
    sociedad_combo = ttk.Combobox(
        form,
        textvariable=sociedad_var,
        values=list(VALID_SOCIEDADES),
        state="readonly",
        width=14,
    )
    sociedad_combo.grid(row=0, column=1, padx=4, pady=6, sticky="w")

    # --- Sección "Lista de activos" (fuente OPCIONAL desde .xlsx) ---
    # Si el usuario carga un .xlsx válido (Activo Fijo, Subnúmero), esos
    # activos reemplazan a los de "Extraer Activos Creados". La lista se
    # valida y se parsea AL SELECCIONAR (no al subir); acá se guarda ya
    # resuelta en `estado_usuario`. Va en su propia sección con título para
    # no confundirla con el selector de anexos (más abajo).
    estado_usuario: dict = {"activos": None, "nombre": None}

    sec_activos = tk.Frame(panel, bg=branding.ISA_BLANCO)
    sec_activos.pack(pady=(6, 0), padx=18)

    tk.Label(
        sec_activos, text="Lista de activos (opcional)",
        bg=branding.ISA_BLANCO, fg=branding.ISA_AZUL,
        font=("Helvetica", 10, "bold"),
    ).pack(pady=(0, 3))

    activos_btns = tk.Frame(sec_activos, bg=branding.ISA_BLANCO)
    activos_btns.pack()

    btn_sel_activos = branding.RoundedButton(
        activos_btns, text="Cargar .xlsx", style="tertiary",
        padx=12, pady=4, font=("Helvetica", 10),
    )
    btn_sel_activos.grid(row=0, column=0, padx=4)

    btn_quitar_activos = branding.RoundedButton(
        activos_btns, text="Quitar", style="tertiary",
        padx=12, pady=4, font=("Helvetica", 10),
    )
    btn_quitar_activos.grid(row=0, column=1, padx=4)

    archivo_activos_status = tk.Label(
        sec_activos, text="Ninguno — se usarán los activos extraídos",
        bg=branding.ISA_BLANCO, fg=branding.ISA_GRIS,
        font=("Helvetica", 9),
    )
    archivo_activos_status.pack(pady=(3, 0))

    def _seleccionar_archivo_activos() -> None:
        p_str = filedialog.askopenfilename(
            title="Seleccionar archivo .xlsx de activos existentes",
            filetypes=[("Excel .xlsx", "*.xlsx")],
        )
        if not p_str:
            return
        p = Path(p_str)
        try:
            from subir_anexos import validar_y_leer_activos_usuario
        except ImportError as exc:
            messagebox.showerror(
                "Error de import",
                f"No se pudo importar subir_anexos:\n{exc}",
            )
            return
        try:
            activos = validar_y_leer_activos_usuario(p)
        except ValueError as exc:
            messagebox.showwarning("Archivo de activos inválido", str(exc))
            return
        estado_usuario["activos"] = activos
        estado_usuario["nombre"] = p.name
        archivo_activos_status.config(
            text=f"{p.name} — {len(activos)} activo(s)",
            fg=branding.ISA_VERDE_OK,
        )

    def _quitar_archivo_activos() -> None:
        estado_usuario["activos"] = None
        estado_usuario["nombre"] = None
        archivo_activos_status.config(
            text="Ninguno — se usarán los activos extraídos",
            fg=branding.ISA_GRIS,
        )

    btn_sel_activos.config(command=_seleccionar_archivo_activos)
    btn_quitar_activos.config(command=_quitar_archivo_activos)

    # --- Sección "Anexos a subir" (los archivos a adjuntar) ---
    archivos_seleccionados: list[Path] = []

    tk.Label(
        panel, text="Anexos a subir",
        bg=branding.ISA_BLANCO, fg=branding.ISA_AZUL,
        font=("Helvetica", 10, "bold"),
    ).pack(pady=(12, 3), padx=18)

    archivos_frame = tk.Frame(panel, bg=branding.ISA_BLANCO)
    archivos_frame.pack(pady=(0, 0), padx=18)

    btn_seleccionar = branding.RoundedButton(
        archivos_frame,
        text="Seleccionar archivos",
        style="tertiary",
        padx=12, pady=4,
        font=("Helvetica", 10),
    )
    btn_seleccionar.grid(row=0, column=0, padx=4, sticky="w")

    btn_quitar = branding.RoundedButton(
        archivos_frame,
        text="Quitar seleccionado",
        style="tertiary",
        padx=12, pady=4,
        font=("Helvetica", 10),
    )
    btn_quitar.grid(row=0, column=1, padx=4, sticky="w")

    archivos_listbox = tk.Listbox(
        panel, height=5, width=58,
        font=("Helvetica", 9),
        selectmode="single",
        highlightthickness=0, bd=1, relief="solid",
    )
    archivos_listbox.pack(pady=(8, 0), padx=18)

    def _refrescar_listbox() -> None:
        archivos_listbox.delete(0, "end")
        for p in archivos_seleccionados:
            archivos_listbox.insert("end", p.name)

    def _seleccionar_archivos() -> None:
        paths = filedialog.askopenfilenames(
            title="Seleccionar archivos a adjuntar",
        )
        for p_str in paths:
            p = Path(p_str)
            if p not in archivos_seleccionados:
                archivos_seleccionados.append(p)
        _refrescar_listbox()

    def _quitar_archivo() -> None:
        sel = archivos_listbox.curselection()
        if not sel:
            return
        idx = sel[0]
        if 0 <= idx < len(archivos_seleccionados):
            archivos_seleccionados.pop(idx)
            _refrescar_listbox()

    btn_seleccionar.config(command=_seleccionar_archivos)
    btn_quitar.config(command=_quitar_archivo)

    # --- Status + botón Subir ---
    status_var = tk.StringVar(value="")

    btn_subir = branding.RoundedButton(
        panel,
        text="Subir Anexos a SAP",
        style="primary",
        padx=18, pady=8, width=235,
    )
    btn_subir.config(
        command=lambda: _subir_anexos_handler(
            root,
            sociedad_var.get(),
            list(archivos_seleccionados),
            estado_usuario["activos"],
            estado_usuario["nombre"],
            status_var,
            btn_subir,
            btn_atras,
        )
    )
    btn_subir.pack(pady=(14, 6), padx=18)

    # Status label legible: 10pt bold + padding inferior amplio para
    # que no quede pegado al borde de la ventana (donde se renderiza
    # recortado / borroso).
    tk.Label(
        panel,
        textvariable=status_var,
        font=("Helvetica", 10, "bold"),
        fg=branding.ISA_VERDE_OK,
        bg=branding.ISA_BLANCO,
        wraplength=520,
    ).pack(pady=(10, 14), padx=18)

    frame_anexos.pack(fill="both", expand=True)

    # Exponer atributos clave para tests.
    frame_anexos.sociedad_var = sociedad_var
    frame_anexos.sociedad_combo = sociedad_combo
    frame_anexos.archivos_listbox = archivos_listbox
    frame_anexos.archivos_seleccionados = archivos_seleccionados
    frame_anexos.btn_seleccionar = btn_seleccionar
    frame_anexos.btn_quitar = btn_quitar
    frame_anexos.btn_subir = btn_subir
    frame_anexos.btn_atras = btn_atras
    frame_anexos.status_var = status_var

    return frame_anexos


def abrir_sox_menu(root: tk.Tk, frame_menu: tk.Frame) -> tk.Frame:
    """Sub-formulario intermedio "Control SOX" — accesible desde el menú
    principal. Muestra un único botón "HUB.PPE.01 Creación de Activos
    Fijos" que abre el formulario con los parámetros (sociedad + fechas).

    Diseñado como contenedor de futuras opciones HUB.PPE.XX que se vayan
    agregando: hoy es un único botón pero la estructura permite añadir
    más sin mover la lógica del menú principal.
    """
    frame_menu.pack_forget()

    frame_sox_menu = tk.Frame(root, bg=branding.ISA_FONDO)
    btn_atras = _crear_header_form(
        root, frame_sox_menu, frame_menu,
        titulo="Control SOX",
        subtitulo="Procesos de control y auditoría",
    )

    # Botón dentro de panel-card blanco — mismo lenguaje visual que
    # el menú principal. A futuro se pueden añadir más botones HUB.PPE
    # acá dentro sin tocar el resto.
    panel = _crear_panel_card(frame_sox_menu)

    btn_hub_ppe_01 = branding.RoundedButton(
        panel,
        text="HUB.PPE.01 Creación de Activos Fijos",
        style="primary",
        padx=18, pady=8, width=340,
        command=lambda: control_sox(root, frame_sox_menu),
    )
    btn_hub_ppe_01.pack(pady=(18, 18), padx=18)

    frame_sox_menu.pack(fill="both", expand=True)

    # Exponer atributos clave para los tests.
    frame_sox_menu.btn_hub_ppe_01 = btn_hub_ppe_01
    frame_sox_menu.btn_atras = btn_atras

    return frame_sox_menu


def control_sox(root: tk.Tk, frame_menu: tk.Frame) -> tk.Frame:
    """Reemplaza la vista del menú principal por el formulario Control SOX
    en la misma ventana (sin abrir un Toplevel).

    Oculta `frame_menu` con `pack_forget` y muestra un nuevo `frame_sox` con
    un botón "← Atrás" arriba que destruye el form SOX y re-muestra el
    menú al ser presionado.

    Devuelve el `Frame` creado para que los tests puedan inspeccionar los
    widgets (las StringVars, los DateEntry, los botones).
    """
    from sox_report import VALID_SOCIEDADES

    # Ocultar el menú — preservamos su estado (status_var, polling del
    # botón Subir, etc.) para no perder progreso si el usuario vuelve.
    frame_menu.pack_forget()

    frame_sox = tk.Frame(root, bg=branding.ISA_FONDO)

    # --- Botón Atrás (esquina superior izquierda, estilo discreto) ---
    btn_atras = branding.RoundedButton(
        frame_sox,
        text="← Atrás",
        style="tertiary",
        font=("Helvetica", 9),
        padx=8,
        pady=2,
    )
    btn_atras.pack(anchor="w", padx=10, pady=(10, 0))

    def volver_al_menu() -> None:
        frame_sox.destroy()
        frame_menu.pack(fill="both", expand=True)

    btn_atras.config(command=volver_al_menu)

    # --- Logo arriba del título (reusa la referencia del root) ---
    if getattr(root, "_logo_ref", None) is not None:
        tk.Label(
            frame_sox, image=root._logo_ref, bg=branding.ISA_FONDO
        ).pack(pady=(4, 6))

    # --- Título ---
    tk.Label(
        frame_sox,
        text="Control SOX",
        font=("Helvetica", 13, "bold"),
        fg=branding.ISA_AZUL,
        bg=branding.ISA_FONDO,
    ).pack(pady=(4, 4))
    tk.Label(
        frame_sox,
        text="Genera el Reporte SOX con los parámetros indicados",
        font=("Helvetica", 10),
        fg=branding.ISA_GRIS,
        bg=branding.ISA_FONDO,
    ).pack(pady=(0, 12))

    # Todo el form + botón + status dentro de un panel-card blanco con
    # borde gris — mismo lenguaje visual que el menú principal.
    panel = _crear_panel_card(frame_sox)

    form = tk.Frame(panel, bg=branding.ISA_BLANCO)
    form.pack(pady=(16, 12), padx=18)

    # --- Sociedades (checkboxes multiselect) ---
    # Multiselect con checkboxes en grilla de 2 columnas: el usuario marca 1
    # o varias sociedades y al Generar se corre el flujo SOX una vez por cada
    # una (un reporte por sociedad, no consolidado). Los checkboxes son más
    # amigables que un Listbox con barras de selección.
    tk.Label(
        form, text="Sociedades:", anchor="e", width=10,
        bg=branding.ISA_BLANCO, fg=branding.ISA_AZUL,
    ).grid(row=0, column=0, padx=4, pady=6, sticky="ne")

    sociedad_box = tk.Frame(form, bg=branding.ISA_BLANCO)
    sociedad_box.grid(row=0, column=1, padx=4, pady=6, sticky="w")

    checks_frame = tk.Frame(sociedad_box, bg=branding.ISA_BLANCO)
    checks_frame.pack(anchor="w")

    _NCOLS_SOC = 2
    soc_vars: dict[str, tk.BooleanVar] = {}
    for i, soc in enumerate(VALID_SOCIEDADES):
        var = tk.BooleanVar(value=False)
        soc_vars[soc] = var
        cb = tk.Checkbutton(
            checks_frame, text=soc, variable=var,
            bg=branding.ISA_BLANCO, fg=branding.ISA_AZUL,
            activebackground=branding.ISA_BLANCO,
            activeforeground=branding.ISA_AZUL,
            selectcolor=branding.ISA_BLANCO,
            font=("Helvetica", 10), anchor="w", width=7,
            highlightthickness=0, bd=0, cursor="hand2", padx=2,
        )
        r, c = divmod(i, _NCOLS_SOC)
        cb.grid(row=r, column=c, sticky="w", padx=(0, 10), pady=1)

    tk.Label(
        sociedad_box,
        text="Selecciona 1 o más sociedades para generar el Reporte SOX",
        fg=branding.ISA_GRIS, bg=branding.ISA_BLANCO,
        font=("Helvetica", 8), justify="left", wraplength=240,
    ).pack(anchor="w", pady=(4, 0))

    def _sociedades_seleccionadas() -> list[str]:
        # Preserva el orden de VALID_SOCIEDADES.
        return [soc for soc, var in soc_vars.items() if var.get()]

    # --- Fechas con calendario emergente (DateEntry de tkcalendar) ---
    # DateEntry abre un popup de calendario al hacer clic en la flecha.
    #
    # IMPORTANTE: NO se usan `validate="key"` + `validatecommand` aquí.
    # La cascada de Tk dispara la `validatecommand` cuando tkcalendar
    # actualiza el textvariable via `set_date()` desde el popup. Esa
    # re-entrancia rompía el flujo `_select` del calendario: a veces el
    # popup quedaba con `_date = None` y los clicks dejaban de surtir
    # efecto hasta reiniciar la app.
    #
    # tkcalendar ya valida el formato `dd.mm.yyyy` al perder foco; lo que
    # el usuario tipea queda saneado en focusout. El submit final pasa
    # además por `validar_fecha` + `validar_rango_fechas` antes de llegar
    # a SAP, así que no perdemos garantías reales — sólo el filtrado
    # cosmético per-keystroke de letras.
    fecha_hoy = datetime.now()

    tk.Label(
        form, text="Desde:", anchor="e", width=10,
        bg=branding.ISA_BLANCO, fg=branding.ISA_AZUL,
    ).grid(row=1, column=0, padx=4, pady=6, sticky="e")
    desde_var = tk.StringVar()
    desde_entry = DateEntry(
        form,
        textvariable=desde_var,
        date_pattern="dd.mm.yyyy",
        width=14,
        background="#1a73e8",
        foreground="white",
        borderwidth=2,
        year=fecha_hoy.year,
        month=fecha_hoy.month,
        day=fecha_hoy.day,
    )
    desde_entry.grid(row=1, column=1, padx=4, pady=6, sticky="w")
    tk.Label(
        form, text="(dd.mm.aaaa)",
        fg=branding.ISA_GRIS_CLARO, bg=branding.ISA_BLANCO,
    ).grid(row=1, column=2, padx=4)

    tk.Label(
        form, text="Hasta:", anchor="e", width=10,
        bg=branding.ISA_BLANCO, fg=branding.ISA_AZUL,
    ).grid(row=2, column=0, padx=4, pady=6, sticky="e")
    hasta_var = tk.StringVar()
    hasta_entry = DateEntry(
        form,
        textvariable=hasta_var,
        date_pattern="dd.mm.yyyy",
        width=14,
        background="#1a73e8",
        foreground="white",
        borderwidth=2,
        year=fecha_hoy.year,
        month=fecha_hoy.month,
        day=fecha_hoy.day,
    )
    hasta_entry.grid(row=2, column=1, padx=4, pady=6, sticky="w")
    tk.Label(
        form, text="(dd.mm.aaaa)",
        fg=branding.ISA_GRIS_CLARO, bg=branding.ISA_BLANCO,
    ).grid(row=2, column=2, padx=4)

    status_var = tk.StringVar()

    btn_generar = branding.RoundedButton(
        panel,
        text="Generar Reporte SOX",
        style="primary",
        padx=18,
        pady=8,
    )
    btn_generar.config(
        command=lambda: _generar_reporte_sox_handler(
            root,
            _sociedades_seleccionadas(),
            desde_var.get(),
            hasta_var.get(),
            status_var,
            btn_generar,
            btn_atras,
        )
    )
    btn_generar.pack(pady=(4, 14), padx=18)

    tk.Label(
        panel,
        textvariable=status_var,
        font=("Helvetica", 9),
        fg=branding.ISA_VERDE_OK,
        bg=branding.ISA_BLANCO,
        wraplength=420,
    ).pack(pady=(0, 14), padx=18)

    frame_sox.pack(fill="both", expand=True)

    # Exponer widgets clave en el frame para que los tests puedan inspeccionar.
    frame_sox.desde_var = desde_var
    frame_sox.hasta_var = hasta_var
    frame_sox.status_var = status_var
    frame_sox.soc_vars = soc_vars
    frame_sox.sociedades_seleccionadas = _sociedades_seleccionadas
    frame_sox.desde_entry = desde_entry
    frame_sox.hasta_entry = hasta_entry
    frame_sox.btn_generar = btn_generar
    frame_sox.btn_atras = btn_atras

    return frame_sox


def _test_conexion_sap_handler() -> None:
    """Handler del botón "Test conexión SAP". Llama a
    `diagnosticar_conexion_sap` y muestra el resultado en un messagebox.
    """
    try:
        from sap_upload import diagnosticar_conexion_sap
    except ImportError as exc:
        messagebox.showerror(
            "Error de import",
            f"No se pudo importar sap_upload:\n{exc}",
        )
        return

    try:
        ok, mensaje = diagnosticar_conexion_sap()
    except Exception as exc:
        _show_unexpected_error("Error en test de conexión SAP", exc)
        return

    _log(f"Test conexión SAP → ok={ok}")
    print(mensaje, flush=True)
    if ok:
        messagebox.showinfo("Test conexión SAP — OK", mensaje)
    else:
        messagebox.showwarning("Test conexión SAP — Problema", mensaje)


def main() -> None:
    # En modo bundled (.exe), si es el primer arranque y la carpeta
    # `entrada/` al lado del ejecutable aún no tiene ningún `.xlsm`, lo
    # extraemos desde el bundle como factory default (`Formato_Dinamico.xlsm`).
    # Después de esto el archivo es externo y editable. En dev mode copia
    # a `<repo>/entrada/` desde `<repo>/resources/`.
    try:
        ruta_formato, recien_creado = asegurar_formato_dinamico()
        if recien_creado:
            _log(
                "Primer arranque: se creó "
                f"{ruta_formato.name} en {ruta_formato.parent}. "
                "Puedes editarlo libremente."
            )
    except Exception as exc:
        # No bloquea el arranque; el handler de "Extraer" reportará
        # el error si el archivo sigue ausente cuando lo necesite.
        _log(f"asegurar_formato_dinamico falló (no crítico): {exc}")

    root = tk.Tk()
    _install_tk_exception_handler(root)
    root.title("Gestión de Activos Fijos")
    # Geometría 620x605 (igual que antes del refresh visual). Las 3
    # cards del menú principal están dimensionadas para caber acá sin
    # forzar agrandar la ventana — ancho ~190 px cada una.
    root.geometry("620x605")
    root.resizable(False, False)
    root.configure(bg=branding.ISA_FONDO)

    # Footer de copyright: se empaca PRIMERO sobre root con side="bottom"
    # para que claimee el strip inferior. Los frames de las sub-vistas
    # se montan después con fill="both"+expand=True y ocupan el espacio
    # por encima del footer, por lo que el copyright persiste visible
    # en todas las pantallas sin tener que llamarlo en cada `abrir_*`.
    _crear_footer_copyright(root)

    # Todos los widgets del menú principal viven dentro de `frame_menu`
    # para que las sub-vistas (Activos Fijos, Control SOX) puedan
    # ocultarlo con pack_forget y mostrar sus propios frames en la misma
    # ventana, preservando el estado del menú.
    frame_menu = tk.Frame(root, bg=branding.ISA_FONDO)

    # Logo Hub de ISA arriba. Si el archivo no existe, _logo_ref es None
    # y simplemente no se renderiza la imagen (graceful fallback). La
    # referencia se guarda en `root` para que GC no libere la imagen y
    # también para que `_crear_header_form` la reuse en sub-vistas.
    root._logo_ref = branding.cargar_logo()
    if root._logo_ref is not None:
        tk.Label(
            frame_menu, image=root._logo_ref, bg=branding.ISA_FONDO,
        ).pack(pady=(18, 8))

    tk.Label(
        frame_menu,
        text="Gestión de Activos Fijos",
        font=("Helvetica", 14, "bold"),
        fg=branding.ISA_AZUL,
        bg=branding.ISA_FONDO,
    ).pack(pady=(4, 20))

    # --- 3 cards visuales con layout 2 + 1 ---
    #   Row arriba:   [ Activos Fijos ] [ Control SOX ]
    #   Row abajo:           [ Reportes ]   (centrada)
    #
    # Las 3 cards se fuerzan a width=205 / height=180 con pack_propagate
    # del Frame interno para que queden visualmente simétricas (mismo
    # tamaño exacto) aunque los textos de descripción varíen en líneas.
    # Borde plano 1 px gris (`ISA_GRIS_BORDE`) — Tk no soporta
    # drop-shadow nativo.
    _CARD_W, _CARD_H = 205, 180

    cards_top = tk.Frame(frame_menu, bg=branding.ISA_FONDO)
    cards_top.pack(pady=(12, 0))

    card_activos, btn_card_activos = _crear_card_visual(
        cards_top,
        titulo="Activos Fijos",
        descripcion=(
            "Consulta y gestiona la información\n"
            "de los activos fijos de la compañía.\n"
            "Visualiza detalles, ubicaciones,\n"
            "responsables y más."
        ),
        btn_texto="Acceder  →",
        btn_style="primary",
        width=_CARD_W, height=_CARD_H,
        command=lambda: abrir_activos_fijos(root, frame_menu),
    )
    card_activos.pack(side="left", padx=8)

    card_sox, btn_card_sox = _crear_card_visual(
        cards_top,
        titulo="Control SOX",
        descripcion=(
            "Revisa y da seguimiento a los\n"
            "controles SOX, evidencia\n"
            "y cumplimiento normativo."
        ),
        btn_texto="Continuar  →",
        btn_style="naranja",
        width=_CARD_W, height=_CARD_H,
        command=lambda: abrir_sox_menu(root, frame_menu),
    )
    card_sox.pack(side="left", padx=8)

    cards_bottom = tk.Frame(frame_menu, bg=branding.ISA_FONDO)
    cards_bottom.pack(pady=(12, 0))

    card_reportes, btn_card_reportes = _crear_card_visual(
        cards_bottom,
        titulo="Reportes",
        descripcion=(
            "Genera y consulta reportes\n"
            "estadísticos y financieros.\n"
            "Obtén información clave para\n"
            "la toma de decisiones."
        ),
        btn_texto="Ver reportes  →",
        btn_style="verde",
        width=_CARD_W, height=_CARD_H,
        disabled=True,
    )
    # Sin `side` ni `padx`: el Frame `cards_bottom` solo contiene esta
    # card, así que pack default la centra horizontalmente.
    card_reportes.pack()
    btn_card_reportes.pack(side="left", padx=10)

    # Botón de diagnóstico "Test conexión SAP": se conserva el código pero
    # NO se hace .pack para ocultarlo de la UI (puede reactivarse en el
    # futuro re-empaquetándolo). El handler `_test_conexion_sap_handler`
    # tampoco se borra.
    btn_test = branding.RoundedButton(
        frame_menu,
        text="Test conexión SAP",
        style="tertiary",
        font=("Helvetica", 9),
        padx=10, pady=2,
        command=_test_conexion_sap_handler,
    )
    # btn_test.pack()  # intencionalmente oculto

    frame_menu.pack(fill="both", expand=True)

    # El splash de PyInstaller (si la app corre como .exe bundled) se
    # cierra justo antes de mainloop para que la transición sea limpia:
    # imagen de carga → ventana principal. En dev mode es no-op.
    root.update_idletasks()
    _cerrar_splash()

    root.mainloop()


if __name__ == "__main__":
    main()
