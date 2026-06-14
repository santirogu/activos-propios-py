"""subir_anexos.py — Carga de archivos como adjuntos a activos fijos en SAP
vía AS02 + GOS (Generic Object Services) + PCATTA_CREA.

Replica `resources/Scriptanexo.vbs` (sin la navegación manual de carpetas
que el usuario hizo durante la grabación — inyectamos `DY_PATH`
directamente en `wnd[3]`). Para cada par `(activo_fijo, subnúmero)` leído
de la hoja `Activos Fijos` del último archivo `ActivosCreados_*.xlsx` en
`salida/`, sube cada uno de los archivos seleccionados como adjunto.

REQUISITOS DE EJECUCIÓN
=======================
Sistema operativo: Windows con SAP GUI for Windows abierto y sesión
iniciada. Mismos requisitos que `sap_upload.py` y `sox_report.py`.

USO
===
    python src/subir_anexos.py SOCIEDAD ARCHIVO1 [ARCHIVO2 ...]

    Ejemplo:
        python src/subir_anexos.py ISA C:\docs\contrato.pdf C:\docs\foto.jpg

También se invoca desde la GUI vía el botón "Subir Anexos" dentro de la
vista "Activos Fijos".
"""

from __future__ import annotations

import sys
import time
from pathlib import Path
from typing import Callable

from openpyxl import load_workbook

# Re-uso de la lista de sociedades válidas y la validación del módulo SOX
# (misma fuente de verdad — ambos módulos consumen la misma lista).
from sox_report import VALID_SOCIEDADES, validar_sociedad
from extraer_activos_creados import ACTIVOS_FIJOS_SHEET_NAME

from paths import SALIDA_DIR  # noqa: E402  (paths centralizado dev/bundled)

# ---------------------------------------------------------------------------
# CONFIGURACIÓN
# ---------------------------------------------------------------------------

# T-code de Cambio Activo Fijo (Asset Master Modify). El prefijo "/n"
# fuerza a SAP a iniciar la transacción FRESCA desde cualquier estado
# previo. Sin él, si la iteración anterior dejó SAP en una pantalla
# detalle de AS02 (porque falló a media ejecución), el `okcd = "as02"`
# crudo no resetea — termina sin pantalla de selección y el siguiente
# findById(ANLA-ANLN1) falla con "control not found".
T_CODE_AS02 = "/nas02"


# Campos del header de AS02 (sociedad + activo + subnúmero).
CAMPO_ANLN1 = "wnd[0]/usr/ctxtANLA-ANLN1"   # activo
CAMPO_ANLN2 = "wnd[0]/usr/ctxtANLA-ANLN2"   # subnúmero
CAMPO_BUKRS = "wnd[0]/usr/ctxtANLA-BUKRS"   # sociedad

# Botón "Generic Object Services" (icono de toolbox) en la barra de
# título de AS02. Se invoca con `pressButton` (NO `pressContextButton`).
SHELL_TITULAR = "wnd[0]/titl/shellcont/shell"
GOS_TOOLBOX = "%GOS_TOOLBOX"

# Tras pressButton("%GOS_TOOLBOX"), SAP despliega una segunda shell con
# la toolbar GOS (botones Crear, Lista, Nota, etc.). Vive en una ruta
# DIFERENTE a la del título (sin `/titl/`).
SHELL_GOS_BAR = "wnd[0]/shellcont/shell"
CREATE_ATTA = "CREATE_ATTA"      # botón "Crear" (abre submenú)
PCATTA_CREA = "PCATTA_CREA"      # item "Crear anexo" del submenú

# Tras `selectContextMenuItem("PCATTA_CREA")`, SAP abre directamente
# `wnd[1]` con el campo DY_PATH editable. NO existe `wnd[2]`, NO hay
# cascada de F4, NO hay DY_FILENAME separado — el path completo va en
# DY_PATH y un único btn[0] confirma y crea el adjunto.
CAMPO_DY_PATH = "wnd[1]/usr/ctxtDY_PATH"

# Botón OK de wnd[1] que confirma el path inyectado y crea el adjunto.
BTN_CONFIRMAR_WND1 = "wnd[1]/tbar[0]/btn[0]"


# ---------------------------------------------------------------------------
# LOGGING
# ---------------------------------------------------------------------------

def _log(mensaje: str) -> None:
    ts = time.strftime("%H:%M:%S")
    print(f"[{ts}] {mensaje}", flush=True)


def _ejecutar(descripcion: str, fn, *args, **kwargs):
    """Wrapper que loguea y re-lanza con contexto si falla."""
    _log(f"  → {descripcion}")
    try:
        return fn(*args, **kwargs)
    except Exception as exc:
        raise RuntimeError(
            f"Falló: {descripcion}\nDetalle técnico SAP: {exc!r}"
        ) from exc


# ---------------------------------------------------------------------------
# CONEXIÓN A SAP (igual al patrón de sap_upload)
# ---------------------------------------------------------------------------

def get_sap_session():
    """Conecta al SAP GUI Scripting Engine y devuelve la primera sesión."""
    try:
        import win32com.client  # type: ignore
    except ImportError as exc:
        raise RuntimeError(
            "Falta la dependencia pywin32. Instalar con: pip install pywin32"
        ) from exc

    try:
        sap_gui_auto = win32com.client.GetObject("SAPGUI")
    except Exception as exc:
        raise RuntimeError(
            "No se pudo conectar a SAP GUI. Verifica:\n"
            "  - SAP GUI for Windows está abierto y con sesión iniciada.\n"
            "  - SAP GUI Scripting habilitado en Options del cliente.\n"
            "  - sapgui/user_scripting = TRUE en el servidor SAP."
        ) from exc

    application = sap_gui_auto.GetScriptingEngine
    if application.Children.Count == 0:
        raise RuntimeError("No hay conexiones SAP activas en este SAP GUI.")
    connection = application.Children(0)
    if connection.Children.Count == 0:
        raise RuntimeError(
            "No hay sesiones activas en la conexión SAP. "
            "Inicia sesión en el sistema SAP antes de correr este script."
        )
    return connection.Children(0)


# ---------------------------------------------------------------------------
# LECTURA DE ACTIVOS DEL EXCEL
# ---------------------------------------------------------------------------

def get_archivo_activos_mas_reciente(salida_dir: Path = SALIDA_DIR) -> Path:
    """Devuelve el archivo `ActivosCreados_*.xlsx` más reciente en
    `salida/` por mtime. Lanza FileNotFoundError si no hay ninguno
    (típicamente porque el usuario aún no corrió "Extraer Activos Creados").
    """
    if not salida_dir.exists():
        raise FileNotFoundError(
            f"No existe la carpeta {salida_dir}. "
            f"Corre primero 'Extraer Activos Creados' para generar el archivo."
        )
    archivos = sorted(
        salida_dir.glob("ActivosCreados_*.xlsx"),
        key=lambda p: p.stat().st_mtime,
    )
    if not archivos:
        raise FileNotFoundError(
            f"No hay archivos ActivosCreados_*.xlsx en {salida_dir}. "
            f"Corre primero 'Extraer Activos Creados' desde la GUI."
        )
    return archivos[-1]


def leer_activos_del_excel(archivo_path: Path) -> list[tuple[int, int]]:
    """Lee la hoja `Activos Fijos` del archivo y devuelve la lista de
    pares `(activo_fijo, subnúmero)` como ints, en el orden de las filas.

    Args:
        archivo_path: ruta al `ActivosCreados_*.xlsx` post-procesado por
            `extraer_activos_creados.procesar_logs` (debe tener la hoja
            `Activos Fijos`).

    Returns:
        Lista de tuplas `(activo, subnúmero)`. Vacía si la hoja existe
        pero no tiene filas de datos.

    Raises:
        FileNotFoundError: si `archivo_path` no existe.
        ValueError: si el workbook no tiene la hoja `Activos Fijos`.
    """
    if not archivo_path.exists():
        raise FileNotFoundError(
            f"No existe el archivo: {archivo_path}"
        )

    wb = load_workbook(archivo_path, data_only=True)
    if ACTIVOS_FIJOS_SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"El archivo {archivo_path.name} no tiene la hoja "
            f"'{ACTIVOS_FIJOS_SHEET_NAME}'. Verifica que sea un "
            f"`ActivosCreados_*.xlsx` post-procesado por "
            f"`procesar_logs`."
        )

    ws = wb[ACTIVOS_FIJOS_SHEET_NAME]
    pares: list[tuple[int, int]] = []
    # Fila 1 = headers; datos desde fila 2.
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) < 2:
            continue
        activo, sub = row[0], row[1]
        if isinstance(activo, int) and isinstance(sub, int):
            pares.append((activo, sub))
    return pares


# ---------------------------------------------------------------------------
# FLUJO SAP — Adjuntar UN archivo a UN activo
# ---------------------------------------------------------------------------

def adjuntar_archivo(
    session, anln1: int, anln2: int, bukrs: str, archivo_path: Path
) -> None:
    """Adjunta `archivo_path` al activo fijo (`anln1`, `anln2`) de la
    sociedad `bukrs` en SAP, vía AS02 + GOS.

    Secuencia 1:1 del recording (líneas 15-29 del `Scriptanexo.vbs`):
      1. okcd = "/nas02" + sendVKey 0 (abre AS02 fresca)
      2. Set ANLN1 (siempre), ANLN2 (sólo si != 0), BUKRS + setFocus +
         caretPosition + sendVKey 0 (carga el activo)
      3. `SHELL_TITULAR.pressButton("%GOS_TOOLBOX")` → despliega la
         toolbar GOS como segunda shell `SHELL_GOS_BAR`
      4. `SHELL_GOS_BAR.pressContextButton("CREATE_ATTA")` → abre
         submenú "Crear"
      5. `SHELL_GOS_BAR.selectContextMenuItem("PCATTA_CREA")` → selecciona
         "Crear anexo" → SAP abre wnd[1] con DY_PATH
      6. Set `wnd[1]/usr/ctxtDY_PATH` = ruta absoluta del archivo +
         setFocus + caretPosition al final
      7. `wnd[1]/tbar[0]/btn[0].press` → confirma y crea el adjunto

    Args:
        session: sesión SAP GUI.
        anln1: número de activo fijo (ej. 8048124).
        anln2: subnúmero (ej. 0, 1, 2...).
        bukrs: código de sociedad (ej. "ISA", "TRAN").
        archivo_path: ruta absoluta del archivo a adjuntar.

    Raises:
        RuntimeError: si cualquier etapa SAP falla (el caller decide si
            soft-fail por activo o hard-fail abortando todo).
    """
    ruta_str = str(archivo_path)
    _log(f"  Adjuntando '{archivo_path.name}' a activo {anln1}-{anln2} ({bukrs})")

    # 1. Maximizar + abrir AS02. Cada línea hace su propio findById
    # para mantener simetría con el .vbs (línea por línea).
    _ejecutar(
        "Maximizar wnd[0]",
        lambda: session.findById("wnd[0]").maximize(),
    )
    _ejecutar(
        f"Escribir T-code '{T_CODE_AS02}' en okcd",
        lambda: setattr(
            session.findById("wnd[0]/tbar[0]/okcd"), "text", T_CODE_AS02
        ),
    )
    _ejecutar(
        "sendVKey 0 después de okcd (abre AS02)",
        lambda: session.findById("wnd[0]").sendVKey(0),
    )

    # 2. Set asset data. ORDEN IMPORTA — match exacto al recording:
    #    ANLN1 → (ANLN2 si != 0) → BUKRS (text + setFocus + caretPosition).
    # ANLN2 se OMITE cuando es 0 (default de SAP, como el recording).
    # Setearlo cuando es 0 disparaba el auto-tab y descolocaba el focus
    # que poníamos después en BUKRS.
    _ejecutar(
        f"Asignar ANLN1 = '{anln1}'",
        lambda: setattr(session.findById(CAMPO_ANLN1), "text", str(anln1)),
    )

    if anln2 != 0:
        _ejecutar(
            f"Asignar ANLN2 = '{anln2}' (subnúmero != 0)",
            lambda: setattr(session.findById(CAMPO_ANLN2), "text", str(anln2)),
        )
    else:
        _log("  → ANLN2 omitido (subnúmero=0, default de SAP)")

    _ejecutar(
        f"Asignar BUKRS = '{bukrs}'",
        lambda: setattr(session.findById(CAMPO_BUKRS), "text", bukrs),
    )
    _ejecutar(
        "Foco en BUKRS",
        lambda: session.findById(CAMPO_BUKRS).setFocus(),
    )
    _ejecutar(
        f"Cursor al final de BUKRS (len={len(bukrs)})",
        lambda: setattr(
            session.findById(CAMPO_BUKRS), "caretPosition", len(bukrs)
        ),
    )
    _ejecutar(
        "sendVKey 0 después de BUKRS (carga el activo)",
        lambda: session.findById("wnd[0]").sendVKey(0),
    )

    # 3. Acceso al menú GOS y selección "Crear → Crear anexo".
    # Match exacto al recording actualizado (líneas 23-25 del .vbs):
    #   a) pressButton "%GOS_TOOLBOX" en wnd[0]/titl/shellcont/shell
    #      → despliega la toolbar GOS como SEGUNDA shell
    #        (wnd[0]/shellcont/shell, sin /titl/).
    #   b) pressContextButton "CREATE_ATTA" en esa segunda shell
    #      → abre el submenú del botón "Crear".
    #   c) selectContextMenuItem "PCATTA_CREA" en la misma shell
    #      → selecciona "Crear anexo".
    _ejecutar(
        f"Press 'Toolbox GOS' (pressButton {GOS_TOOLBOX})",
        lambda: session.findById(SHELL_TITULAR).pressButton(GOS_TOOLBOX),
    )
    _ejecutar(
        f"Abrir submenú 'Crear' (pressContextButton {CREATE_ATTA})",
        lambda: session.findById(SHELL_GOS_BAR).pressContextButton(CREATE_ATTA),
    )
    _ejecutar(
        f"Seleccionar 'Crear anexo' (selectContextMenuItem {PCATTA_CREA})",
        lambda: session.findById(SHELL_GOS_BAR).selectContextMenuItem(PCATTA_CREA),
    )

    # 4. Tras `selectContextMenuItem("PCATTA_CREA")`, SAP abre wnd[1]
    # directamente con el campo DY_PATH. Inyectamos la ruta absoluta,
    # ponemos foco + caret al final, y confirmamos con btn[0].
    # Líneas 26-29 del recording actualizado.
    _ejecutar(
        f"Asignar wnd[1]/DY_PATH = '{ruta_str}'",
        lambda: setattr(session.findById(CAMPO_DY_PATH), "text", ruta_str),
    )
    _ejecutar(
        "Foco en wnd[1]/DY_PATH",
        lambda: session.findById(CAMPO_DY_PATH).setFocus(),
    )
    _ejecutar(
        f"Cursor al final del path (len={len(ruta_str)})",
        lambda: setattr(
            session.findById(CAMPO_DY_PATH), "caretPosition", len(ruta_str)
        ),
    )
    _ejecutar(
        f"Pulsar OK en wnd[1] ({BTN_CONFIRMAR_WND1}) → crea el adjunto",
        lambda: session.findById(BTN_CONFIRMAR_WND1).press(),
    )


# ---------------------------------------------------------------------------
# ORQUESTADOR
# ---------------------------------------------------------------------------

def subir_anexos(
    session,
    sociedad: str,
    archivos: list[Path],
    archivo_activos: Path | None = None,
    progress_callback: Callable[[int, int, str], None] | None = None,
) -> dict:
    """Adjunta cada archivo de `archivos` a cada par `(activo, subnúmero)`
    leído de la hoja `Activos Fijos` del archivo `archivo_activos`.

    **Soft-fail por iteración**: si una combinación `(activo, archivo)`
    falla en SAP, se loguea y se continúa con la siguiente. El total de
    éxitos/fallos se reporta al final.

    Args:
        session: sesión SAP GUI activa.
        sociedad: código de sociedad (validado contra `VALID_SOCIEDADES`).
        archivos: lista de Paths absolutos a los archivos a subir.
        archivo_activos: ruta al `ActivosCreados_*.xlsx` con la hoja
            `Activos Fijos`. Si None, se usa el más reciente en `salida/`.
        progress_callback: opcional. Si se pasa, se llama con
            `(intento_actual, total_intentos, descripcion)` antes de
            cada attachment — útil para que el handler GUI actualice
            un status label durante el worker.

    Returns:
        Dict con stats:
          - `exitosos`: cuántos `(activo, archivo)` se adjuntaron OK.
          - `fallidos`: cuántos fallaron.
          - `total_intentos`: total = activos × archivos.
          - `detalles_fallos`: lista de tuplas
            `(activo, subnúmero, archivo_path, mensaje_error)`.

    Raises:
        ValueError: si `sociedad` inválida, o si `Activos Fijos` no
            existe o está vacía.
        FileNotFoundError: si no hay archivos `ActivosCreados_*.xlsx`.
    """
    sociedad_norm = validar_sociedad(sociedad)

    if not archivos:
        raise ValueError(
            "Debes seleccionar al menos un archivo a adjuntar."
        )

    if archivo_activos is None:
        archivo_activos = get_archivo_activos_mas_reciente()
    activos = leer_activos_del_excel(archivo_activos)
    if not activos:
        raise ValueError(
            f"El archivo {archivo_activos.name} tiene la hoja "
            f"'{ACTIVOS_FIJOS_SHEET_NAME}' pero sin filas de datos."
        )

    total = len(activos) * len(archivos)
    _log(
        f"=== Subiendo {len(archivos)} archivo(s) a {len(activos)} "
        f"activo(s) — {total} attachments en total ==="
    )

    exitosos = 0
    detalles_fallos: list[tuple[int, int, str, str]] = []
    intento = 0

    for activo, sub in activos:
        for archivo in archivos:
            intento += 1
            desc = f"activo {activo}-{sub}, archivo '{archivo.name}'"
            if progress_callback is not None:
                try:
                    progress_callback(intento, total, desc)
                except Exception:
                    pass

            try:
                adjuntar_archivo(session, activo, sub, sociedad_norm, archivo)
                exitosos += 1
                _log(f"OK ({intento}/{total}): {desc}")
            except Exception as exc:
                detalles_fallos.append(
                    (activo, sub, str(archivo), str(exc))
                )
                _log(f"FALLO ({intento}/{total}): {desc} — {exc!r}")

    _log(
        f"=== Finalizado: {exitosos} OK, {len(detalles_fallos)} fallos ==="
    )

    return {
        "exitosos": exitosos,
        "fallidos": len(detalles_fallos),
        "total_intentos": total,
        "detalles_fallos": detalles_fallos,
    }


# ---------------------------------------------------------------------------
# ENTRY POINT CLI
# ---------------------------------------------------------------------------

def main(argv=None) -> int:
    argv = argv if argv is not None else sys.argv[1:]
    print("=" * 70, flush=True)
    print("Subir Anexos a Activos Fijos SAP", flush=True)
    print("=" * 70, flush=True)

    if len(argv) < 2:
        print(
            "Uso: python src/subir_anexos.py SOCIEDAD ARCHIVO1 [ARCHIVO2 ...]\n"
            f"  Sociedades válidas: {', '.join(VALID_SOCIEDADES)}",
            file=sys.stderr,
        )
        return 2

    sociedad = argv[0]
    archivos = [Path(a) for a in argv[1:]]

    # Validar archivos existen
    for a in archivos:
        if not a.exists():
            print(f"ERROR: archivo no existe: {a}", file=sys.stderr)
            return 1

    try:
        validar_sociedad(sociedad)
    except ValueError as exc:
        print(f"ERROR de validación: {exc}", file=sys.stderr)
        return 1

    try:
        session = get_sap_session()
    except RuntimeError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    try:
        stats = subir_anexos(session, sociedad, archivos)
    except Exception as exc:
        print(f"\nERROR durante el flujo: {exc}", file=sys.stderr)
        return 1

    print(flush=True)
    print("=" * 70, flush=True)
    print(
        f"Resumen: {stats['exitosos']}/{stats['total_intentos']} OK, "
        f"{stats['fallidos']} fallos",
        flush=True,
    )
    if stats["detalles_fallos"]:
        print("\nFallos:", flush=True)
        for activo, sub, archivo, error in stats["detalles_fallos"][:10]:
            print(f"  {activo}-{sub} / {Path(archivo).name}: {error}",
                  flush=True)
        if len(stats["detalles_fallos"]) > 10:
            print(f"  ... y {len(stats['detalles_fallos']) - 10} más",
                  flush=True)
    print("=" * 70, flush=True)
    return 0 if stats["fallidos"] == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
