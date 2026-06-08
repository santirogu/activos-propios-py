"""extraer_activos_creados.py — Extracción del log SM35P para un usuario
SAP, vía SAP GUI Scripting.

Replica `resources/ScriptSM35P.vbs`. El Usuario SAP que el usuario final
ingresa en el formulario se usa como filtro en el campo CREATOR de
SM35P (Monitor de logs BDC) con un wildcard "*" prefijo, para listar
todas las sesiones BDC creadas por él. Se toma el primer log de la
tabla resultante, se abre su detalle, y se exporta vía la cadena de
toolbar grabada.

REQUISITOS DE EJECUCIÓN
=======================
Sistema operativo: Windows con SAP GUI for Windows abierto y sesión
iniciada. Mismos requisitos que `sap_upload.py` y `sox_report.py`.

USO
===
    python src/extraer_activos_creados.py USUARIO_SAP

    Ejemplo:
        python src/extraer_activos_creados.py 1017209574

También se invoca desde la GUI vía el botón "Ejecutar" de la vista
"Extraer Activos Creados" dentro de Activos Fijos.

LIMITACIONES ACTUALES
=====================
El recording NO fija path/filename de salida — SAP guarda el .xlsx en
su ruta default (típicamente la última usada en la sesión). Si se
necesita control de path, ajustar `exportar_log` añadiendo
DY_PATH/DY_FILENAME en el diálogo wnd[1] o wnd[2].
"""

from __future__ import annotations

import re
import sys
import time
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

PROJECT_ROOT = Path(__file__).resolve().parent.parent
SALIDA_DIR = PROJECT_ROOT / "salida"

# Nombre estándar del archivo final. Patrón:
#   ActivosCreados_{USUARIO}_{YYYYMMDD_HHMMSS}.xlsx
NOMBRE_PREFIJO = "ActivosCreados"
NOMBRE_EXTENSION = ".xlsx"

# Nombres de hojas tras el post-procesamiento.
LOGS_SHEET_NAME = "Logs"
ACTIVOS_FIJOS_SHEET_NAME = "Activos Fijos"

# Header de la columna que contiene los mensajes del log SM35P. Si el
# header exacto no se encuentra, se usa el índice de columna por default
# (COL_MENSAJE_LOG_DEFAULT, 1-based).
HEADER_MENSAJE_LOG = "Mensaje de log"
COL_MENSAJE_LOG_DEFAULT = 2  # columna B

# Regex que detecta menciones de activos fijos en el campo "Mensaje de log".
# Formato observado en el log SAP: "El act.fj. 8048124 0 se ha creado".
#   - "act.fj." literal (acepta también "act. fj." con espacio opcional)
#   - espacio + número grande (código del activo, grupo 1)
#   - espacio + número (subnúmero, grupo 2)
# `re.IGNORECASE` para tolerar variantes "Act.fj.", "ACT.FJ.", etc.
PATRON_ACTIVO_LOG = re.compile(
    r"act\.\s*fj\.\s+(\d+)\s+(\d+)",
    re.IGNORECASE,
)

# Headers de la hoja "Activos Fijos" (en orden).
ACTIVOS_FIJOS_HEADERS = ("Activos Fijos", "Subnúmero")

# ---------------------------------------------------------------------------
# CONFIGURACIÓN
# ---------------------------------------------------------------------------

# T-code de la transacción de logs BDC (Monitor de Sesiones BDC).
T_CODE_SM35P = "sm35p"

# Campo CREATOR (filtro por usuario, con wildcard prefijo "*").
CAMPO_CREATOR = "wnd[0]/usr/subSCR_INFO:RSBDC_PROTOCOL:0201/txtD0100-CREATOR"

# Primera celda de la tabla de protocolos (columna EDATE, fila 0).
# Hacer F2 (sendVKey 2) sobre esta celda abre el detalle del log.
CELDA_PRIMER_REGISTRO = (
    "wnd[0]/usr/tabsTAB_PROTOCOL/tabpALL_PROT/"
    "ssubSCR_CONTENT:RSBDC_PROTOCOL:0200/"
    "tblRSBDC_PROTOCOLTC_PROTOCOL/"
    "txtLIST_BDCLD-EDATE[0,0]"
)

# Botones de toolbar de la cadena de exportación. Los índices son
# específicos de la pantalla de detalle del log de SM35P (recording) y
# NO son estándar SAP.
BTN_EXPORTAR_TBAR0 = "wnd[0]/tbar[0]/btn[86]"
BTN_EXPORTAR_TBAR1 = "wnd[0]/tbar[1]/btn[43]"

# Campos del diálogo "Save list as file" (wnd[1]) que SAP abre tras la
# cadena de exportación. Los seteamos para forzar que el archivo caiga
# en `salida/` con el nombre estándar, en lugar de la ruta default que
# el recording usaba (la última carpeta usada por el usuario en SAP).
CAMPO_DY_PATH = "wnd[1]/usr/ctxtDY_PATH"
CAMPO_DY_FILENAME = "wnd[1]/usr/ctxtDY_FILENAME"
BTN_CONFIRMAR_WND1 = "wnd[1]/tbar[0]/btn[11]"


# ---------------------------------------------------------------------------
# LOGGING (mismo patrón que sap_upload.py / sox_report.py)
# ---------------------------------------------------------------------------

def _log(mensaje: str) -> None:
    ts = time.strftime("%H:%M:%S")
    print(f"[{ts}] {mensaje}", flush=True)


def _ejecutar(descripcion: str, fn, *args, **kwargs):
    """Ejecuta `fn(*args, **kwargs)` logueando la operación. Re-lanza
    con contexto si falla (las excepciones COM del SAP Frontend Server
    suelen venir con descripción vacía)."""
    _log(f"  → {descripcion}")
    try:
        return fn(*args, **kwargs)
    except Exception as exc:
        raise RuntimeError(
            f"Falló: {descripcion}\n"
            f"Detalle técnico SAP: {exc!r}"
        ) from exc


# ---------------------------------------------------------------------------
# VALIDACIONES
# ---------------------------------------------------------------------------

def _nombre_archivo_extraccion(usuario_sap: str) -> str:
    """Construye el nombre estándar del archivo extraído:
    `ActivosCreados_{USUARIO}_{YYYYMMDD_HHMMSS}.xlsx`."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{NOMBRE_PREFIJO}_{usuario_sap}_{ts}{NOMBRE_EXTENSION}"


def validar_usuario_sap(usuario: str) -> str:
    """Valida y normaliza el Usuario SAP ingresado por el usuario final.

    Acepta cualquier string no-vacío después de strip(). NO fuerza
    mayúsculas porque los IDs SAP varían (números puros como `1017209574`
    o alfanuméricos como `INTC37089` con casing específico). Sí elimina
    whitespace y rechaza inputs vacíos / no-string.

    Args:
        usuario: ID del usuario tal cual lo escribió el usuario final.

    Returns:
        El usuario normalizado (strip + sin transformación de casing).

    Raises:
        ValueError: si el input es None, no-string, o vacío tras strip.
    """
    if not isinstance(usuario, str) or not usuario.strip():
        raise ValueError("Debes ingresar un Usuario SAP.")
    return usuario.strip()


# ---------------------------------------------------------------------------
# CONEXIÓN A SAP (igual a sap_upload.get_sap_session)
# ---------------------------------------------------------------------------

def get_sap_session():
    """Conecta al SAP GUI Scripting Engine y devuelve la primera sesión."""
    try:
        import win32com.client  # type: ignore
    except ImportError as exc:
        raise RuntimeError(
            "Falta la dependencia pywin32. Instalar con: pip install pywin32"
        ) from exc

    try:
        sap_gui_auto = win32com.client.GetObject("SAPGUI")
    except Exception as exc:
        raise RuntimeError(
            "No se pudo conectar a SAP GUI. Verifica:\n"
            "  - SAP GUI for Windows está abierto y con sesión iniciada.\n"
            "  - SAP GUI Scripting habilitado en Options del cliente.\n"
            "  - sapgui/user_scripting = TRUE en el servidor SAP."
        ) from exc

    application = sap_gui_auto.GetScriptingEngine
    if application.Children.Count == 0:
        raise RuntimeError("No hay conexiones SAP activas en este SAP GUI.")
    connection = application.Children(0)
    if connection.Children.Count == 0:
        raise RuntimeError(
            "No hay sesiones activas en la conexión SAP. "
            "Inicia sesión en el sistema SAP antes de correr este script."
        )
    return connection.Children(0)


# ---------------------------------------------------------------------------
# PASOS DEL FLUJO SAP
# ---------------------------------------------------------------------------

def abrir_sm35p(session) -> None:
    """Maximiza la ventana y abre la T-code SM35P (monitor de logs BDC)."""
    _log(f"Paso 1/4: Abriendo transacción {T_CODE_SM35P.upper()}...")
    wnd = _ejecutar(
        "Localizar ventana principal wnd[0]",
        session.findById, "wnd[0]",
    )
    _ejecutar("Maximizar ventana principal", wnd.maximize)
    okcd = _ejecutar(
        "Localizar casilla de comandos (wnd[0]/tbar[0]/okcd)",
        session.findById, "wnd[0]/tbar[0]/okcd",
    )
    _ejecutar(
        f"Escribir T-code '{T_CODE_SM35P}' en okcd",
        lambda: setattr(okcd, "text", T_CODE_SM35P),
    )
    _ejecutar("Enviar Enter (sendVKey 0)", wnd.sendVKey, 0)


def filtrar_por_usuario(session, usuario_sap: str) -> None:
    """Filtra los logs por CREATOR con wildcard prefijo `*<usuario>` y
    aplica con Enter."""
    valor_filtro = f"*{usuario_sap}"
    _log(f"Paso 2/4: Filtrando por CREATOR='{valor_filtro}'...")
    creator = _ejecutar(
        f"Localizar campo CREATOR ({CAMPO_CREATOR})",
        session.findById, CAMPO_CREATOR,
    )
    _ejecutar(
        f"Asignar CREATOR = '{valor_filtro}'",
        lambda: setattr(creator, "text", valor_filtro),
    )
    _ejecutar("Foco en campo CREATOR", creator.setFocus)
    _ejecutar(
        f"Cursor al final (caretPosition={len(valor_filtro)})",
        lambda: setattr(creator, "caretPosition", len(valor_filtro)),
    )
    wnd = _ejecutar(
        "Localizar wnd[0] para enviar Enter",
        session.findById, "wnd[0]",
    )
    _ejecutar("Pulsar Enter para aplicar filtro", wnd.sendVKey, 0)


def abrir_primer_registro(session) -> None:
    """Abre el detalle del primer log (fila 0) de la tabla de protocolos
    mediante foco en la celda EDATE[0,0] + F2 (sendVKey 2)."""
    _log("Paso 3/4: Abriendo detalle del primer registro...")
    celda = _ejecutar(
        f"Localizar primer registro de la tabla ({CELDA_PRIMER_REGISTRO})",
        session.findById, CELDA_PRIMER_REGISTRO,
    )
    _ejecutar("Foco en primer registro", celda.setFocus)
    _ejecutar(
        "Cursor en celda (caretPosition=5)",
        lambda: setattr(celda, "caretPosition", 5),
    )
    wnd = _ejecutar(
        "Localizar wnd[0] para F2",
        session.findById, "wnd[0]",
    )
    _ejecutar("Pulsar F2 (sendVKey 2) para abrir detalle", wnd.sendVKey, 2)


def exportar_log(session, carpeta_destino: str, nombre_archivo: str) -> None:
    """Exporta el log abierto al archivo `carpeta_destino/nombre_archivo`.

    Variante optimizada del recording: en vez de hacer F4 + picker
    (wnd[2]) + confirmar 2 veces, inyectamos `DY_PATH` y `DY_FILENAME`
    DIRECTAMENTE en el diálogo wnd[1] y confirmamos una sola vez con
    btn[11]. El picker del recording era una conveniencia del usuario
    para navegar; programáticamente no lo necesitamos.

    Secuencia:
      1. `wnd[0]/tbar[0]/btn[86]` press — abre menú/acción de exportar.
      2. `wnd[0]/tbar[1]/btn[43]` press — sub-acción que abre el save dialog.
      3. Set `wnd[1]/usr/ctxtDY_PATH` con `carpeta_destino`.
      4. Set `wnd[1]/usr/ctxtDY_FILENAME` con `nombre_archivo`.
      5. `wnd[1]/tbar[0]/btn[11]` press — confirma save.

    Args:
        session: sesión SAP GUI.
        carpeta_destino: ruta absoluta (ej. `C:\\Users\\xxx\\salida`).
        nombre_archivo: nombre con extensión (ej. `ActivosCreados_USR_20260601_143022.xlsx`).
    """
    _log(f"Paso 4/4: Exportando log → {carpeta_destino}\\{nombre_archivo}")
    btn1 = _ejecutar(
        f"Localizar botón exportar tbar[0] ({BTN_EXPORTAR_TBAR0})",
        session.findById, BTN_EXPORTAR_TBAR0,
    )
    _ejecutar("Pulsar btn[86] de tbar[0]", btn1.press)

    btn2 = _ejecutar(
        f"Localizar botón exportar tbar[1] ({BTN_EXPORTAR_TBAR1})",
        session.findById, BTN_EXPORTAR_TBAR1,
    )
    _ejecutar("Pulsar btn[43] de tbar[1]", btn2.press)

    # Inyectar DY_PATH y DY_FILENAME en el diálogo wnd[1] (salta F4/picker).
    path_field = _ejecutar(
        f"Localizar campo ruta ({CAMPO_DY_PATH})",
        session.findById, CAMPO_DY_PATH,
    )
    _ejecutar(
        f"Asignar ruta = '{carpeta_destino}'",
        lambda: setattr(path_field, "text", carpeta_destino),
    )

    nombre_field = _ejecutar(
        f"Localizar campo nombre ({CAMPO_DY_FILENAME})",
        session.findById, CAMPO_DY_FILENAME,
    )
    _ejecutar(
        f"Asignar nombre = '{nombre_archivo}'",
        lambda: setattr(nombre_field, "text", nombre_archivo),
    )
    _ejecutar(
        "Cursor al final del nombre",
        lambda: setattr(nombre_field, "caretPosition", len(nombre_archivo)),
    )

    btn_save = _ejecutar(
        f"Localizar botón confirmar save ({BTN_CONFIRMAR_WND1})",
        session.findById, BTN_CONFIRMAR_WND1,
    )
    _ejecutar("Pulsar OK para guardar", btn_save.press)


# ---------------------------------------------------------------------------
# ORQUESTADOR
# ---------------------------------------------------------------------------

def _esperar_archivo_listo(
    archivo: Path, timeout_seg: float = 10.0, poll_seg: float = 0.5
) -> bool:
    """Espera a que `archivo` exista y su tamaño se estabilice (no esté
    siendo escrito todavía). SAP a veces tarda 1-3s en cerrar el handle
    del archivo tras la exportación. Returns True si quedó listo dentro
    del timeout, False en caso contrario.

    Helper duplicado de `sox_report._esperar_archivo_listo` para
    mantener cada módulo SAP self-contained — la función es trivial y
    el costo de duplicación menor.
    """
    inicio = time.time()
    tamano_previo = -1
    while time.time() - inicio < timeout_seg:
        if archivo.exists():
            tamano = archivo.stat().st_size
            if tamano > 0 and tamano == tamano_previo:
                return True
            tamano_previo = tamano
        time.sleep(poll_seg)
    return False


def procesar_logs(archivo_path: Path) -> dict[str, int]:
    """Post-procesa el .xlsx exportado por SAP:

      1. Renombra la hoja única exportada por SAP a `"Logs"`.
      2. Parsea la columna `"Mensaje de log"` buscando `PATRON_ACTIVO_LOG`
         y extrae todos los pares `(activo_fijo, subnúmero)`.
      3. Deduplica los pares preservando orden de primera aparición.
      4. Crea (o reemplaza) una hoja `"Activos Fijos"` con 2 columnas:
         `Activos Fijos` + `Subnúmero`, headers en bold, datos como ints.

    Si la hoja `Activos Fijos` ya existe (corrida previa), se borra y
    recrea — idempotente.

    Args:
        archivo_path: ruta al .xlsx generado por SAP.

    Returns:
        Dict con stats:
          - `menciones_total`: cuántas veces el regex matcheó (incluye
            duplicados).
          - `activos_unicos`: cuántos pares (activo, sub) únicos hay.
          - `descartadas`: filas donde el regex no matcheó pero el
            mensaje sí contenía "act" (no se incrementa por filas no
            relacionadas).

    Raises:
        FileNotFoundError: si `archivo_path` no existe.
    """
    if not archivo_path.exists():
        raise FileNotFoundError(
            f"No existe el archivo a procesar: {archivo_path}"
        )

    _log(f"Post-procesamiento: procesando '{archivo_path.name}'...")
    _esperar_archivo_listo(archivo_path)

    wb = load_workbook(archivo_path)

    # Paso 1: renombrar la hoja única a "Logs". Si ya hay una hoja
    # llamada "Logs", se respeta (probablemente segundo run sobre el
    # mismo archivo).
    ws_logs = wb.active
    if ws_logs.title != LOGS_SHEET_NAME:
        # Si ya existe una hoja con ese nombre (raro pero posible), la
        # borramos para evitar el conflicto del rename.
        if LOGS_SHEET_NAME in wb.sheetnames:
            del wb[LOGS_SHEET_NAME]
        ws_logs.title = LOGS_SHEET_NAME

    # Paso 2: localizar la columna de mensaje del log por header. Si no
    # se encuentra, fallback a columna B (COL_MENSAJE_LOG_DEFAULT).
    header_row = next(
        ws_logs.iter_rows(min_row=1, max_row=1, values_only=True), tuple()
    )
    col_mensaje_idx = COL_MENSAJE_LOG_DEFAULT
    for idx, value in enumerate(header_row, start=1):
        if value == HEADER_MENSAJE_LOG:
            col_mensaje_idx = idx
            break

    # Paso 3: escanear filas de datos (a partir de fila 2), aplicar
    # regex, acumular y deduplicar manteniendo orden.
    menciones_total = 0
    vistos: set[tuple[int, int]] = set()
    activos_unicos: list[tuple[int, int]] = []

    for row in ws_logs.iter_rows(min_row=2, values_only=True):
        if col_mensaje_idx > len(row):
            continue
        mensaje = row[col_mensaje_idx - 1]
        if not isinstance(mensaje, str):
            continue
        for match in PATRON_ACTIVO_LOG.finditer(mensaje):
            menciones_total += 1
            par = (int(match.group(1)), int(match.group(2)))
            if par not in vistos:
                vistos.add(par)
                activos_unicos.append(par)

    # Paso 4: crear hoja "Activos Fijos" (idempotente).
    if ACTIVOS_FIJOS_SHEET_NAME in wb.sheetnames:
        del wb[ACTIVOS_FIJOS_SHEET_NAME]
    ws_af = wb.create_sheet(ACTIVOS_FIJOS_SHEET_NAME)

    bold = Font(bold=True)
    for col_idx, header in enumerate(ACTIVOS_FIJOS_HEADERS, start=1):
        cell = ws_af.cell(1, col_idx, header)
        cell.font = bold

    for fila_idx, (activo, sub) in enumerate(activos_unicos, start=2):
        ws_af.cell(fila_idx, 1, activo)
        ws_af.cell(fila_idx, 2, sub)

    wb.save(archivo_path)

    _log(
        f"Post-procesamiento OK: {menciones_total} menciones, "
        f"{len(activos_unicos)} activos únicos escritos en "
        f"'{ACTIVOS_FIJOS_SHEET_NAME}'."
    )
    return {
        "menciones_total": menciones_total,
        "activos_unicos": len(activos_unicos),
    }


def extraer_activos_creados(
    session,
    usuario_sap: str,
    carpeta_destino: str | None = None,
    nombre_archivo: str | None = None,
) -> tuple[str, str]:
    """Ejecuta el flujo completo de SM35P para extraer los activos creados
    por el `usuario_sap` indicado, guardando el resultado en
    `carpeta_destino/nombre_archivo`.

    Args:
        session: sesión SAP GUI activa.
        usuario_sap: ID del usuario SAP SIN wildcard — el flujo añade el
            `*` prefijo automáticamente al aplicar el filtro CREATOR.
        carpeta_destino: ruta absoluta donde guardar el .xlsx
            (default: `<PROJECT_ROOT>/salida`).
        nombre_archivo: nombre del archivo de salida
            (default: `ActivosCreados_{USUARIO}_{YYYYMMDD_HHMMSS}.xlsx`).

    Returns:
        (carpeta, nombre): ruta absoluta de la carpeta destino y nombre
        del archivo creado por SAP. El handler GUI los usa para mostrar
        al usuario dónde quedó el archivo.

    Raises:
        ValueError: si `usuario_sap` no pasa `validar_usuario_sap`.
        RuntimeError: si alguna etapa SAP falla (re-lanzada con contexto).
    """
    usuario_norm = validar_usuario_sap(usuario_sap)

    if carpeta_destino is None:
        SALIDA_DIR.mkdir(parents=True, exist_ok=True)
        carpeta_destino = str(SALIDA_DIR)
    if nombre_archivo is None:
        nombre_archivo = _nombre_archivo_extraccion(usuario_norm)

    inicio = time.monotonic()
    _log("=== Iniciando extracción de activos creados (SM35P) ===")
    abrir_sm35p(session)
    filtrar_por_usuario(session, usuario_norm)
    abrir_primer_registro(session)
    exportar_log(session, carpeta_destino, nombre_archivo)

    # Etapa post-SAP (pure Python): renombrar hoja a "Logs" + crear
    # hoja "Activos Fijos" con los pares (activo, subnúmero) parseados
    # de los mensajes del log.
    archivo_path = Path(carpeta_destino) / nombre_archivo
    procesar_logs(archivo_path)

    duracion = time.monotonic() - inicio
    _log(f"=== Extracción finalizada en {duracion:.1f}s ===")

    return carpeta_destino, nombre_archivo


# ---------------------------------------------------------------------------
# ENTRY POINT CLI
# ---------------------------------------------------------------------------

def main(argv=None) -> int:
    argv = argv if argv is not None else sys.argv[1:]
    print("=" * 70, flush=True)
    print("Extracción de log SM35P por Usuario SAP", flush=True)
    print("=" * 70, flush=True)

    if len(argv) != 1:
        print(
            "Uso: python src/extraer_activos_creados.py USUARIO_SAP\n"
            "Ejemplo: python src/extraer_activos_creados.py 1017209574",
            file=sys.stderr,
        )
        return 2

    usuario = argv[0]
    try:
        validar_usuario_sap(usuario)
    except ValueError as exc:
        print(f"ERROR de validación: {exc}", file=sys.stderr, flush=True)
        return 1

    try:
        session = get_sap_session()
    except RuntimeError as exc:
        print(f"ERROR: {exc}", file=sys.stderr, flush=True)
        return 1

    try:
        carpeta, nombre = extraer_activos_creados(session, usuario)
    except Exception as exc:
        print(f"\nERROR durante el flujo SM35P: {exc}", file=sys.stderr, flush=True)
        return 1

    print(flush=True)
    print("=" * 70, flush=True)
    print(f"Log extraído: {carpeta}\\{nombre}", flush=True)
    print("=" * 70, flush=True)
    return 0


if __name__ == "__main__":
    sys.exit(main())
